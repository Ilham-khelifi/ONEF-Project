import sys
import os

from Models import init_db

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QPushButton, QTableWidget,
                             QTableWidgetItem, QVBoxLayout, QHBoxLayout, QLineEdit, QLabel,
                             QFrame, QDialog, QMessageBox, QCheckBox, QComboBox, QDateEdit,
                             QScrollArea, QSizePolicy, QStackedWidget, QToolButton, QMenu,
                             QListWidget, QListWidgetItem, QGridLayout, QSpinBox, QHeaderView, QAbstractItemView,QFileDialog)
from PyQt5.QtGui import QIcon, QPixmap, QFont, QColor, QPalette, QBrush
from PyQt5.QtCore import Qt, QSize, QDate, QTimer, QLocale, pyqtSignal

from Absence import AddForm
from Controllers.Evaluation import EvaluationController
from Controllers.EmployeController import EmployeeController   
from DatabaseConnection import db
from MessageBox import StyledMessageDialog
from TablePaginator1 import tablepaginator
from Views.Historique import HistoryDialog
from topbar import create_top_bar
import openpyxl
from weasyprint import HTML, CSS
import html

# Define color constants based on the screenshots
DARK_BG = "#263238"      # Dark background color
DARKER_BG = "#26282b"    # Darker background for contrast
MEDIUM_BG = "#37474f"    # Medium background for containers
LIGHT_BG = "#455a64"     # Light background for hover states
ORANGE = "#ff6a0e"       # Primary accent color
YELLOW = "#ffc20e"       # Secondary accent color
YELLOW_BTN = "#e6b800"   # Button color
WHITE = "#ffffff"        # Text color
BLACK = "#000000"        # Black text color
GRAY = "#8f8f8f"         # Gray for secondary text
GREEN = "#4CAF50"        # Success color
RED = "#f44336"          # Error/Delete color
FORM_BG = "#1e2327"      # Form background color from screenshot
TABLE_ROW_BG = "#2c2c2c" # Table row background from screenshot
DELETE_DIALOG_BG = "#1e2327"  # Delete dialog background
DELETE_DIALOG_BOX = "#37474f"  # Delete dialog message box
DELETE_BTN_GREEN = "#3f861e"   # Delete button green
DELETE_BTN_RED = "#f50707"     # Cancel button red




class Evaluation(QMainWindow):
    def __init__(self, current_user_data=None, session=None):
        """
        Initialize the main application window and set up the UI components
        """
        super().__init__()
        self.setWindowTitle("نظام إدارة الموارد البشرية - إدارة التنقيط")
        self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE}; font-size: 14px;")
        self.current_user_data = current_user_data or {}
        current_user_account_number = None
        if self.current_user_data:
            current_user_account_number = self.current_user_data.get('account_number')
            print(f"DEBUG - Utilisateur actuel dans EvaluationManagement: {current_user_account_number}")
        else:
            print("DEBUG - Aucun utilisateur actuel défini dans EvaluationManagement")
        # MODIFICATION: Utiliser la session partagée ou créer une nouvelle si nécessaire
        if session:
            self.session = session
            print("DEBUG - Utilisation de la session partagée dans gestionComptes")
        else:
            print("DEBUG - Création d'une nouvelle session dans gestionComptes")
            self.session = init_db('mysql+pymysql://hr:hr@localhost/HR')
        self.controller = EvaluationController(self.session, current_user_account_number)
        
        # Central widget
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        # Main horizontal layout
        self.main_layout = QHBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)

        # Sidebar and toggle
        #self.sidebar, self.sidebar_toggle = create_sidebar(self, self.main_layout)
        #self.main_layout.addWidget(self.sidebar)

        # Scrollable content area
        self.content_scroll = QScrollArea()
        self.content_scroll.setWidgetResizable(True)
        self.content_scroll.setFrameShape(QFrame.NoFrame)
        self.content_scroll.setStyleSheet(f"background-color: {DARK_BG};")

        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(0)

        self.content_scroll.setWidget(self.content_widget)
        self.main_layout.addWidget(self.content_scroll)

        # Top bar
        #self.top_bar = create_top_bar(self, self.content_layout, self.sidebar_toggle)
        #self.content_layout.addWidget(self.top_bar)

        # Stacked widget (for switching views)
        self.stacked_widget = QStackedWidget()
        self.content_layout.addWidget(self.stacked_widget)

        # Main page setup
        self.main_page = QWidget()
        self.main_page_layout = QVBoxLayout(self.main_page)
        self.main_page_layout.setContentsMargins(15, 15, 15, 15)

        # Title container for label + refresh button side by side
        title_container = QWidget()
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 10)  # margin below the title row
        title_layout.setSpacing(10)

        # Table title label
        table_title = QLabel("تفاصيل جدول التنقيط  ")
        table_title.setStyleSheet("font-size: 18px; font-weight: bold; color: white;")
        table_title.setAlignment(Qt.AlignLeft)

        # Refresh button
        refresh_btn = QPushButton("⟳")  # Unicode refresh icon
        refresh_btn.setFixedSize(40, 40)
        refresh_btn.setToolTip("تحديث الجدول")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 17px;
                color: #ffffff;
                font-weight: bold;
                font-size: 20px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        refresh_btn.clicked.connect(self.reload_table)

        # Add title and button with stretch in between to push button right
        title_layout.addWidget(table_title)
        title_layout.addStretch()
        title_layout.addWidget(refresh_btn)

        # Add the title container (label + button) to the main page layout
        self.main_page_layout.addWidget(title_container)

        # Action bar and table
        self.create_table()
        self.action_bar = ActionBar(self, self.table)
        self.main_page_layout.insertWidget(0, self.action_bar)

        # Buttons
        self.create_action_buttons()

        # Add main page to stacked widget
        self.stacked_widget.addWidget(self.main_page)
        self.stacked_widget.setCurrentIndex(0)


    def create_table_title(self):
        """
        Create a title for the main table
        """
        title_container = QWidget()
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 10)

        # Create title label
        title_label = QLabel("تفاصيل جدول التنقيط")
        title_label.setStyleSheet(f"""
            color: {WHITE};
            font-size: 18px;
            font-weight: bold;
        """)
        title_label.setLayoutDirection(Qt.RightToLeft)
        title_layout.addWidget(title_label)

        self.main_page_layout.addWidget(title_container)

    def create_table(self):
        table_container = QWidget()
        table_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(15, 15, 15, 15)

        self.table = QTableWidget()
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                border: none;
                gridline-color: {LIGHT_BG};
                font-size: 16px;
            }}
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid {LIGHT_BG};
            }}
            QTableWidget::item:selected {{
                background-color: {ORANGE};
                color: {WHITE};
            }}
            QHeaderView::section {{
                background-color: {DARKER_BG};
                color: {WHITE};
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 16px;
            }}
            QScrollBar:vertical {{
                background: {MEDIUM_BG};
                width: 0px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {LIGHT_BG};
                min-height: 20px;
                border-radius: 6px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)

        columns = [
            "رقم التنقيط", "السنة", "رقم الموظف","اللقب", "الإسم",
            "النقطة السنوية", "نقطة\nالمردودية 1", "نقطة\nالمردودية 2",
            "نقطة\nالمردودية 3", "نقطة\nالمردودية 4", "المعدل"
        ]
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionsMovable(False)
        header.setSectionsClickable(False)

        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)

        for i in range(len(columns)):
            self.table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

        # === Load data from database using the controller ===
        table_data = self.controller.load_evaluations_with_employee_names_current_year()
        print("Sample row:", table_data[0] if table_data else "No data")

        self.table.setRowCount(len(table_data) if table_data else 0)

        for row_idx, row_data in enumerate(table_data):
            def safe_text(value):
                return str(value) if value is not None else ""

            self.table.setItem(row_idx, 0, QTableWidgetItem(safe_text(row_data["idEvaluation"])))
            self.table.setItem(row_idx, 1, QTableWidgetItem(safe_text(row_data["Annee"])))
            self.table.setItem(row_idx, 2, QTableWidgetItem(safe_text(row_data["idemploye"])))
            self.table.setItem(row_idx, 3, QTableWidgetItem(safe_text(row_data["nom"])))
            self.table.setItem(row_idx, 4, QTableWidgetItem(safe_text(row_data["prenom"])))
            self.table.setItem(row_idx, 5, QTableWidgetItem(safe_text(row_data["NoteAnnuelle"])))
            self.table.setItem(row_idx, 6, QTableWidgetItem(safe_text(row_data["Note1"])))
            self.table.setItem(row_idx, 7, QTableWidgetItem(safe_text(row_data["Note2"])))
            self.table.setItem(row_idx, 8, QTableWidgetItem(safe_text(row_data["Note3"])))
            self.table.setItem(row_idx, 9, QTableWidgetItem(safe_text(row_data["Note4"])))
            self.table.setItem(row_idx, 10, QTableWidgetItem(safe_text(row_data["Moyenne"])))

            for col_idx in range(len(columns)):
                self.table.item(row_idx, col_idx).setTextAlignment(Qt.AlignCenter)

        # Add table to layout
        table_layout.addWidget(self.table)

        # Integrate pagination with loaded data
        paginator = tablepaginator(self.table, rows_per_page=10)

        table_layout.addWidget(paginator)

        # Add container to main layout
        self.main_page_layout.addWidget(table_container)

    def reload_table(self):
        # Clear existing rows
        self.table.setRowCount(0)

        # Load data from the database
        table_data = self.controller.load_evaluations_with_employee_names_current_year()
        print("Reloaded data:", table_data[0] if table_data else "No data")

        self.table.setRowCount(len(table_data) if table_data else 0)

        for row_idx, row_data in enumerate(table_data):
            def safe_text(value):
                return str(value) if value is not None else ""

            self.table.setItem(row_idx, 0, QTableWidgetItem(safe_text(row_data["idEvaluation"])))
            self.table.setItem(row_idx, 1, QTableWidgetItem(safe_text(row_data["Annee"])))
            self.table.setItem(row_idx, 2, QTableWidgetItem(safe_text(row_data["idemploye"])))
            self.table.setItem(row_idx, 3, QTableWidgetItem(safe_text(row_data["nom"])))
            self.table.setItem(row_idx, 4, QTableWidgetItem(safe_text(row_data["prenom"])))
            self.table.setItem(row_idx, 5, QTableWidgetItem(safe_text(row_data["NoteAnnuelle"])))
            self.table.setItem(row_idx, 6, QTableWidgetItem(safe_text(row_data["Note1"])))
            self.table.setItem(row_idx, 7, QTableWidgetItem(safe_text(row_data["Note2"])))
            self.table.setItem(row_idx, 8, QTableWidgetItem(safe_text(row_data["Note3"])))
            self.table.setItem(row_idx, 9, QTableWidgetItem(safe_text(row_data["Note4"])))
            self.table.setItem(row_idx, 10, QTableWidgetItem(safe_text(row_data["Moyenne"])))

            for col_idx in range(self.table.columnCount()):
                self.table.item(row_idx, col_idx).setTextAlignment(Qt.AlignCenter)
    def update_data(self):
        self.reload_table()
    def create_action_buttons(self):
            """
            Create action buttons for view details, previous years, and activity log
            All buttons have the same size
            """
            # Buttons container
            buttons_widget = QWidget()
            buttons_layout = QHBoxLayout(buttons_widget)
            buttons_layout.setContentsMargins(0, 20, 0, 0)  # Increased top margin
            buttons_layout.setSpacing(25)  # Increased spacing

            # Create buttons - REMOVED DELETE BUTTON as requested
            edit_btn = QPushButton("تعديل")
            add_btn = QPushButton("إضافة")
            previous_years_btn = QPushButton("عرض تفاصيل \nالسنوات السابقة")
            history_btn = QPushButton("سجل الأنشطة")

            # Style buttons - all with the same fixed size
            for btn in [edit_btn, add_btn , previous_years_btn, history_btn]:
                btn.setFixedSize(140, 40)  # Same fixed size for all buttons
                btn.setStyleSheet(f"""
                       QPushButton {{
                           background-color: {ORANGE};
                           color: {WHITE};
                           border: none;
                           border-radius: 5px;
                           font-size: 16px;
                           font-weight: bold;
                       }}
                       QPushButton:hover {{
                           background-color: #e05d00;
                       }}
                       QPushButton:pressed {{
                           background-color: #cc5200;
                       }}
                       QPushButton:disabled {{
                           background-color: #a0a0a0;
                           color: #e0e0e0;
                       }}
                   """)

            # Connect buttons to actions
            edit_btn.clicked.connect(self.on_edit_button_clicked)
            add_btn.clicked.connect(self.create_add_window )
            previous_years_btn.clicked.connect(self.show_previous_years)
            history_btn.clicked.connect(self.show_history)

            # Add buttons to layout - right to left for Arabic
            buttons_layout.addWidget(previous_years_btn)
            buttons_layout.addWidget(edit_btn)
            buttons_layout.addWidget(add_btn)

            buttons_layout.addWidget(history_btn)

            self.main_page_layout.addWidget(buttons_widget)

    def show_previous_years(self):
        self.previous_years = PreviousYears(self, session=self.session)
        self.previous_years.closed.connect(self.reload_table)  # Now reliably connected
        self.previous_years.show()

    def show_history(self):
        """Show the activity history using the  history dialog with shared session"""
        try:
            from Views.Historique import HistoryDialog
            
            print(f"DEBUG - Ouverture historique avec user_data: {self.current_user_data}")
            
            self.history_dialog = HistoryDialog(
                parent=self,
                current_user_data=self.current_user_data,
                session=self.session,  # AJOUT: Passer la session partagée
                module_name="إدارة التقيمات",
                gestion_filter="إدارة التقييمات" # Filtre pour le module actuel
            )
            self.history_dialog.show()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "خطأ",
                f"حدث خطأ غير متوقع:\n{str(e)}"
            )
            print(f"Erreur dans show_history: {e}")   


    def on_edit_button_clicked(self):
        indexes = self.table.selectionModel().selectedRows()

        if len(indexes) == 0:
            dialog = StyledMessageDialog(self, title="تحذير", message="يرجى تحديد صف للتعديل.", message_type="warning")
            dialog.exec_()
            return

        selected_row = indexes[0].row()

        def get_cell_text(row, col):
            item = self.table.item(row, col)
            if item:
                return item.text()
            widget = self.table.cellWidget(row, col)
            if widget and hasattr(widget, 'text'):
                return widget.text()
            return ""

        data = {
            "رقم التنقيط": get_cell_text(selected_row, 0),
            "السنة": get_cell_text(selected_row, 1),
            "رقم الموظف": get_cell_text(selected_row, 2),
            "الإسم": get_cell_text(selected_row, 3),
            "اللقب": get_cell_text(selected_row, 4),
            "النقطة السنوية": get_cell_text(selected_row, 5),
            "نقطة المردودية 1": get_cell_text(selected_row, 6),
            "نقطة المردودية 2": get_cell_text(selected_row, 7),
            "نقطة المردودية 3": get_cell_text(selected_row, 8),
            "نقطة المردودية 4": get_cell_text(selected_row, 9),
        }

        self.edit_window = EditForm(self, data, selected_row, self.table, session=self.session)
        if self.edit_window.exec_() == QDialog.Accepted:
            self.reload_table()

    def create_add_window(self):
        indexes = self.table.selectionModel().selectedRows()

        if len(indexes) == 0:
            dialog = StyledMessageDialog(self, title="تحذير", message="يرجى تحديد موظف من الجدول", message_type="warning")
            dialog.exec_()
            return

        selected_row = indexes[0].row()

        def get_cell_text(row, col):
            item = self.table.item(row, col)
            return item.text() if item else ""

        # Collect the data from the selected row based on your columns
        data = {
            "رقم التنقيط": get_cell_text(selected_row, 0),
            "السنة": get_cell_text(selected_row, 1),
            "رقم الموظف": get_cell_text(selected_row, 2),
            "إسم الموظف": get_cell_text(selected_row, 3),
            "لقب الموظف": get_cell_text(selected_row, 4),
            "النقطة السنوية": get_cell_text(selected_row, 5),
            "نقطة المردودية 1": get_cell_text(selected_row, 6),
            "نقطة المردودية 2": get_cell_text(selected_row, 7),
            "نقطة المردودية 3": get_cell_text(selected_row, 8),
            "نقطة المردودية 4": get_cell_text(selected_row, 9),
        }

        # Pass the data dict to the AddForm constructor
        self.add_form = AddForm(self,session=self.session)

        # Populate the form fields from data
        for key, value in data.items():
            if key in self.add_form.add_form_fields:
                self.add_form.add_form_fields[key].setText(value)

        if self.add_form.exec_() == QDialog.Accepted:
            self.reload_table()

    # --- MÉTHODES D'EXPORTATION ET D'IMPRESSION POUR LA CLASSE Evaluation (Année Courante) ---
    def _get_table_data_as_lists(self):
        """Helper function to extract data ONLY from VISIBLE rows in self.table (année courante)."""
        if not hasattr(self, 'table') or not self.table or self.table.rowCount() == 0:
            print(f"Avertissement dans _get_table_data_as_lists pour {self.__class__.__name__}: self.table non disponible ou vide.")
            return None, None

        headers = [self.table.horizontalHeaderItem(i).text()
                   for i in range(self.table.columnCount())]

        visible_row_data = []
        for row in range(self.table.rowCount()):
            # Vérifie si la ligne est cachée (par le paginateur associé à self.table)
            if not self.table.isRowHidden(row): 
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                visible_row_data.append(row_data)
        return headers, visible_row_data

    def export_data_to_excel(self):
        print(f"{self.__class__.__name__}: export_data_to_excel (année courante) called")
        headers, data = self._get_table_data_as_lists()

        if not data:
            QMessageBox.information(self, "لا بيانات للتصدير", "لا توجد صفوف ظاهرة في جدول التنقيط الحالي للتصدير.")
            return

        filePath, _ = QFileDialog.getSaveFileName(self, "تصدير إلى Excel (التنقيط الحالي)", 
                                                  os.path.expanduser("~/Documents/Evaluations_Annee_Courante.xlsx"),
                                                  "Excel Workbook (*.xlsx);;All Files (*)")
        if not filePath: return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "تنقيط السنة الحالية"
            sheet.sheet_view.rightToLeft = True
            sheet.append(headers)
            for row_values in data: sheet.append(row_values)
            for col_idx, column_cells in enumerate(sheet.columns):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = length + 5
            workbook.save(filePath)
            QMessageBox.information(self, "نجاح التصدير", f"تم تصدير البيانات بنجاح إلى:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ أثناء تصدير الملف: {e}")
            print(f"Error exporting Evaluations (current year) to Excel: {e}")

    def print_data_to_pdf(self):
        print(f"{self.__class__.__name__}: print_data_to_pdf (année courante) called")
        headers, data = self._get_table_data_as_lists()

        if not data:
            QMessageBox.information(self, "لا بيانات للطباعة", "لا توجد صفوف ظاهرة في جدول التنقيط الحالي للطباعة.")
            return

        filePath, _ = QFileDialog.getSaveFileName(self, "طباعة إلى PDF (التنقيط الحالي)", 
                                                  os.path.expanduser("~/Documents/Evaluations_Annee_Courante.pdf"),
                                                  "PDF Document (*.pdf);;All Files (*)")
        if not filePath: return

        html_content = "<html><head><meta charset='UTF-8'>"
        html_content += """
        <style>
            @font-face { font-family: 'DejaVu Sans'; }
            body { direction: rtl; font-family: 'Arial', 'DejaVu Sans', sans-serif; font-size: 9pt; }
            table { width: 100%; border-collapse: collapse; margin-top: 10px; }
            th, td { border: 1px solid #333; padding: 4px; text-align: right; word-wrap: break-word; }
            th { background-color: #f0f0f0; font-weight: bold; }
            caption { font-size: 1.1em; font-weight: bold; margin-bottom: 8px; text-align: center; }
        </style>
        </head><body>
        """
        html_content += "<table><caption>تفاصيل جدول التنقيط (السنة الحالية)</caption><thead><tr>"
        for header in headers: html_content += f"<th>{header}</th>"
        html_content += "</tr></thead><tbody>"
        for row_values in data:
            html_content += "<tr>"
            for cell_value in row_values: html_content += f"<td>{html.escape(str(cell_value))}</td>"
            html_content += "</tr>"
        html_content += "</tbody></table></body></html>"

        try:
            # A3 landscape pour de nombreuses colonnes
            css_style = CSS(string='@page { size: A3 landscape; margin: 0.7cm; }') 
            HTML(string=html_content).write_pdf(filePath, stylesheets=[css_style])
            QMessageBox.information(self, "نجاح الطباعة", f"تم إنشاء ملف PDF بنجاح:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في الطباعة", f"حدث خطأ أثناء إنشاء ملف PDF: {e}\n"
                                 "Vérifiez WeasyPrint et ses dépendances (Pango, Cairo).")
            print(f"Error printing Evaluations (current year) to PDF: {e}")






class ActionBar(QWidget):
    def __init__(self, parent, table):
        super().__init__(parent)
        self.table = table
        self.setFixedHeight(60)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 0, 0, 0)
        layout.setSpacing(10)

        # Filter section container
        filter_container = QWidget()
        filter_container_layout = QHBoxLayout(filter_container)
        filter_container_layout.setContentsMargins(0, 0, 0, 0)
        filter_container_layout.setSpacing(10)


        # Filter button with icon and text
        self.filter_btn = QPushButton("ترشيح")
        self.filter_btn.setIcon(QIcon("pics/filter.png"))  # Replace with a valid icon path
        self.filter_btn.setIconSize(QSize(20, 20))
        self.filter_btn.setFixedHeight(35)
        self.filter_btn.setToolTip("ترشيح")
        self.filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 5px;
                color: #ffffff;
                font-weight: bold;
                font-size: 16px;
                padding: 0 10px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        self.filter_btn.clicked.connect(self.show_filter_dialog)


        filter_container_layout.addWidget(self.filter_btn)

        layout.addWidget(filter_container )
        layout.addStretch()
        self.setLayout(layout)

        self.filter_checkboxes = {}  # Store column checkboxes


    def show_filter_dialog(self):
        """
        Show the filter dialog with checkboxes for column selection
        """
        self.filter_dialog = QDialog(self)
        self.filter_dialog.setWindowTitle("ترشيح")
        self.filter_dialog.setFixedWidth(350)
        self.filter_dialog.setStyleSheet("background-color: #37474f; color: #ffffff;")

        dialog_layout = QVBoxLayout(self.filter_dialog)
        dialog_layout.setContentsMargins(20, 20, 20, 20)

        title_label = QLabel("اختر الأعمدة للترشيح:")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 15px;")
        dialog_layout.addWidget(title_label)

        columns = [
             "رقم التنقيط", "السنة", "رقم الموظف", "الإسم", "اللقب",
            "النقطة السنوية", "نقطة المردودية 1", "نقطة المردودية 2",
            "نقطة المردودية 3", "نقطة المردودية 4", "المعدل"
        ]

        for column in columns:
            checkbox = QCheckBox(column)
            checkbox.setStyleSheet("""
                QCheckBox {
                    font-size: 16px;
                    padding: 8px 0;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
                QCheckBox::indicator:checked {
                    background-color: #ff6a0e;
                    border: 2px solid #ffffff;
                }
            """)
            dialog_layout.addWidget(checkbox)
            self.filter_checkboxes[column] = checkbox

        filter_value_layout = QHBoxLayout()
        filter_value_layout.setSpacing(10)

        filter_value_label = QLabel("قيمة الترشيح:")
        filter_value_label.setStyleSheet("font-size: 16px;")

        self.filter_value_input = QLineEdit()
        self.filter_value_input.setPlaceholderText("أدخل قيمة للترشيح...")
        self.filter_value_input.setStyleSheet("""
            QLineEdit {
                background-color: #ffffff;
                border: none;
                border-radius: 5px;
                padding: 10px;
                color: #000000;
                font-size: 16px;
                text-align: right;
            }
        """)
        self.filter_value_input.setLayoutDirection(Qt.RightToLeft)

        filter_value_layout.addWidget(filter_value_label)
        filter_value_layout.addWidget(self.filter_value_input)
        dialog_layout.addSpacing(15)
        dialog_layout.addLayout(filter_value_layout)

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        apply_btn = QPushButton("تطبيق")
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff6a0e;
                color: #ffffff;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-size: 16px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #e05d00;
            }
        """)
        apply_btn.clicked.connect(self.apply_filter)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                color: #ffffff;
                border: 1px solid #455a64;
                border-radius: 5px;
                padding: 10px;
                font-size: 16px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        cancel_btn.clicked.connect(self.filter_dialog.reject)

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(apply_btn)
        dialog_layout.addSpacing(15)
        dialog_layout.addLayout(buttons_layout)

        self.filter_dialog.exec_()  # Ensure modal behavior

    def apply_filter(self):
        """
        Apply the filter to the table based on selected columns and input value
        """
        if not self.table:
            print("Error: Table reference is missing.")
            return

        selected_filter_columns = [
            column for column, checkbox in self.filter_checkboxes.items() if checkbox.isChecked()
        ]

        filter_value = self.filter_value_input.text().strip().lower()

        if not filter_value or not selected_filter_columns:
            # Reset filter if no value or no columns selected
            for row in range(self.table.rowCount()):
                self.table.setRowHidden(row, False)
            self.filter_dialog.accept()
            return

        column_indices = [
            col_idx for col_idx in range(self.table.columnCount())
            if self.table.horizontalHeaderItem(col_idx).text() in selected_filter_columns
        ]

        for row in range(self.table.rowCount()):
            row_visible = any(
                filter_value in self.table.item(row, col_idx).text().lower()
                for col_idx in column_indices
            )
            self.table.setRowHidden(row, not row_visible)

        self.filter_dialog.accept()

class AddForm(QDialog):
    """
    Dialog for adding new evaluation data
    """
    def __init__(self, parent=None, session=None):
        super().__init__(parent)
        self.setWindowTitle("إضافة تنقيط")
        self.setFixedSize(500, 700)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)

        current_user_account_number = None
        if hasattr(parent, 'current_user_data') and parent.current_user_data:
            current_user_account_number = parent.current_user_data.get('account_number')

        self.current_user_account_number = current_user_account_number 
        
        self.session = session
        self.controller = EvaluationController(self.session,current_user_account_number)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        
        scroll_area.setStyleSheet(f"background-color: {DARK_BG};")

        center_container = QWidget()
        center_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        center_layout = QHBoxLayout(center_container)

        form_container_wrapper = QWidget()
        form_container_wrapper.setFixedWidth(500)
        form_wrapper_layout = QVBoxLayout(form_container_wrapper)
        form_wrapper_layout.setContentsMargins(20, 20, 20, 20)
        form_wrapper_layout.setSpacing(15)

        self.title_label = QLabel("إضافة تنقيط")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px 0;")

        form_container = QWidget()
        form_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")
        container_layout = QVBoxLayout(form_container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.setSpacing(15)

        fields = [

            {"name": "السنة"},
            {"name": "إسم الموظف"},
            {"name": "لقب الموظف"},
            {"name": "النقطة السنوية"},
            {"name": "نقطة المردودية 1"},
            {"name": "نقطة المردودية 2"},
            {"name": "نقطة المردودية 3"},
            {"name": "نقطة المردودية 4"},

        ]

        self.add_form_fields = {}

        for field in fields:
            field_layout = QVBoxLayout()
            field_layout.setSpacing(5)


            label = QLabel(field["name"])
            label.setStyleSheet("font-size: 16px; font-weight: bold;")

            input_field = QLineEdit()
            
            input_field.setAlignment(Qt.AlignRight)
            input_field.setLayoutDirection(Qt.RightToLeft)
            input_field.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {DARKER_BG};
                    border: none;
                    border-radius: 5px;
                    padding: 8px;
                    color: {WHITE};
                    font-size: 16px;
                    text-align: right;
                }}
            """)

            if field["name"] == "السنة":
                current_year = str(datetime.now().year)
                input_field.setText(current_year)

            field_layout.addWidget(label)
            field_layout.addWidget(input_field)
            container_layout.addLayout(field_layout)

            self.add_form_fields[field["name"]] = input_field

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        save_btn = QPushButton("حفظ")
        save_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #3d8b40;
            }}
        """)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)



        save_btn.clicked.connect(self.save_new_entry)
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)

        container_layout.addLayout(buttons_layout)

        form_wrapper_layout.addWidget(self.title_label)
        form_wrapper_layout.addWidget(form_container)

        center_layout.addStretch()
        center_layout.addWidget(form_container_wrapper)
        center_layout.addStretch()

        scroll_area.setWidget(center_container)
        main_layout.addWidget(scroll_area)

    def save_new_entry(self):
        # Extract fields and strip whitespace
        annee = self.add_form_fields["السنة"].text().strip()
        name = self.add_form_fields["إسم الموظف"].text().strip()
        lastname = self.add_form_fields["لقب الموظف"].text().strip()

        note_annuelle = self.add_form_fields["النقطة السنوية"].text().strip()

        # Simple helper for showing error dialog and returning early
        def show_error(msg):
            dialog = StyledMessageDialog(self, title="خطأ", message=msg, message_type="warning")
            dialog.exec_()

        # Validate mandatory alphabetic fields
        if not (name and name.isalpha()):
            show_error("الاسم يجب أن يحتوي على حروف فقط ويجب ألا يكون فارغًا.")
            return
        if not (lastname and lastname.isalpha()):
            show_error("اللقب يجب أن يحتوي على حروف فقط ويجب ألا يكون فارغًا.")
            return

        # Validate year
        if not annee.isdigit() or len(annee) > 4:
            show_error("يرجى إدخال السنة بشكل صحيح (أربعة أرقام كحد أقصى).")
            return

        current_year = datetime.now().year
        if int(annee) > current_year:
            show_error(f"لا يمكن أن تكون السنة أكبر من {current_year}.")
            return

        # Validate that "النقطة السنوية" is numeric
        try:
            note_annuelle_float = float(note_annuelle)
        except ValueError:
            show_error("يرجى إدخال النقطة السنوية كرقم صحيح أو عشري.")
            return

        # Extract optional notes safely and validate numeric values for last four notes
        def get_and_validate_note(field_name):
            field = self.add_form_fields.get(field_name)
            if field is None:
                return None
            text = field.text().strip()
            if text == "":
                return None
            try:
                return float(text)
            except ValueError:
                show_error(f"يرجى إدخال قيمة عددية صحيحة أو عشرية في الحقل '{field_name}'.")
                return None

        note1 = get_and_validate_note("نقطة المردودية 1")
        if note1 is None and self.add_form_fields.get("نقطة المردودية 1") and self.add_form_fields[
            "نقطة المردودية 1"].text().strip() != "":
            return

        note2 = get_and_validate_note("نقطة المردودية 2")
        if note2 is None and self.add_form_fields.get("نقطة المردودية 2") and self.add_form_fields[
            "نقطة المردودية 2"].text().strip() != "":
            return

        note3 = get_and_validate_note("نقطة المردودية 3")
        if note3 is None and self.add_form_fields.get("نقطة المردودية 3") and self.add_form_fields[
            "نقطة المردودية 3"].text().strip() != "":
            return

        note4 = get_and_validate_note("نقطة المردودية 4")
        if note4 is None and self.add_form_fields.get("نقطة المردودية 4") and self.add_form_fields[
            "نقطة المردودية 4"].text().strip() != "":
            return

        # Check if sum of notes > note_annuelle
        notes_sum = sum(filter(None, [note1, note2, note3, note4]))  # sum ignores None
        if notes_sum > note_annuelle_float:
            show_error("مجموع نقاط المردودية لا يمكن أن يتجاوز النقطة السنوية.")
            return

        # Save evaluation via controller, catch errors
        
        try:
            controller = EvaluationController(self.session,self.current_user_account_number)
            status = controller.save_evaluation_for_employee(
                name=name.strip(),
                lastname=lastname.strip(),
                annee=annee,
                note_annuelle=str(note_annuelle_float),  # save as string if needed
                note1=str(note1) if note1 is not None else None,
                note2=str(note2) if note2 is not None else None,
                note3=str(note3) if note3 is not None else None,
                note4=str(note4) if note4 is not None else None,
            )

            if status == "not_found":
                show_error("الموظف غير موجود.")
                return
            elif status == "already_exists":
                show_error("الموظف لديه تقييم لهذه السنة. لا يمكن إضافة تقييم آخر.")
                return
            elif status != "success":
                show_error("فشل غير معروف أثناء حفظ التقييم.")
                return

        except Exception as e:
            dialog = StyledMessageDialog(self, title="خطأ", message=f"حدث خطأ أثناء الحفظ: {str(e)}",
                                         message_type="error")
            dialog.exec_()
            return

        # Show success dialog
        dialog = StyledMessageDialog(self, title="تم الحفظ", message="تم إضافة التنقيط بنجاح.",
                                     message_type="information")
        dialog.exec_()
        self.accept()

class PreviousYears(QMainWindow):
    closed = pyqtSignal()
    data_changed = pyqtSignal()
    def __init__(self, parent=None,session=None):
        super().__init__(parent)
        self.setWindowTitle("سجل التنقيط")
        self.setGeometry(150, 150, 1000, 600)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")
        current_user_account_number = None
        if hasattr(parent, 'current_user_data') and parent.current_user_data:
            current_user_account_number = parent.current_user_data.get('account_number')

        self.session = session
        # CORRECTION: Passer current_user_account_number au contrôleur
        self.controller = EvaluationController(self.session, current_user_account_number)
        self.setWindowFlags(
            Qt.Window |
            Qt.WindowCloseButtonHint |
            Qt.WindowMaximizeButtonHint |
            Qt.WindowMinimizeButtonHint
        )

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Create and add the top bar
        self.top_bar_widget, self.export_button_topbar, self.print_button_topbar = create_top_bar(
            self,           # parent
            main_layout,    # layout où ajouter la topbar
            None,           # sidebar_toggle (pas nécessaire ici a priori)
            excel_export_action_callback=self.export_data_to_excel, # Callback Excel
            pdf_print_action_callback=self.print_data_to_pdf       # Callback PDF
        )

        # Action bar setup
        action_bar = QWidget()
        action_layout = QHBoxLayout(action_bar)
        action_layout.setContentsMargins(0, 0, 0, 15)

        # Filter section container
        filter_container = QWidget()
        filter_container_layout = QHBoxLayout(filter_container)
        filter_container_layout.setContentsMargins(0, 0, 0, 0)
        filter_container_layout.setSpacing(10)

        # Filter button with icon and text
        self.filter_btn = QPushButton("ترشيح")
        self.filter_btn.setIcon(QIcon("pics/filter.png"))  # Replace with your icon path
        self.filter_btn.setIconSize(QSize(20, 20))
        self.filter_btn.setFixedHeight(35)
        self.filter_btn.setToolTip("ترشيح")
        self.filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 5px;
                color: #ffffff;
                font-weight: bold;
                font-size: 16px;
                padding: 0 10px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        self.filter_btn.clicked.connect(self.show_filter_dialog)

        filter_container_layout.addWidget(self.filter_btn)

        # Add filter container to the action layout
        action_layout.addWidget(filter_container)

        # Add a stretch to push other widgets (if any) to the right or left as you want
        action_layout.addStretch()

        # Add action_bar to main layout
        main_layout.addWidget(action_bar)

        # Title container with label + refresh button side by side
        title_container = QWidget()
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 15)
        title_layout.setSpacing(10)

        title_label = QLabel("سجل التنقيط السنوي")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: white;")
        title_label.setAlignment(Qt.AlignLeft)

        refresh_btn = QPushButton("⟳")
        refresh_btn.setFixedSize(40, 40)
        refresh_btn.setToolTip("تحديث الجدول")
        refresh_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {MEDIUM_BG};
                border: none;
                border-radius: 17px;
                color: {WHITE};
                font-weight: bold;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: {LIGHT_BG};
            }}
            QPushButton:pressed {{
                background-color: {DARKER_BG};
            }}
        """)
        refresh_btn.clicked.connect(self.reload_table_past)

        title_layout.addWidget(title_label)
        title_layout.addStretch()
        title_layout.addWidget(refresh_btn)

        main_layout.addWidget(title_container)

        # Table container with rounded corners
        table_container = QWidget()
        table_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(15, 15, 15, 15)

        # Table
        self.consultation_table = QTableWidget()
        self.consultation_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                border: none;
                gridline-color: {LIGHT_BG};
                font-size: 16px; 
            }}
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid {LIGHT_BG};
            }}
            QTableWidget::item:selected {{
                background-color: {ORANGE};
                color: {WHITE};
            }}
            QHeaderView::section {{
                background-color: {DARKER_BG};
                color: {WHITE};
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 16px; 
            }}
            QScrollBar:vertical {{
                background: {MEDIUM_BG};
                width: 12px;
            }}
            QScrollBar::handle:vertical {{
                background: {LIGHT_BG};
                border-radius: 6px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)

        columns = [
            "رقم التنقيط", "السنة", "رقم الموظف", "الإسم", "اللقب",
            "النقطة السنوية", "نقطة\nالمردودية 1", "نقطة\nالمردودية 2",
            "نقطة\nالمردودية 3", "نقطة\nالمردودية 4", "المعدل"
        ]
        self.consultation_table.setColumnCount(len(columns))
        self.consultation_table.setHorizontalHeaderLabels(columns)

        header = self.consultation_table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionsMovable(False)
        header.setSectionsClickable(False)

        self.consultation_table.verticalHeader().setVisible(False)
        self.consultation_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.consultation_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.consultation_table.setEditTriggers(QTableWidget.NoEditTriggers)

        for i in range(len(columns)):
            self.consultation_table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

        # Align existing rows (optional)
        for row in range(self.consultation_table.rowCount()):
            for col in range(self.consultation_table.columnCount()):
                item = self.consultation_table.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

        # Add table to container
        table_layout.addWidget(self.consultation_table)

        # Paginator widget
        self.paginator = tablepaginator(table=self.consultation_table, rows_per_page=10)
        table_layout.addWidget(self.paginator)

        # Add table container to the main layout
        main_layout.addWidget(table_container)

        # Buttons layout
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(20)
        buttons_layout.setContentsMargins(0, 20, 0, 0)

        edit_btn = QPushButton("تعديل")
        close_btn = QPushButton("إغلاق")
        delete_btn = QPushButton("حذف")
        for btn in (edit_btn, delete_btn,close_btn):
            btn.setFixedSize(160, 45)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
            """)

        edit_btn.clicked.connect(self.on_edit_button_clicked2)
        delete_btn.clicked.connect(self.show_delete_dialog)
        close_btn.clicked.connect(self.close)

        buttons_layout.addWidget(edit_btn)

        buttons_layout.addWidget(delete_btn)
        buttons_layout.addWidget(close_btn)
        main_layout.addLayout(buttons_layout)

        # Initial load
        self.reload_table_past()
        self.paginator.update_total_rows()

    def reload_table_past(self):
        # Clear existing rows
        self.consultation_table.setRowCount(0)

        # Load data from the database
        table_data = self.controller.load_evaluations_with_employee_names()
        print("Reloaded data:", table_data[0] if table_data else "No data")

        if not table_data:
            return

        self.consultation_table.setRowCount(len(table_data))

        for row_idx, row_data in enumerate(table_data):
            def safe_text(value):
                return str(value) if value is not None else ""

            self.consultation_table.setItem(row_idx, 0, QTableWidgetItem(safe_text(row_data["idEvaluation"])))
            self.consultation_table.setItem(row_idx, 1, QTableWidgetItem(safe_text(row_data["Annee"])))
            self.consultation_table.setItem(row_idx, 2, QTableWidgetItem(safe_text(row_data["idemploye"])))
            self.consultation_table.setItem(row_idx, 3, QTableWidgetItem(safe_text(row_data["nom"])))
            self.consultation_table.setItem(row_idx, 4, QTableWidgetItem(safe_text(row_data["prenom"])))
            self.consultation_table.setItem(row_idx, 5, QTableWidgetItem(safe_text(row_data["NoteAnnuelle"])))
            self.consultation_table.setItem(row_idx, 6, QTableWidgetItem(safe_text(row_data["Note1"])))
            self.consultation_table.setItem(row_idx, 7, QTableWidgetItem(safe_text(row_data["Note2"])))
            self.consultation_table.setItem(row_idx, 8, QTableWidgetItem(safe_text(row_data["Note3"])))
            self.consultation_table.setItem(row_idx, 9, QTableWidgetItem(safe_text(row_data["Note4"])))
            self.consultation_table.setItem(row_idx, 10, QTableWidgetItem(safe_text(row_data["Moyenne"])))

            for col_idx in range(self.consultation_table.columnCount()):
                self.consultation_table.item(row_idx, col_idx).setTextAlignment(Qt.AlignCenter)
    # --- MÉTHODES D'EXPORTATION ET D'IMPRESSION POUR PreviousYears ---
    def _get_table_data_as_lists(self): # Opère sur self.consultation_table
        """Helper function to extract data ONLY from VISIBLE rows in self.consultation_table."""
        if not hasattr(self, 'consultation_table') or not self.consultation_table or self.consultation_table.rowCount() == 0:
            print(f"Avertissement dans _get_table_data_as_lists pour {self.__class__.__name__}: self.consultation_table non disponible ou vide.")
            return None, None

        headers = [self.consultation_table.horizontalHeaderItem(i).text()
                   for i in range(self.consultation_table.columnCount())]

        visible_row_data = []
        for row in range(self.consultation_table.rowCount()):
            if not self.consultation_table.isRowHidden(row):
                row_data = []
                for col in range(self.consultation_table.columnCount()):
                    item = self.consultation_table.item(row, col)
                    row_data.append(item.text() if item else "")
                visible_row_data.append(row_data)
        return headers, visible_row_data

    def export_data_to_excel(self):
        print(f"{self.__class__.__name__}: export_data_to_excel (années précédentes) called")
        headers, data = self._get_table_data_as_lists()

        if not data:
            QMessageBox.information(self, "لا بيانات للتصدير", "لا توجد صفوف ظاهرة في جدول التنقيط للسنوات السابقة للتصدير.")
            return

        filePath, _ = QFileDialog.getSaveFileName(self, "تصدير إلى Excel (تنقيط سابق)", 
                                                  os.path.expanduser("~/Documents/Evaluations_Annees_Precedentes.xlsx"),
                                                  "Excel Workbook (*.xlsx);;All Files (*)")
        if not filePath: return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "تنقيط سنوات سابقة"
            sheet.sheet_view.rightToLeft = True
            sheet.append(headers)
            for row_values in data: sheet.append(row_values)
            for col_idx, column_cells in enumerate(sheet.columns):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = length + 5
            workbook.save(filePath)
            QMessageBox.information(self, "نجاح التصدير", f"تم تصدير البيانات بنجاح إلى:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ أثناء تصدير الملف: {e}")
            print(f"Error exporting Evaluations (previous years) to Excel: {e}")

    def print_data_to_pdf(self):
        print(f"{self.__class__.__name__}: print_data_to_pdf (années précédentes) called")
        headers, data = self._get_table_data_as_lists()

        if not data:
            QMessageBox.information(self, "لا بيانات للطباعة", "لا توجد صفوف ظاهرة في جدول التنقيط للسنوات السابقة للطباعة.")
            return

        filePath, _ = QFileDialog.getSaveFileName(self, "طباعة إلى PDF (تنقيط سابق)", 
                                                  os.path.expanduser("~/Documents/Evaluations_Annees_Precedentes.pdf"),
                                                  "PDF Document (*.pdf);;All Files (*)")
        if not filePath: return

        html_content = "<html><head><meta charset='UTF-8'>"
        html_content += """
        <style>
            @font-face { font-family: 'DejaVu Sans'; }
            body { direction: rtl; font-family: 'Arial', 'DejaVu Sans', sans-serif; font-size: 8pt; } /* Police un peu plus petite */
            table { width: 100%; border-collapse: collapse; margin-top: 10px; }
            th, td { border: 1px solid #333; padding: 3px; text-align: right; word-wrap: break-word; } /* Padding réduit */
            th { background-color: #f0f0f0; font-weight: bold; }
            caption { font-size: 1.0em; font-weight: bold; margin-bottom: 7px; text-align: center; }
        </style>
        </head><body>
        """
        html_content += "<table><caption>سجل التنقيط (السنوات السابقة)</caption><thead><tr>"
        for header in headers: html_content += f"<th>{header}</th>"
        html_content += "</tr></thead><tbody>"
        for row_values in data:
            html_content += "<tr>"
            for cell_value in row_values: html_content += f"<td>{html.escape(str(cell_value))}</td>"
            html_content += "</tr>"
        html_content += "</tbody></table></body></html>"

        try:
            css_style = CSS(string='@page { size: A3 landscape; margin: 0.7cm; }') # A3 landscape
            HTML(string=html_content).write_pdf(filePath, stylesheets=[css_style])
            QMessageBox.information(self, "نجاح الطباعة", f"تم إنشاء ملف PDF بنجاح:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في الطباعة", f"حدث خطأ أثناء إنشاء ملف PDF: {e}\n"
                                 "Vérifiez WeasyPrint et ses dépendances.")
            print(f"Error printing Evaluations (previous years) to PDF: {e}")

    def closeEvent(self, event):
        self.closed.emit()
        super().closeEvent(event)

    def on_edit_button_clicked2(self):
        table = self.consultation_table
        indexes = table.selectionModel().selectedRows()

        if len(indexes) == 0:
            dialog = StyledMessageDialog(self, title=" تحذير", message="يرجى تحديد صف للتعديل.",
                                         message_type="warning")
            dialog.exec_()
            return

        selected_row = indexes[0].row()

        fields = [
            {"name": "السنة"},
            {"name": "رقم الموظف"},
            {"name": "الإسم"},
            {"name": "اللقب"},
            {"name": "النقطة السنوية"},
            {"name": "نقطة المردودية 1"},
            {"name": "نقطة المردودية 2"},
            {"name": "نقطة المردودية 3"},
            {"name": "نقطة المردودية 4"},
        ]

        form_fields_indices = {
            "السنة": 1,
            "رقم الموظف": 2,
            "الإسم": 3,
            "اللقب": 4,
            "النقطة السنوية": 5,
            "نقطة المردودية 1": 6,
            "نقطة المردودية 2": 7,
            "نقطة المردودية 3": 8,
            "نقطة المردودية 4": 9,
            "رقم التنقيط": 0,  # Added to ensure evaluation ID column is known
        }

        # Charger les données de la ligne sélectionnée dans un dict
        data = {}
        for field in fields:
            name = field["name"]
            col_idx = form_fields_indices.get(name, None)
            if col_idx is not None:
                item = table.item(selected_row, col_idx)
                data[name] = item.text() if item else ""
            else:
                data[name] = ""

        # Also load the evaluation_id from the table
        eval_col = form_fields_indices.get("رقم التنقيط")
        if eval_col is not None:
            item = table.item(selected_row, eval_col)
            data["رقم التنقيط"] = item.text() if item else ""
        else:
            data["رقم التنقيط"] = ""

        self.edit_window = QWidget()
        self.edit_window.setWindowTitle("تعديل البيانات")
        self.edit_window.setFixedSize(500, 700)
        self.edit_window.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")
        self.edit_window.setWindowFlags(
            self.edit_window.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)

        main_layout = QVBoxLayout(self.edit_window)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setStyleSheet(f"background-color: {DARK_BG};")

        center_container = QWidget()
        center_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        center_layout = QHBoxLayout(center_container)

        form_container_wrapper = QWidget()
        form_container_wrapper.setFixedWidth(500)
        form_wrapper_layout = QVBoxLayout(form_container_wrapper)
        form_wrapper_layout.setContentsMargins(20, 20, 20, 20)
        form_wrapper_layout.setSpacing(15)

        self.title_label = QLabel("تعديل البيانات")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin: 10px 0; color: white;")

        form_container = QWidget()
        form_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")

        container_layout = QVBoxLayout(form_container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.setSpacing(15)

        self.edit_form_fields = {}

        editable_fields = {
            "السنة",
            "النقطة السنوية",
            "نقطة المردودية 1",
            "نقطة المردودية 2",
            "نقطة المردودية 3",
            "نقطة المردودية 4"
        }

        for field in fields:
            name = field["name"]
            

            field_layout = QVBoxLayout()
            field_layout.setSpacing(5)

            label = QLabel(name)
            label.setStyleSheet("font-size: 16px; font-weight: bold; color: white;")

            input_field = QLineEdit()
          
            input_field.setText(data.get(name, ""))
            input_field.setAlignment(Qt.AlignRight)
            input_field.setLayoutDirection(Qt.RightToLeft)

            # Read-only sauf pour les champs éditables
            input_field.setReadOnly(name not in editable_fields)

            input_field.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {DARKER_BG};
                    border: none;
                    border-radius: 5px;
                    padding: 8px;
                    color: {WHITE};
                    font-size: 16px;
                }}
            """)

            field_layout.addWidget(label)
            field_layout.addWidget(input_field)
            container_layout.addLayout(field_layout)

            self.edit_form_fields[name] = input_field

        # Add hidden or read-only field for evaluation id if needed, to keep track
        # or store it separately if your logic requires it (optional).

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        save_btn = QPushButton("حفظ")
        save_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #3d8b40;
            }}
        """)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)

        cancel_btn.clicked.connect(self.edit_window.close)
        save_btn.clicked.connect(lambda: self.save_edited_data(selected_row, form_fields_indices=form_fields_indices))

        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)

        container_layout.addLayout(buttons_layout)

        form_wrapper_layout.addWidget(self.title_label)
        form_wrapper_layout.addWidget(form_container)

        center_layout.addStretch()
        center_layout.addWidget(form_container_wrapper)
        center_layout.addStretch()

        scroll_area.setWidget(center_container)

        main_layout.addWidget(scroll_area)

        self.edit_window.show()

    def save_edited_data(self, row, form_fields_indices=None):
        if form_fields_indices is None:
            form_fields_indices = self.form_fields_indices  # your map

        fields = list(self.edit_form_fields.items())

        # Validation du champ "السنة"
        year_field = self.edit_form_fields.get("السنة")
        if year_field is None:
            return

        year_text = year_field.text().strip()

        if not year_text.isdigit():
            dialog = StyledMessageDialog(self.edit_window, title="خطأ", message="يرجى إدخال السنة كعدد صحيح.",
                                         message_type="warning")
            dialog.exec_()
            self.edit_window.activateWindow()
            self.edit_window.raise_()
            return

        entered_year = int(year_text)
        current_year = datetime.now().year

        if entered_year > current_year:
            dialog = StyledMessageDialog(self.edit_window, title="خطأ",
                                         message=f"لا يمكن أن تكون السنة أكبر من {current_year}.",
                                         message_type="warning")
            dialog.exec_()
            self.edit_window.activateWindow()
            self.edit_window.raise_()
            return

        numeric_fields = [
            "النقطة السنوية",
            "نقطة المردودية 1",
            "نقطة المردودية 2",
            "نقطة المردودية 3",
            "نقطة المردودية 4",
        ]

        for field_name in numeric_fields:
            value = self.edit_form_fields[field_name].text().strip()
            if value != "":
                try:
                    float(value)
                except ValueError:
                    dialog = StyledMessageDialog(self.edit_window, title="خطأ",
                                                 message=f"يرجى إدخال قيمة عددية صحيحة أو عشرية في الحقل '{field_name}'.",
                                                 message_type="warning")
                    dialog.exec_()
                    self.edit_window.activateWindow()
                    self.edit_window.raise_()
                    return

        field_dict = {name: field.text().strip() for name, field in fields}

        evaluation_id = field_dict.get("رقم التنقيط")

        if not evaluation_id:
            eval_col = form_fields_indices.get("رقم التنقيط")
            if eval_col is None:
                dialog = StyledMessageDialog(self.edit_window, title="خطأ",
                                             message="تعذر العثور على عمود 'رقم التنقيط' في الجدول.",
                                             message_type="error")
                dialog.exec_()
                return

            item = self.consultation_table.item(row, eval_col)
            if item:
                evaluation_id = item.text()
            else:
                dialog = StyledMessageDialog(self.edit_window, title="خطأ",
                                             message="معرف التنقيط غير موجود في الجدول.",
                                             message_type="error")
                dialog.exec_()
                return

        # التحقق من أن مجموع نقاط المردودية لا يتجاوز النقطة السنوية
        try:
            note_annuelle = float(field_dict.get("النقطة السنوية", "0") or "0")
            note1 = float(field_dict.get("نقطة المردودية 1", "0") or "0")
            note2 = float(field_dict.get("نقطة المردودية 2", "0") or "0")
            note3 = float(field_dict.get("نقطة المردودية 3", "0") or "0")
            note4 = float(field_dict.get("نقطة المردودية 4", "0") or "0")

            total = note1 + note2 + note3 + note4

            if total > note_annuelle + 0.01:  # Margin for rounding error
                dialog = StyledMessageDialog(
                    self.edit_window,
                    title="خطأ",
                    message=f"مجموع نقاط المردودية ({total:.2f}) لا يجب أن يتجاوز النقطة السنوية ({note_annuelle:.2f}).",
                    message_type="warning"
                )
                dialog.exec_()
                return

        except ValueError:
            dialog = StyledMessageDialog(self.edit_window, title="خطأ",
                                         message="يرجى التأكد من أن جميع النقاط مدخلة بشكل صحيح كأرقام.",
                                         message_type="warning")
            dialog.exec_()
            return

        try:
            success = self.controller.update_evaluation_past(
                evaluation_id=evaluation_id,
                annee=field_dict.get("السنة"),
                note_annuelle=field_dict.get("النقطة السنوية"),
                note1=field_dict.get("نقطة المردودية 1"),
                note2=field_dict.get("نقطة المردودية 2"),
                note3=field_dict.get("نقطة المردودية 3"),
                note4=field_dict.get("نقطة المردودية 4"),
            )
        except Exception as e:
            dialog = StyledMessageDialog(self.edit_window, title="خطأ", message=str(e), message_type="error")
            dialog.exec_()
            self.edit_window.activateWindow()
            self.edit_window.raise_()
            return

        if success:
            if row >= self.consultation_table.rowCount():
                self.consultation_table.insertRow(row)

            for field_name, field in fields:
                new_text = field.text().strip()
                col_idx = form_fields_indices.get(field_name)
                print(f"Champ: {field_name}, Colonne: {col_idx}, Valeur: {new_text}")
                if col_idx is not None:
                    item = QTableWidgetItem(new_text)
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.consultation_table.setItem(row, col_idx, item)

            dialog = StyledMessageDialog(self.edit_window, title="نجاح", message="تم تحديث البيانات بنجاح.",
                                         message_type="success")
            dialog.exec_()
            self.reload_table_past()
            self.edit_window.close()

        else:
            dialog = StyledMessageDialog(self.edit_window, title="خطأ", message="حدث خطأ أثناء تحديث البيانات.",
                                         message_type="error")
            dialog.exec_()

    def show_delete_dialog(self):
        table = self.consultation_table
        indexes = table.selectionModel().selectedRows()

        if not indexes:
            dialog = QDialog(self)
            dialog.setWindowTitle("  تحذير")
            dialog.setFixedSize(450, 180)
            dialog.setStyleSheet(f"background-color: {DARKER_BG};")

            layout = QVBoxLayout(dialog)
            layout.setContentsMargins(25, 25, 25, 25)
            layout.setSpacing(20)

            title_label = QLabel(" تحذير")
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setStyleSheet(f"color: {YELLOW}; font-size: 18px; font-weight: bold;")

            message_label = QLabel("يرجى  تحديد  صف  للحذف.")
            message_label.setAlignment(Qt.AlignCenter)
            message_label.setStyleSheet(f"color: {WHITE}; font-size: 16px;")

            layout.addWidget(title_label)
            layout.addWidget(message_label)

            dialog.exec_()
            return

        selected_row = indexes[0].row()
        consultation_id = table.model().index(selected_row, 0).data()

        confirm_dialog = QDialog(self)
        confirm_dialog.setWindowTitle("تأكيد الحذف")
        confirm_dialog.setFixedSize(450, 220)
        confirm_dialog.setStyleSheet(f"background-color: {DARKER_BG};")

        layout = QVBoxLayout(confirm_dialog)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(30)

        title_label = QLabel("تأكيد الحذف")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"color: {YELLOW}; font-size: 18px; font-weight: bold;")

        message_label = QLabel("هل أنت متأكد من حذف هذا السجل؟")
        message_label.setAlignment(Qt.AlignCenter)
        message_label.setStyleSheet(f"color: {WHITE}; font-size: 16px;")

        layout.addWidget(title_label)
        layout.addWidget(message_label)

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(20)

        confirm_btn = QPushButton("احذف")
        confirm_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                font-size: 16px;
                font-weight: bold;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #3d8b40;
            }}
        """)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                font-size: 16px;
                font-weight: bold;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)

        confirm_btn.clicked.connect(lambda: self.confirm_delete_and_show_message(confirm_dialog, selected_row))
        cancel_btn.clicked.connect(confirm_dialog.reject)

        buttons_layout.addWidget(confirm_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

        confirm_dialog.exec_()

    def confirm_delete_and_show_message(self, dialog, row_index):
        item = self.consultation_table.item(row_index, 0)
        if item is None:
            print("No item found at the selected row.")
            return

        consultation_id_text = item.text()
        if not consultation_id_text.isdigit():
            print(f" Invalid consultation ID: {consultation_id_text}")
            return

        consultation_id = int(consultation_id_text)

        # Delete from DB via controller
        self.controller.delete_evaluation(consultation_id)

        # Remove row from table
        self.consultation_table.removeRow(row_index)

        # Close dialog
        dialog.accept()

        # Show success message
        success_dialog = QDialog(self)
        success_dialog.setWindowTitle("تم الحذف")
        success_dialog.setFixedSize(450, 180)
        success_dialog.setStyleSheet(f"background-color: {DARKER_BG};")

        layout = QVBoxLayout(success_dialog)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)

        title_label = QLabel("تم الحذف")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"color: {GREEN}; font-size: 18px; font-weight: bold;")

        message_label = QLabel("تم حذف السجل بنجاح.")
        message_label.setAlignment(Qt.AlignCenter)
        message_label.setStyleSheet(f"color: {WHITE}; font-size: 16px;")

        layout.addWidget(title_label)
        layout.addWidget(message_label)

        success_dialog.exec_()



    def show_filter_dialog(self):
        """
        Show the filter dialog with styled checkboxes and a value input for filtering
        """
        self.filter_dialog = QDialog(self)
        self.filter_dialog.setWindowTitle("ترشيح")
        self.filter_dialog.setFixedWidth(350)
        self.filter_dialog.setStyleSheet("background-color: #37474f; color: #ffffff;")

        dialog_layout = QVBoxLayout(self.filter_dialog)
        dialog_layout.setContentsMargins(20, 20, 20, 20)

        title_label = QLabel("اختر الأعمدة للترشيح:")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 15px;")
        dialog_layout.addWidget(title_label)

        self.filter_checkboxes = {}

        # Use actual column headers from the table
        columns = [
            "رقم التنقيط", "السنة", "رقم الموظف", "الإسم", "اللقب",
            "النقطة السنوية", "نقطة المردودية 1", "نقطة المردودية 2",
            "نقطة المردودية 3", "نقطة المردودية 4", "المعدل"
        ]
        for column in columns:
            checkbox = QCheckBox(column)
            checkbox.setStyleSheet("""
                QCheckBox {
                    font-size: 16px;
                    padding: 8px 0;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
                QCheckBox::indicator:checked {
                    background-color: #ff6a0e;
                    border: 2px solid #ffffff;
                }
            """)
            dialog_layout.addWidget(checkbox)
            self.filter_checkboxes[column] = checkbox

        # Value input
        filter_value_layout = QHBoxLayout()
        filter_value_layout.setSpacing(10)

        filter_value_label = QLabel("قيمة الترشيح:")
        filter_value_label.setStyleSheet("font-size: 16px;")

        self.filter_value_input = QLineEdit()
        self.filter_value_input.setPlaceholderText("أدخل قيمة للترشيح...")
        self.filter_value_input.setStyleSheet("""
            QLineEdit {
                background-color: #ffffff;
                border: none;
                border-radius: 5px;
                padding: 10px;
                color: #000000;
                font-size: 16px;
                text-align: right;
            }
        """)
        self.filter_value_input.setLayoutDirection(Qt.RightToLeft)

        filter_value_layout.addWidget(filter_value_label)
        filter_value_layout.addWidget(self.filter_value_input)
        dialog_layout.addSpacing(15)
        dialog_layout.addLayout(filter_value_layout)

        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        apply_btn = QPushButton("تطبيق")
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff6a0e;
                color: #ffffff;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-size: 16px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #e05d00;
            }
        """)
        apply_btn.clicked.connect(self.apply_filter)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                color: #ffffff;
                border: 1px solid #455a64;
                border-radius: 5px;
                padding: 10px;
                font-size: 16px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        cancel_btn.clicked.connect(self.filter_dialog.reject)

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(apply_btn)
        dialog_layout.addSpacing(15)
        dialog_layout.addLayout(buttons_layout)

        self.filter_dialog.exec_()

    def apply_filter(self):
        """
        Apply the filter to the table based on selected columns and input value.
        """
        if not self.consultation_table:
            print("Error: Table reference is missing.")
            return

        selected_filter_columns = [
            column for column, checkbox in self.filter_checkboxes.items() if checkbox.isChecked()
        ]

        print("Selected filter columns:", selected_filter_columns)

        filter_value = self.filter_value_input.text().strip().lower()
        print("Filter value:", filter_value)

        # If no filter value is provided or no columns selected, reset the filter
        if not filter_value or not selected_filter_columns:
            print("Resetting filter")
            for row in range(self.consultation_table.rowCount()):
                self.consultation_table.setRowHidden(row, False)
            self.filter_dialog.accept()
            return

        # Get the column indices based on selected columns
        column_indices = [
            col_idx for col_idx in range(self.consultation_table.columnCount())
            if self.consultation_table.horizontalHeaderItem(col_idx).text() in selected_filter_columns
        ]
        print("Column indices for filtering:", column_indices)

        # Apply the filter on each row
        for row in range(self.consultation_table.rowCount()):
            row_visible = False
            for col_idx in column_indices:
                item = self.consultation_table.item(row, col_idx)
                if item and filter_value in item.text().lower():
                    row_visible = True
                    break
            print(f"Row {row} visibility: {row_visible}")
            self.consultation_table.setRowHidden(row, not row_visible)

        self.filter_dialog.accept()


class EditForm(QDialog):
    """
    Dialog for editing existing data
    """
    def __init__(self, parent, data, row_index, table,session=None):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("تعديل البيانات")
        self.setFixedSize(500, 700)  # Set fixed size
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")

        self.session = session
        current_user_account_number = parent.current_user_data.get('account_number') if hasattr(parent, 'current_user_data') else None
        self.controller = EvaluationController(self.session, current_user_account_number)



        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)

        self.table = table
        self.row_index = row_index
        self.data = data
        self.evaluation_id = data.get("رقم التنقيط")  # Keep track of the ID here

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # Create a scroll area to allow scrolling
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)# Make the widget inside scrollable
        scroll_area.setStyleSheet(f"background-color: {DARK_BG};")

        # Container for form fields and labels
        center_container = QWidget()
        center_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        center_layout = QHBoxLayout(center_container)

        form_container_wrapper = QWidget()
        form_container_wrapper.setFixedWidth(500)
        form_wrapper_layout = QVBoxLayout(form_container_wrapper)
        form_wrapper_layout.setContentsMargins(20, 20, 20, 20)
        form_wrapper_layout.setSpacing(15)

        self.title_label = QLabel("تعديل البيانات")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin: 10px 0;")

        form_container = QWidget()
        form_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")

        container_layout = QVBoxLayout(form_container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.setSpacing(15)

        fields = [
            {"name": "رقم التنقيط"},
            {"name": "السنة"},
            {"name": "رقم الموظف"},
            {"name": "الإسم"},
            {"name": "اللقب"},
            {"name": "النقطة السنوية"},
            {"name": "نقطة المردودية 1"},
            {"name": "نقطة المردودية 2"},
            {"name": "نقطة المردودية 3"},
            {"name": "نقطة المردودية 4"},
        ]

        self.edit_form_fields = {}

        for i, field in enumerate(fields):
            field_layout = QVBoxLayout()
            field_layout.setSpacing(5)

            label = QLabel(field["name"])
            label.setStyleSheet("font-size: 16px; font-weight: bold;")

            input_field = QLineEdit()
            
            input_field.setAlignment(Qt.AlignRight)

            if field["name"] in data:
                input_field.setText(data[field["name"]])

            # Hide the first field (ID) and its label
            if field["name"] == "رقم التنقيط":
                label.hide()
                input_field.hide()

            # Make all fields not editable except the last 5 and the year
            if i < len(fields) - 5 and field["name"] != "السنة":
                input_field.setReadOnly(True)

            input_field.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {DARKER_BG};
                    border: none;
                    border-radius: 5px;
                    padding: 8px;
                    color: {WHITE};
                    font-size: 16px;
                    text-align: right;
                }}
            """)
            input_field.setLayoutDirection(Qt.RightToLeft)

            field_layout.addWidget(label)
            field_layout.addWidget(input_field)
            container_layout.addLayout(field_layout)

            self.edit_form_fields[field["name"]] = input_field

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        save_btn = QPushButton("حفظ")
        save_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #3d8b40;
            }}
        """)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)

        save_btn.clicked.connect(self.save_edited_row)
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)

        container_layout.addLayout(buttons_layout)

        form_wrapper_layout.addWidget(self.title_label)
        form_wrapper_layout.addWidget(form_container)

        center_layout.addStretch()
        center_layout.addWidget(form_container_wrapper)
        center_layout.addStretch()

        # Add the form to the scroll area
        scroll_area.setWidget(center_container)

        main_layout.addWidget(scroll_area)

    def save_edited_row(self):
        """
        Save the data from the form fields to the table, with validation.
        """

        # Validate the year field first
        year_input = self.edit_form_fields["السنة"].text().strip()

        if not year_input.isdigit():
            dialog = StyledMessageDialog(self, title="خطأ", message="يرجى إدخال السنة كعدد صحيح.",
                                         message_type="warning")
            dialog.exec_()
            return

        entered_year = int(year_input)
        current_year = datetime.now().year

        if entered_year > current_year:
            dialog = StyledMessageDialog(self, title="خطأ", message=f"لا يمكن أن تكون السنة أكبر من {current_year}.",
                                         message_type="warning")
            dialog.exec_()
            return

        # Validate that each note does not exceed annual score and sum of notes too
        try:
            annual_note = float(self.edit_form_fields["النقطة السنوية"].text().strip())

            notes = []
            for key in ["نقطة المردودية 1", "نقطة المردودية 2", "نقطة المردودية 3", "نقطة المردودية 4"]:
                note_text = self.edit_form_fields[key].text().strip()
                note_value = float(note_text) if note_text else 0.0

                # Validate individual note against annual note
                if note_value > annual_note:
                    dialog = StyledMessageDialog(
                        self,
                        title="خطأ",
                        message=f"قيمة '{key}' ({note_value}) لا يجب أن تتجاوز النقطة السنوية ({annual_note}).",
                        message_type="warning"
                    )
                    dialog.exec_()
                    return

                notes.append(note_value)

            total_notes = sum(notes)
            if total_notes > annual_note:
                dialog = StyledMessageDialog(self, title="خطأ",
                                             message=f"مجموع نقاط المردودية ({total_notes}) لا يجب أن يتجاوز النقطة السنوية ({annual_note}).",
                                             message_type="warning")
                dialog.exec_()
                return

        except ValueError:
            dialog = StyledMessageDialog(self, title="خطأ",
                                         message="ييرجى على الأقل إدخال المعدل السنوي، ويجب أن يكون رقمًا",
                                         message_type="warning")
            dialog.exec_()
            return

        # Prepare data as strings to pass to update_evaluation
        data = {
            "Annee": self.edit_form_fields["السنة"].text().strip(),
            "NoteAnnuelle": self.edit_form_fields["النقطة السنوية"].text().strip(),
            "Note1": self.edit_form_fields["نقطة المردودية 1"].text().strip() or None,
            "Note2": self.edit_form_fields["نقطة المردودية 2"].text().strip() or None,
            "Note3": self.edit_form_fields["نقطة المردودية 3"].text().strip() or None,
            "Note4": self.edit_form_fields["نقطة المردودية 4"].text().strip() or None,
        }

        try:
            result = self.controller.update_evaluation(
                self.evaluation_id,
                data["Annee"],
                data["NoteAnnuelle"],
                data["Note1"],
                data["Note2"],
                data["Note3"],
                data["Note4"]
            )
        except Exception as e:
            dialog = StyledMessageDialog(self, title="خطأ",
                                         message=f"حدث خطأ أثناء تحديث التقييم:\n{str(e)}",
                                         message_type="critical")
            dialog.exec_()
            return

        # Handle the returned status
        if result == "already_exists":
            dialog = StyledMessageDialog(self, title="تنبيه",
                                         message="الموظف لديه تقييم لهذه السنة بالفعل ولا يمكن تكراره.",
                                         message_type="warning")
            dialog.exec_()
            return
        elif result == "not_found":
            dialog = StyledMessageDialog(self, title="خطأ",
                                         message="الموظف لا يملك تنقيطًا، يُرجى إضافته أولاً.",
                                         message_type="critical")
            dialog.exec_()
            return
        elif result == "success":
            self.accept()  # Close dialog on success

if __name__ == '__main__':
    app = QApplication(sys.argv)

    # Set application-wide font for Arabic support
    font = QFont("Arial", 10)
    app.setFont(font)

    # Set right-to-left layout for Arabic
    app.setLayoutDirection(Qt.RightToLeft)

    window = Evaluation()
    window.show()

    sys.exit(app.exec_())