
import re
import sys
import os

from Models import init_db
from Views.topbar import create_top_bar


sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QPushButton, QTableWidget,
                             QTableWidgetItem, QVBoxLayout, QHBoxLayout, QLineEdit, QLabel,
                             QFrame, QDialog, QMessageBox, QCheckBox, QComboBox, QDateEdit,
                             QScrollArea, QSizePolicy, QStackedWidget, QToolButton, QMenu,
                             QListWidget, QListWidgetItem, QGridLayout, QHeaderView, QCompleter,QFileDialog)
from PyQt5.QtGui import QIcon, QPixmap, QFont, QColor, QPalette, QBrush
from PyQt5.QtCore import Qt, QSize, QDate, QTimer, QLocale, pyqtSignal, QStringListModel
from sqlalchemy.orm import Session

import Controllers
from Controllers import EmployeController
from Controllers.Absence import AbsenceController
from Controllers.EmployeController import EmployeeController
from Controllers.Evaluation import EvaluationController
from DatabaseConnection import db
from MessageBox import StyledMessageDialog
from Views.TablePaginator1 import  tablepaginator
from weasyprint import HTML,CSS
import openpyxl
import html
from Models import Employe

# Define color constants based on the screenshots
DARK_BG = "#263238"  # Dark background color
DARKER_BG = "#26282b"  # Darker background for contrast
MEDIUM_BG = "#37474f"  # Medium background for containers
LIGHT_BG = "#455a64"  # Light background for hover states
ORANGE = "#ff6a0e"  # Primary accent color
YELLOW = "#ffc20e"  # Secondary accent color
YELLOW_BTN = "#e6b800"  # Button color
WHITE = "#ffffff"  # Text color
BLACK = "#000000"  # Black text color
GRAY = "#8f8f8f"  # Gray for secondary text
GREEN = "#4CAF50"  # Success color
RED = "#f44336"  # Error/Delete color


class AbsenceManagementSystem(QMainWindow):

    def __init__(self, current_user_data=None, session=None):
        super().__init__()
        self.setWindowTitle("نظام إدارة الموارد البشرية - إدارة الغيابات")
        self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")

        self.current_user_data = current_user_data or {}
        current_user_account_number = None
        if self.current_user_data:
            current_user_account_number = self.current_user_data.get('account_number')
            print(f"DEBUG - Utilisateur actuel dans AbsenceManagement: {current_user_account_number}")
        else:
            print("DEBUG - Aucun utilisateur actuel défini dans AbsenceManagement")     
        # MODIFICATION: Utiliser la session partagée ou créer une nouvelle si nécessaire
        if session:
            self.session = session
            print("DEBUG - Utilisation de la session partagée dans gestionComptes")
        else:
            print("DEBUG - Création d'une nouvelle session dans gestionComptes")
            self.session = init_db('mysql+pymysql://hr:hr@localhost/HR')
        self.controller = AbsenceController(self.session, current_user_account_number)

        # Central widget
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        # Main horizontal layout
        self.main_layout = QHBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)

        # Sidebar and toggle
        #self.sidebar, self.sidebar_toggle = create_sidebar(self, self.main_layout)
        #self.main_layout.addWidget(self.sidebar)

        # Scrollable content area
        self.content_scroll = QScrollArea()
        self.content_scroll.setWidgetResizable(True)
        self.content_scroll.setFrameShape(QFrame.NoFrame)
        self.content_scroll.setStyleSheet(f"background-color: {DARK_BG};")

        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(0)

        self.content_scroll.setWidget(self.content_widget)
        self.main_layout.addWidget(self.content_scroll)

        # Top bar
        #self.top_bar = create_top_bar(self, self.content_layout, self.sidebar_toggle)
        #self.content_layout.addWidget(self.top_bar)

        # Stacked widget
        self.stacked_widget = QStackedWidget()
        self.content_layout.addWidget(self.stacked_widget)

        # Main page setup
        self.main_page = QWidget()
        self.main_page_layout = QVBoxLayout(self.main_page)
        self.main_page_layout.setContentsMargins(15, 15, 15, 15)

        # Title + refresh button
        title_container = QWidget()
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 10)

        table_title = QLabel("تفاصيل جدول الغيابات")
        table_title.setStyleSheet("font-size: 18px; font-weight: bold; color: white;")
        table_title.setAlignment(Qt.AlignLeft)

        refresh_btn = QPushButton("⟳")
        refresh_btn.setFixedSize(40, 40)
        refresh_btn.setToolTip("تحديث")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 17px;
                color: #ffffff;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        refresh_btn.clicked.connect(self.load_data)

        title_layout.addWidget(table_title)
        title_layout.addStretch()
        title_layout.addWidget(refresh_btn)
        self.main_page_layout.addWidget(title_container)

        # Table + action bar
        self.create_table()
        self.action_bar = ActionBar(self, self.table)
        self.main_page_layout.insertWidget(0, self.action_bar)

        # Action buttons
        self.create_action_buttons()

        self.stacked_widget.addWidget(self.main_page)
        self.stacked_widget.setCurrentIndex(0)

    def create_table(self):
        # Container for styling
        table_container = QWidget()
        table_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(15, 15, 15, 15)

        # Create table
        self.table = QTableWidget()
        self.table.setStyleSheet(f"""
           QTableWidget {{
               background-color: {MEDIUM_BG};
               color: {WHITE};
               border: none;
               gridline-color: {LIGHT_BG};
               font-size: 16px;
           }}
           QTableWidget::item {{
               padding: 8px;
               border-bottom: 1px solid {LIGHT_BG};
           }}
           QTableWidget::item:selected {{
               background-color: {ORANGE};
               color: {WHITE};
           }}
           QHeaderView::section {{
               background-color: {DARKER_BG};
               color: {WHITE};
               padding: 8px;
               border: none;
               font-weight: bold;
               font-size: 16px;
           }}
           
           QScrollBar:vertical, QScrollBar:horizontal {{
                width: 0px;
                height: 0px;
            }}
           QScrollBar::handle:vertical {{
               background: {LIGHT_BG};
               min-height: 20px;
               border-radius: 6px;
           }}
       """)

        columns = [
            "السنة", "الشهر", "رقم الموظف", "الإسم", "اللقب",
            "الغيابات\nالمأجورة", "الغيابات\nالغير المأجورة",
            "الغيابات\nالمرضية", "المجموع"
        ]
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionsMovable(False)
        header.setSectionsClickable(False)
        header.setStretchLastSection(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)

        for i in range(len(columns)):
            self.table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

        table_layout.addWidget(self.table)

        # Add container to main layout
        self.main_page_layout.addWidget(table_container)

        # Create paginator and connect to table
        self.paginator = tablepaginator(self.table, rows_per_page=10)
        self.main_page_layout.addWidget(self.paginator)

        # Load data
        self.load_data()

    def load_data(self):
        # Clear current rows
        self.table.setRowCount(0)

        # Load data from controller
        absences = self.controller.load_table()  # Expected to be a list of dicts

        for row_data in absences:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            values = [
                row_data.get("annee", ""),
                row_data.get("mois", ""),
                row_data.get("idemploye", ""),
                row_data.get("prenom", ""),
                row_data.get("nom", ""),
                row_data.get("justifiees", 0),
                row_data.get("non_justifiees", 0),
                row_data.get("maladies", 0),
                row_data.get("total", 0)
            ]

            for col, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                self.table.setItem(row_position, col, item)

        # After loading all rows, update paginator
        self.paginator.update_total_rows()
        self.paginator.update_page(1)
        
    def update_data(self):
        self.load_data()
    def create_action_buttons(self):
        """
        Create action buttons for edit, delete, and history
        All buttons have the same size
        """
        # Buttons container
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 20, 0, 0)  # Increased top margin
        buttons_layout.setSpacing(25)  # Increased spacing

        # Create buttons
        add_btn = QPushButton("إضافة")
        details_btn = QPushButton("تفاصيل")
        history_btn = QPushButton("سجل الأنشطة")

        # Style buttons - all with the same fixed size
        for btn in [add_btn, details_btn, history_btn]:
            btn.setFixedSize(140, 40)  # Same fixed size for all buttons
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
            """)

        # Connect buttons to actions

        details_btn.clicked.connect(self.previous_years_page)
        history_btn.clicked.connect(self.show_history)  # Show history dialog
        add_btn.clicked.connect(self.create_add_window)

        # Add buttons to layout - right to left for Arabic
        buttons_layout.addWidget(details_btn)
        buttons_layout.addWidget(add_btn)
        buttons_layout.addWidget(history_btn)

        self.main_page_layout.addWidget(buttons_widget)
        
        
        
    

    def previous_years_page(self):

        self.previous_years_page = ConsultationPage(self)
        self.previous_years_page.closed.connect(self.load_data)
        self.previous_years_page.show()

    def create_add_window(self):
        self.add_window = AddForm(self)
        self.add_window.accepted.connect(self.load_data)  # Load data after successful add
        self.add_window.show()
    def show_history(self):
        """Show the activity history using the  history dialog with shared session"""
        try:
            from Views.Historique import HistoryDialog
            
            print(f"DEBUG - Ouverture historique avec user_data: {self.current_user_data}")
            
            self.history_dialog = HistoryDialog(
                parent=self,
                current_user_data=self.current_user_data,
                session=self.session,  # AJOUT: Passer la session partagée
                module_name="إدارة الغيابات",  # Arabic for "Absence Management"
                gestion_filter= "إدارة الغيابات"  # Arabic for "Absence Management"
            )
            self.history_dialog.show()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "خطأ",
                f"حدث خطأ غير متوقع:\n{str(e)}"
            )
            print(f"Erreur dans show_history: {e}")   
    def _get_table_data_as_lists(self):
        """Helper function to extract data ONLY from VISIBLE rows in self.table."""
        if not hasattr(self, 'table') or not self.table or self.table.rowCount() == 0:
            # Gérer le cas où la table n'existe pas encore ou est vide
            # Cela peut arriver si la méthode est appelée avant que create_table() ne soit exécutée
            # ou si load_data() n'a pas encore été appelée ou n'a retourné aucune donnée.
            print(f"Avertissement dans _get_table_data_as_lists pour {self.__class__.__name__}: self.table non disponible ou vide.")
            return None, None

        headers = [self.table.horizontalHeaderItem(i).text()
                   for i in range(self.table.columnCount())]

        visible_row_data = []
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                visible_row_data.append(row_data)
        return headers, visible_row_data

    def export_data_to_excel(self):
        print(f"{self.__class__.__name__}: export_data_to_excel (visible rows only) called")
        # Cette ligne va maintenant fonctionner si _get_table_data_as_lists est définie :
        headers, data = self._get_table_data_as_lists() 

        if not data: # Vérifier si data est None ou vide
            QMessageBox.information(self, "لا بيانات للتصدير", "لا توجد صفوف ظاهرة في الجدول حاليًا للتصدير.")
            return

        # ... (le reste de votre code export_data_to_excel) ...

    def print_data_to_pdf(self):
        print(f"{self.__class__.__name__}: print_data_to_pdf (visible rows only) called")
        # Cette ligne va maintenant fonctionner si _get_table_data_as_lists est définie :
        headers, data = self._get_table_data_as_lists()

        if not data: # Vérifier si data est None ou vide
            QMessageBox.information(self, "لا بيانات للطباعة", "لا توجد صفوف ظاهرة في الجدول حاليًا للطباعة.")
            return
    def export_data_to_excel(self):
        print(f"{self.__class__.__name__}: export_data_to_excel (visible rows only) called")
        headers, data = self._get_table_data_as_lists()

        if not data:
            QMessageBox.information(self, "لا بيانات للتصدير", "لا توجد صفوف ظاهرة في الجدول حاليًا للتصدير.")
            return

        filePath, _ = QFileDialog.getSaveFileName(self, "تصدير إلى Excel (الصفحة الحالية)", 
                                                  os.path.expanduser("~/Documents/Absences_Page_Actuelle.xlsx"),
                                                  "Excel Workbook (*.xlsx);;All Files (*)")
        if not filePath:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "الغيابات - الصفحة الحالية" # Adapter
            sheet.sheet_view.rightToLeft = True
            sheet.append(headers)

            for row_values in data:
                sheet.append(row_values)

            for col_idx, column_cells in enumerate(sheet.columns):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = length + 5

            workbook.save(filePath)
            QMessageBox.information(self, "نجاح التصدير", f"تم تصدير الصفوف الظاهرة بنجاح إلى:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ أثناء تصدير الملف: {e}")
            print(f"Error exporting Absences to Excel: {e}")

    def print_data_to_pdf(self):
        print(f"{self.__class__.__name__}: print_data_to_pdf (visible rows only) called")
        headers, data = self._get_table_data_as_lists()

        if not data:
            QMessageBox.information(self, "لا بيانات للطباعة", "لا توجد صفوف ظاهرة في الجدول حاليًا للطباعة.")
            return

        filePath, _ = QFileDialog.getSaveFileName(self, "طباعة إلى PDF (الصفحة الحالية)", 
                                                  os.path.expanduser("~/Documents/Absences_Page_Actuelle.pdf"),
                                                  "PDF Document (*.pdf);;All Files (*)")
        if not filePath:
            return

        html_content = "<html><head><meta charset='UTF-8'>"
        html_content += """
        <style>
            @font-face { font-family: 'DejaVu Sans'; src: url('fonts/DejaVuSans.ttf'); } /* Ajustez si police non trouvée */
            body { direction: rtl; font-family: 'Arial', 'DejaVu Sans', sans-serif; font-size: 9pt; }
            table { width: 100%; border-collapse: collapse; margin-top: 10px; }
            th, td { border: 1px solid #333; padding: 4px; text-align: right; word-wrap: break-word; }
            th { background-color: #f0f0f0; font-weight: bold; }
            caption { font-size: 1.1em; font-weight: bold; margin-bottom: 8px; text-align: center; }
        </style>
        </head><body>
        """
        html_content += "<table><caption>جدول الغيابات (الصفحة الحالية)</caption><thead><tr>" # Adapter
        for header in headers:
            html_content += f"<th>{header}</th>"
        html_content += "</tr></thead><tbody>"

        for row_values in data:
            html_content += "<tr>"
            for cell_value in row_values:
                html_content += f"<td>{html.escape(str(cell_value))}</td>"
            html_content += "</tr>"
        html_content += "</tbody></table></body></html>"

        try:
            css_style = CSS(string='@page { size: A4 landscape; margin: 1cm; }') # Landscape pour plus de colonnes
            HTML(string=html_content).write_pdf(filePath, stylesheets=[css_style])
            QMessageBox.information(self, "نجاح الطباعة", f"تم إنشاء ملف PDF للصفوف الظاهرة بنجاح:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في الطباعة", f"حدث خطأ أثناء إنشاء ملف PDF: {e}\n"
                                 "تأكد أن WeasyPrint et ses dépendances (Pango, Cairo) sont correctement installés.")
            print(f"Error printing Absences to PDF: {e}")

# ... (votre classe ActionBar reste la même) ...
class ActionBar(QWidget):
    def __init__(self, parent, table):

        super().__init__(parent)
        self.table = table
        self.setFixedHeight(60)

        # Layout
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 15)

        # Filter section container
        filter_container = QWidget()
        filter_container_layout = QHBoxLayout(filter_container)
        filter_container_layout.setContentsMargins(0, 0, 0, 0)
        filter_container_layout.setSpacing(10)

        # Filter button with icon and text
        self.filter_btn = QPushButton("ترشيح")
        self.filter_btn.setIcon(QIcon("pics/filter.png"))  # Replace with a valid icon path
        self.filter_btn.setIconSize(QSize(20, 20))
        self.filter_btn.setFixedHeight(35)
        self.filter_btn.setToolTip("ترشيح")
        self.filter_btn.setStyleSheet("""
                   QPushButton {
                       background-color: #37474f;
                       border: none;
                       border-radius: 5px;
                       color: #ffffff;
                       font-weight: bold;
                       font-size: 16px;
                       padding: 0 10px;
                   }
                   QPushButton:hover {
                       background-color: #455a64;
                   }
               """)
        self.filter_btn.clicked.connect(self.show_filter_dialog)

        filter_container_layout.addWidget(self.filter_btn)

        layout.addWidget(filter_container)
        layout.addStretch()
        self.setLayout(layout)

        self.filter_checkboxes = {}  # Store column checkboxes


    def show_filter_dialog(self):
        """
        Show the filter dialog with checkboxes for column selection
        """
        self.filter_dialog = QDialog(self)
        self.filter_dialog.setWindowTitle("ترشيح")
        self.filter_dialog.setFixedWidth(350)
        self.filter_dialog.setStyleSheet("background-color: #37474f; color: #ffffff;")

        dialog_layout = QVBoxLayout(self.filter_dialog)
        dialog_layout.setContentsMargins(20, 20, 20, 20)

        title_label = QLabel("اختر الأعمدة للترشيح:")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 15px;")
        dialog_layout.addWidget(title_label)

        columns = [
            "السنة", "الشهر", "رقم الموظف", "الإسم", "اللقب",
            "الغيابات المأجورة", "الغيابات الغير المأجورة",
            "الغيابات المرضية", "المجموع"
        ]

        for column in columns:
            checkbox = QCheckBox(column)
            checkbox.setStyleSheet("""
                QCheckBox {
                    font-size: 16px;
                    padding: 8px 0;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
                QCheckBox::indicator:checked {
                    background-color: #ff6a0e;
                    border: 2px solid #ffffff;
                }
            """)
            dialog_layout.addWidget(checkbox)
            self.filter_checkboxes[column] = checkbox

        filter_value_layout = QHBoxLayout()
        filter_value_layout.setSpacing(10)

        filter_value_label = QLabel("قيمة الترشيح:")
        filter_value_label.setStyleSheet("font-size: 16px;")

        self.filter_value_input = QLineEdit()
        self.filter_value_input.setPlaceholderText("أدخل قيمة للترشيح...")
        self.filter_value_input.setStyleSheet("""
            QLineEdit {
                background-color: #ffffff;
                border: none;
                border-radius: 5px;
                padding: 10px;
                color: #000000;
                font-size: 16px;
                text-align: right;
            }
        """)
        self.filter_value_input.setLayoutDirection(Qt.RightToLeft)

        filter_value_layout.addWidget(filter_value_label)
        filter_value_layout.addWidget(self.filter_value_input)
        dialog_layout.addSpacing(15)
        dialog_layout.addLayout(filter_value_layout)

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        apply_btn = QPushButton("تطبيق")
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff6a0e;
                color: #ffffff;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-size: 16px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #e05d00;
            }
        """)
        apply_btn.clicked.connect(self.apply_filter)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                color: #ffffff;
                border: 1px solid #455a64;
                border-radius: 5px;
                padding: 10px;
                font-size: 16px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        cancel_btn.clicked.connect(self.filter_dialog.reject)

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(apply_btn)
        dialog_layout.addSpacing(15)
        dialog_layout.addLayout(buttons_layout)

        self.filter_dialog.exec_()  # Ensure modal behavior

    def apply_filter(self):
        """
        Apply the filter to the table based on selected columns and input value
        """
        if not self.table:
            print("Error: Table reference is missing.")
            return

        selected_filter_columns = [
            column for column, checkbox in self.filter_checkboxes.items() if checkbox.isChecked()
        ]

        filter_value = self.filter_value_input.text().strip().lower()

        if not filter_value or not selected_filter_columns:
            # Reset filter if no value or no columns selected
            for row in range(self.table.rowCount()):
                self.table.setRowHidden(row, False)
            self.filter_dialog.accept()
            return

        column_indices = [
            col_idx for col_idx in range(self.table.columnCount())
            if self.table.horizontalHeaderItem(col_idx).text() in selected_filter_columns
        ]

        for row in range(self.table.rowCount()):
            row_visible = any(
                filter_value in self.table.item(row, col_idx).text().lower()
                for col_idx in column_indices
            )
            self.table.setRowHidden(row, not row_visible)

        self.filter_dialog.accept()


class ConsultationPage(QMainWindow):
    closed = pyqtSignal()
    def __init__(self, parent=None, stacked_widget=None, main_page=None, edit_callback=None):
        super().__init__(parent)
        self.stacked_widget = stacked_widget
        self.main_page = main_page
        self.edit_callback = edit_callback

        self.setWindowTitle("تفاصيل غيابات السنوات السابقة")
        self.setGeometry(150, 150, 1000, 600)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")

        self.db_session = db.get_session()
        current_user_account_number = None
        if hasattr(parent, 'current_user_data') and parent.current_user_data:
            current_user_account_number = parent.current_user_data.get('account_number')
        
        self.controller = AbsenceController(self.db_session,current_user_account_number)



        self.setWindowFlags(
            Qt.Window |
            Qt.WindowCloseButtonHint |
            Qt.WindowMaximizeButtonHint |
            Qt.WindowMinimizeButtonHint
        )

        # Central widget and layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.layout = QVBoxLayout(central_widget)  # Uniform layout usage
        self.layout.setContentsMargins(20, 20, 20, 20)

        # Add top bar using the shared create_top_bar function
        self.top_bar_widget, self.export_button_topbar, self.print_button_topbar = create_top_bar(
            self, 
            self.layout, 
            None, # Pas de sidebar_toggle pour cette fenêtre
            excel_export_action_callback=self.export_data_to_excel, # NOUVEAU
            pdf_print_action_callback=self.print_data_to_pdf       # NOUVEAU
        )

        # Create components
        self.create_action_bar()
        self.create_header()
        self.create_table()
        self.create_paginator()
        self.create_edit_button()

    def create_action_bar(self):
        action_bar = QWidget()
        action_layout = QHBoxLayout(action_bar)
        action_layout.setContentsMargins(0, 0, 0, 15)

        # Filter container widget and layout
        filter_container = QWidget()
        filter_container_layout = QHBoxLayout(filter_container)
        filter_container_layout.setContentsMargins(0, 0, 0, 0)
        filter_container_layout.setSpacing(10)

        # Single filter button with icon and text
        self.filter_btn = QPushButton("ترشيح")
        self.filter_btn.setIcon(QIcon("pics/filter.png"))  # Use your icon path here
        self.filter_btn.setIconSize(QSize(20, 20))
        self.filter_btn.setFixedHeight(35)
        self.filter_btn.setToolTip("ترشيح")
        self.filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 5px;
                color: #ffffff;
                font-weight: bold;
                font-size: 16px;
                padding: 0 10px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)

        # Connect signal
        self.filter_btn.clicked.connect(self.show_filter_dialog)

        # Add the button to filter container layout
        filter_container_layout.addWidget(self.filter_btn)

        # Add filter container to action layout
        action_layout.addWidget(filter_container)

        # Add stretch to push content to the left if needed
        action_layout.addStretch()

        # Add the action bar to main layout (assuming self.layout exists)
        self.layout.addWidget(action_bar)

    def create_header(self):
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)

        title_label = QLabel("تفاصيل غيابات السنوات السابقة")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: white;")

        refresh_btn = QPushButton("⟳")
        refresh_btn.setFixedSize(40, 40)
        refresh_btn.setToolTip("تحديث")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 17px;
                color: #ffffff;
                font-weight: bold;
                font-size: 20px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        refresh_btn.clicked.connect(self.load_table)  # or whatever refresh method you want

        header_layout.addWidget(title_label)
        header_layout.addStretch()
        header_layout.addWidget(refresh_btn)

        self.layout.addWidget(header_widget)



    def create_table(self):
        # === Table ===
        self.table = QTableWidget()
        self.table.setStyleSheet(f"""
               QTableWidget {{
                   background-color: {MEDIUM_BG};
                   color: {WHITE};
                   border: none;
                   gridline-color: {LIGHT_BG};
                   font-size: 16px;
               }}
               QTableWidget::item {{
                   padding: 8px;
                   border-bottom: 1px solid {LIGHT_BG};
               }}
               QTableWidget::item:selected {{
                   background-color: {ORANGE};
                   color: {WHITE};
               }}
               QHeaderView::section {{
                   background-color: {DARKER_BG};
                   color: {WHITE};
                   padding: 8px;
                   border: none;
                   font-weight: bold;
                   font-size: 16px;
               }}
               QScrollBar:vertical, QScrollBar:horizontal {{
                   width: opx;
                   margin: 0px;
               }}
              
           """)

        columns = [
            "رقم الغياب", "السنة", "الشهر", "رقم الموظف", "اللقب", "الاسم",
            "نوع الغياب", "رقم القرار", "تاريخ\nالقرار", "تاريخ\nالبداية", "تاريخ\nالنهاية", "سبب الغياب"
        ]
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionsMovable(False)
        header.setStretchLastSection(False)
        header.setSectionsClickable(False)

        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)

        for i in range(len(columns)):
            self.table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

        self.load_table()

        self.layout.addWidget(self.table)

    def load_table(self):
        self.table.setRowCount(0)
        table_data = self.controller.load_absences_with_employee_names()
        if not table_data:
            return
        self.table.setRowCount(len(table_data))

        months_ar = [
            "", "جانفي", "فيفري", "مارس", "أفريل", "ماي", "جوان",
            "جويلية", "أوت", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"
        ]

        for row_idx, row_data in enumerate(table_data):
            def safe_text(value):
                return str(value) if value is not None else ""

            date_debut = safe_text(row_data.get("DateDebut", ""))
            year = date_debut[:4] if len(date_debut) >= 4 else ""
            month_number = int(date_debut[5:7]) if len(date_debut) >= 7 else 0
            month_name = months_ar[month_number] if 1 <= month_number <= 12 else ""

            row_values = [
                safe_text(row_data.get("idAbsence")),
                year,
                month_name,
                safe_text(row_data.get("idemploye")),
                safe_text(row_data.get("prenom")),
                safe_text(row_data.get("nom")), 
                safe_text(row_data.get("Type")),
                safe_text(row_data.get("NumeroDecision")),
                safe_text(row_data.get("DateDecision")),
                date_debut,
                safe_text(row_data.get("DateFin")),
                safe_text(row_data.get("Raison") or row_data.get("Raison2")),
            ]

            for col_idx, value in enumerate(row_values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, col_idx, item)

    def create_paginator(self):
        # Add paginator below the table
        self.paginator = tablepaginator(self.table, rows_per_page=10)
        self.layout.addWidget(self.paginator)
        self.paginator.connect_to_table(self.table)

    def create_edit_button(self):
        close_btn = QPushButton("إغلاق")
        edit_btn = QPushButton("تعديل")
        delete_btn = QPushButton("حذف")

        for btn in (edit_btn, close_btn, delete_btn):
            btn.setFixedSize(160, 45)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
            """)

        buttons_layout = QHBoxLayout()

        buttons_layout.addWidget(edit_btn)
        buttons_layout.addStretch(1)  # space between edit and delete
        buttons_layout.addWidget(delete_btn)
        buttons_layout.addStretch(1)  # space between delete and close
        buttons_layout.addWidget(close_btn)

        self.layout.addLayout(buttons_layout)

        # Connect button signals
        edit_btn.clicked.connect(self.on_edit_button_clicked)
        delete_btn.clicked.connect(self.on_delete_button_clicked)
        close_btn.clicked.connect(self.close)

    def closeEvent(self, event):
        self.closed.emit()
        event.accept()  #

    def on_edit_button_clicked(self):
        indexes = self.table.selectionModel().selectedRows()

        if len(indexes) == 0:
            dialog = StyledMessageDialog(self, title="تحذير", message="يرجى تحديد صف للتعديل.", message_type="warning")
            dialog.exec_()
            return

        selected_row = indexes[0].row()

        def get_cell_text(row, col):
            item = self.table.item(row, col)
            if item:
                return item.text()
            widget = self.table.cellWidget(row, col)
            if widget and hasattr(widget, 'text'):
                return widget.text()
            return ""

        data = {
            "الإسم": get_cell_text(selected_row, 5),  # Column with first name
            "اللقب": get_cell_text(selected_row, 4),  # Column with last name
            "نوع الغياب": get_cell_text(selected_row, 6),  # Absence type combo
            "رقم القرار": get_cell_text(selected_row, 7),  # Decision number
            "تاريخ القرار": get_cell_text(selected_row, 8),  # Decision date
            "تاريخ البداية": get_cell_text(selected_row, 9),  # Start date
            "تاريخ النهاية": get_cell_text(selected_row, 10),  # End date
            "سبب الغياب": get_cell_text(selected_row, 11),  # Absence reason
            "إضافة سبب آخر للغياب": get_cell_text(selected_row, 12),  # Additional reason
        }

        self.edit_window = EditForm(self, data, selected_row, self.table)
        if self.edit_window.exec_() == QDialog.Accepted:
            self.load_table()

    def on_delete_button_clicked(self):
        indexes = self.table.selectionModel().selectedRows()

        if not indexes:
            dialog = QDialog(self)
            dialog.setWindowTitle("  تحذير")
            dialog.setFixedSize(450, 180)
            dialog.setStyleSheet(f"background-color: {DARKER_BG};")

            layout = QVBoxLayout(dialog)
            layout.setContentsMargins(25, 25, 25, 25)
            layout.setSpacing(20)

            title_label = QLabel(" تحذير")
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setStyleSheet(f"color: {YELLOW}; font-size: 18px; font-weight: bold;")

            message_label = QLabel("يرجى تحديد صف للحذف.")
            message_label.setAlignment(Qt.AlignCenter)
            message_label.setStyleSheet(f"color: {WHITE}; font-size: 16px;")

            layout.addWidget(title_label)
            layout.addWidget(message_label)

            dialog.exec_()
            return

        selected_row = indexes[0].row()
        absence_id = self.table.item(selected_row, 0).text()

        confirm_dialog = QDialog(self)
        confirm_dialog.setWindowTitle("تأكيد الحذف")
        confirm_dialog.setFixedSize(450, 220)
        confirm_dialog.setStyleSheet(f"background-color: {DARKER_BG};")

        layout = QVBoxLayout(confirm_dialog)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(30)

        title_label = QLabel("تأكيد الحذف")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"color: {YELLOW}; font-size: 18px; font-weight: bold;")

        message_label = QLabel("هل أنت متأكد من الحذف؟")
        message_label.setAlignment(Qt.AlignCenter)
        message_label.setStyleSheet(f"color: {WHITE}; font-size: 16px;")

        layout.addWidget(title_label)
        layout.addWidget(message_label)

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(20)

        confirm_btn = QPushButton("تأكيد")
        confirm_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                font-size: 16px;
                font-weight: bold;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #3d8b40;
            }}
        """)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                font-size: 16px;
                font-weight: bold;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)

        confirm_btn.clicked.connect(lambda: self.confirm_delete_and_show_message(confirm_dialog, selected_row))

        cancel_btn.clicked.connect(confirm_dialog.reject)

        buttons_layout.addWidget(confirm_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

        confirm_dialog.exec_()

    def confirm_delete_and_show_message(self, dialog, row_index):
        item = self.table.item(row_index, 0)
        if item is None:
            print("No item found at the selected row.")
            return
        print("The item has been found")

        consultation_id_text = item.text()
        if not consultation_id_text.isdigit():
            print(f"Invalid consultation ID: {consultation_id_text}")
            return

        consultation_id = int(consultation_id_text)

        print("I am about to call the controller")
        try:
            # Delete from DB via controller
            self.controller.delete_absence(absence_id=consultation_id)
        except ValueError as e:
            # Show error dialog if absence not found or other error occurs
            error_dialog = QDialog(self)
            error_dialog.setWindowTitle("خطأ في الحذف")
            error_dialog.setFixedSize(450, 180)
            error_dialog.setStyleSheet(f"background-color: {DARKER_BG};")

            layout = QVBoxLayout(error_dialog)
            layout.setContentsMargins(25, 25, 25, 25)
            layout.setSpacing(20)

            title_label = QLabel("خطأ")
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setStyleSheet(f"color: {RED}; font-size: 18px; font-weight: bold;")

            message_label = QLabel(str(e))
            message_label.setAlignment(Qt.AlignCenter)
            message_label.setStyleSheet(f"color: {WHITE}; font-size: 16px;")

            layout.addWidget(title_label)
            layout.addWidget(message_label)

            error_dialog.exec_()
            return

        # Remove row from table
        self.table.removeRow(row_index)

        # Close dialog
        dialog.accept()

        # Show success message dialog
        success_dialog = QDialog(self)
        success_dialog.setWindowTitle("تم الحذف")
        success_dialog.setFixedSize(450, 180)
        success_dialog.setStyleSheet(f"background-color: {DARKER_BG};")

        layout = QVBoxLayout(success_dialog)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)

        title_label = QLabel("تم الحذف")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"color: {GREEN}; font-size: 18px; font-weight: bold;")

        message_label = QLabel("تم حذف السجل بنجاح.")
        message_label.setAlignment(Qt.AlignCenter)
        message_label.setStyleSheet(f"color: {WHITE}; font-size: 16px;")

        layout.addWidget(title_label)
        layout.addWidget(message_label)

        success_dialog.exec_()

    def show_filter_dialog(self):
        """
        Show the filter dialog with styled checkboxes and a value input for filtering
        """
        self.filter_dialog = QDialog(self)
        self.filter_dialog.setWindowTitle("ترشيح")
        self.filter_dialog.setFixedWidth(350)
        self.filter_dialog.setStyleSheet("background-color: #37474f; color: #ffffff;")

        dialog_layout = QVBoxLayout(self.filter_dialog)
        dialog_layout.setContentsMargins(20, 20, 20, 20)

        title_label = QLabel("اختر الأعمدة للترشيح:")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 15px;")
        dialog_layout.addWidget(title_label)

        self.filter_checkboxes = {}

        # Use actual column headers from the table
        columns = [
            "رقم الغياب", "السنة", "الشهر", "رقم الموظف", "اللقب", "الاسم",
            "نوع الغياب", "رقم القرار", "تاريخ القرار", "تاريخ البداية", "تاريخ النهاية", "سبب الغياب"
        ]

        for column in columns:
            checkbox = QCheckBox(column)
            checkbox.setStyleSheet("""
                QCheckBox {
                    font-size: 24px;
                    padding: 8px 0;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                }
                QCheckBox::indicator:checked {
                    background-color: #ff6a0e;
                    border: 2px solid #ffffff;
                }
            """)
            dialog_layout.addWidget(checkbox)
            self.filter_checkboxes[column] = checkbox

        # Value input
        filter_value_layout = QHBoxLayout()
        filter_value_layout.setSpacing(10)

        filter_value_label = QLabel("قيمة الترشيح:")
        filter_value_label.setStyleSheet("font-size: 16px;")

        self.filter_value_input = QLineEdit()
        self.filter_value_input.setPlaceholderText("أدخل قيمة للترشيح...")
        self.filter_value_input.setStyleSheet("""
            QLineEdit {
                background-color: #ffffff;
                border: none;
                border-radius: 5px;
                padding: 10px;
                color: #000000;
                font-size: 16px;
                text-align: right;
            }
        """)
        self.filter_value_input.setLayoutDirection(Qt.RightToLeft)

        filter_value_layout.addWidget(filter_value_label)
        filter_value_layout.addWidget(self.filter_value_input)
        dialog_layout.addSpacing(15)
        dialog_layout.addLayout(filter_value_layout)

        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        apply_btn = QPushButton("تطبيق")
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff6a0e;
                color: #ffffff;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-size: 16px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #e05d00;
            }
        """)
        apply_btn.clicked.connect(self.apply_filter)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                color: #ffffff;
                border: 1px solid #455a64;
                border-radius: 5px;
                padding: 10px;
                font-size: 16px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        cancel_btn.clicked.connect(self.filter_dialog.reject)

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(apply_btn)
        dialog_layout.addSpacing(15)
        dialog_layout.addLayout(buttons_layout)

        self.filter_dialog.exec_()

    def apply_filter(self):
        """
        Apply the filter to the table based on selected columns and input value
        """
        if not self.table:
            print("Error: Table reference is missing.")
            return

        selected_filter_columns = [
            column for column, checkbox in self.filter_checkboxes.items() if checkbox.isChecked()
        ]

        filter_value = self.filter_value_input.text().strip().lower()

        if not filter_value or not selected_filter_columns:
            # Reset filter if no value or no columns selected
            for row in range(self.table.rowCount()):
                self.table.setRowHidden(row, False)
            self.filter_dialog.accept()
            return

        column_indices = [
            col_idx for col_idx in range(self.table.columnCount())
            if self.table.horizontalHeaderItem(col_idx).text() in selected_filter_columns
        ]

        for row in range(self.table.rowCount()):
            row_visible = any(
                filter_value in self.table.item(row, col_idx).text().lower()
                for col_idx in column_indices
                if self.table.item(row, col_idx)  # Avoid None
            )
            self.table.setRowHidden(row, not row_visible)

        self.filter_dialog.accept()

    def load_table(self):
        # ... (votre méthode load_table existante) ...
        super_return_value = super().load_table() if hasattr(super(), 'load_table') else None # Au cas où
        
         # Mettre à jour l'état des boutons
        return super_return_value

    # --- MÉTHODES D'EXPORTATION ET D'IMPRESSION POUR ConsultationPage ---
    # Elles sont identiques à celles de AbsenceManagementSystem, car elles lisent self.table
    # et utilisent les mêmes bibliothèques.

    def _get_table_data_as_lists(self): # Identique
        """Helper function to extract data ONLY from VISIBLE rows in self.table."""
        if not hasattr(self, 'table') or not self.table or self.table.rowCount() == 0:
            return None, None
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        visible_row_data = []
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                visible_row_data.append(row_data)
        return headers, visible_row_data

    def export_data_to_excel(self): # Identique, mais avec des noms de fichiers/titres différents
        print(f"{self.__class__.__name__}: export_data_to_excel (visible rows only) called")
        headers, data = self._get_table_data_as_lists()
        if not data:
            QMessageBox.information(self, "لا بيانات للتصدير", "لا توجد صفوف ظاهرة في الجدول حاليًا للتصدير.")
            return
        filePath, _ = QFileDialog.getSaveFileName(self, "تصدير إلى Excel (غيابات سابقة)", 
                                                  os.path.expanduser("~/Documents/Absences_Anciennes_Page_Actuelle.xlsx"),
                                                  "Excel Workbook (*.xlsx);;All Files (*)")
        if not filePath: return
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "غيابات سابقة - صفحة حالية" # Adapter
            sheet.sheet_view.rightToLeft = True
            sheet.append(headers)
            for row_values in data: sheet.append(row_values)
            for col_idx, column_cells in enumerate(sheet.columns):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = length + 5
            workbook.save(filePath)
            QMessageBox.information(self, "نجاح التصدير", f"تم تصدير الصفوف الظاهرة بنجاح إلى:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ أثناء تصدير الملف: {e}")
            print(f"Error exporting ConsultationPage to Excel: {e}")

    def print_data_to_pdf(self): # Identique, mais avec des noms de fichiers/titres différents
        print(f"{self.__class__.__name__}: print_data_to_pdf (visible rows only) called")
        headers, data = self._get_table_data_as_lists()
        if not data:
            QMessageBox.information(self, "لا بيانات للطباعة", "لا توجد صفوف ظاهرة في الجدول حاليًا للطباعة.")
            return
        filePath, _ = QFileDialog.getSaveFileName(self, "طباعة إلى PDF (غيابات سابقة)", 
                                                  os.path.expanduser("~/Documents/Absences_Anciennes_Page_Actuelle.pdf"),
                                                  "PDF Document (*.pdf);;All Files (*)")
        if not filePath: return
        html_content = "<html><head><meta charset='UTF-8'>"
        html_content += """
        <style>
            @font-face { font-family: 'DejaVu Sans'; src: url('fonts/DejaVuSans.ttf'); }
            body { direction: rtl; font-family: 'Arial', 'DejaVu Sans', sans-serif; font-size: 8pt; } /* Police plus petite pour plus de données */
            table { width: 100%; border-collapse: collapse; margin-top: 10px; }
            th, td { border: 1px solid #333; padding: 3px; text-align: right; word-wrap: break-word; }
            th { background-color: #f0f0f0; font-weight: bold; }
            caption { font-size: 1.0em; font-weight: bold; margin-bottom: 7px; text-align: center; }
        </style>
        </head><body>
        """
        html_content += "<table><caption>تفاصيل غيابات السنوات السابقة (الصفحة الحالية)</caption><thead><tr>" # Adapter
        for header in headers: html_content += f"<th>{header}</th>"
        html_content += "</tr></thead><tbody>"
        for row_values in data:
            html_content += "<tr>"
            for cell_value in row_values: html_content += f"<td>{html.escape(str(cell_value))}</td>"
            html_content += "</tr>"
        html_content += "</tbody></table></body></html>"
        try:
            css_style = CSS(string='@page { size: A3 landscape; margin: 0.7cm; }') # A3 landscape pour beaucoup de colonnes
            HTML(string=html_content).write_pdf(filePath, stylesheets=[css_style])
            QMessageBox.information(self, "نجاح الطباعة", f"تم إنشاء ملف PDF للصفوف الظاهرة بنجاح:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في الطباعة", f"حدث خطأ أثناء إنشاء ملف PDF: {e}\n"
                                 "Vérifiez l'installation de WeasyPrint et de ses dépendances (Pango, Cairo).")
            print(f"Error printing ConsultationPage to PDF: {e}")




class EditForm(QDialog):
    """
    Dialog for editing existing absence data
    """

    def __init__(self, parent=None, data=None, row_index=None, table=None):
        super().__init__(parent)
        self.setWindowTitle("تعديل البيانات")
        self.setFixedSize(600, 800)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)

        self.db_session = db.get_session()
        current_user_account_number = None
        if hasattr(parent, 'current_user_data') and parent.current_user_data:
            current_user_account_number = parent.current_user_data.get('account_number')
        self.controller = AbsenceController(self.db_session,current_user_account_number)

        self.data = data or {}
        self.row_index = row_index
        self.table = table

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setStyleSheet(f"background-color: {DARK_BG};")

        center_container = QWidget()
        center_layout = QHBoxLayout(center_container)

        form_container_wrapper = QWidget()
        form_container_wrapper.setFixedWidth(500)
        form_wrapper_layout = QVBoxLayout(form_container_wrapper)
        form_wrapper_layout.setContentsMargins(20, 20, 20, 20)
        form_wrapper_layout.setSpacing(15)

        self.title_label = QLabel("تعديل البيانات")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px 0;")

        form_container = QWidget()
        form_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")

        container_layout = QVBoxLayout(form_container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.setSpacing(15)

        # Define fields WITHOUT employee ID
        fields = [
            {"name": "الإسم", "is_date": False},
            {"name": "اللقب", "is_date": False},
            {"name": "نوع الغياب", "is_date": False, "is_combo": True},
            {"name": "رقم القرار", "is_date": False},
            {"name": "تاريخ القرار", "is_date": True},
            {"name": "تاريخ البداية", "is_date": True},
            {"name": "تاريخ النهاية", "is_date": True},
            {"name": "سبب الغياب", "is_date": False},
            {"name": "إضافة سبب آخر للغياب", "is_date": False},
        ]

        absence_types = ["مبرر", "غير مبرر", "إجازة مرضية"]

        self.edit_form_fields = {}
        current_date = QDate.currentDate()

        for field in fields:
            field_layout = QVBoxLayout()
            field_layout.setSpacing(5)

            label = QLabel(field["name"])
            label.setStyleSheet("font-size: 16px; font-weight: bold;")

            if field.get("is_date"):
                input_field = QDateEdit()
                input_field.setDisplayFormat("yyyy-MM-dd")
                input_field.setCalendarPopup(True)
                input_field.setDate(current_date)

                if self.data.get(field["name"]):
                    date_val = QDate.fromString(self.data[field["name"]], "yyyy-MM-dd")
                    if date_val.isValid():
                        input_field.setDate(date_val)

                input_field.setStyleSheet(f"""
                    QDateEdit {{
                        background-color: {DARKER_BG};
                        border: none;
                        border-radius: 5px;
                        padding: 8px;
                        color: {WHITE};
                        font-size: 16px;
                        text-align: right;
                    }}
                """)
            elif field.get("is_combo"):
                input_field = QComboBox()
                input_field.addItems(absence_types)
                current_text = self.data.get(field["name"], absence_types[0])
                if current_text in absence_types:
                    input_field.setCurrentText(current_text)
                else:
                    input_field.setCurrentIndex(0)

                input_field.setStyleSheet(f"""
                    QComboBox {{
                        background-color: {DARKER_BG};
                        border: none;
                        border-radius: 5px;
                        padding: 8px;
                        color: {WHITE};
                        font-size: 16px;
                    }}
                """)
            else:
                input_field = QLineEdit()
                
                input_field.setAlignment(Qt.AlignRight)
                input_field.setLayoutDirection(Qt.RightToLeft)
                input_field.setText(self.data.get(field["name"], ""))
                input_field.setStyleSheet(f"""
                    QLineEdit {{
                        background-color: {DARKER_BG};
                        border: none;
                        border-radius: 5px;
                        padding: 8px;
                        color: {WHITE};
                        font-size: 16px;
                        text-align: right;
                    }}
                """)

                # Make first name and last name fields read-only
                if field["name"] in ["الإسم", "اللقب"]:
                    input_field.setReadOnly(True)

            field_layout.addWidget(label)
            field_layout.addWidget(input_field)
            container_layout.addLayout(field_layout)

            self.edit_form_fields[field["name"]] = input_field

        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        save_btn = QPushButton("حفظ")
        save_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #3d8b40;
            }}
        """)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)

        save_btn.clicked.connect(self.save_edited_row)
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)

        container_layout.addLayout(buttons_layout)

        form_wrapper_layout.addWidget(self.title_label)
        form_wrapper_layout.addWidget(form_container)

        center_layout.addStretch()
        center_layout.addWidget(form_container_wrapper)
        center_layout.addStretch()

        scroll_area.setWidget(center_container)
        main_layout.addWidget(scroll_area)

    def save_edited_row(self):
        try:
            date_decision = self.edit_form_fields["تاريخ القرار"].date()
            date_start = self.edit_form_fields["تاريخ البداية"].date()
            date_end = self.edit_form_fields["تاريخ النهاية"].date()
        except KeyError:
            StyledMessageDialog(self, title="خطأ",
                                message="تأكد من وجود جميع الحقول المطلوبة.",
                                message_type="warning").exec_()
            return

        # Always check date_start <= date_end regardless of type
        if date_start > date_end:
            StyledMessageDialog(self, title="خطأ في التاريخ",
                                message="يجب أن يكون تاريخ البداية قبل تاريخ النهاية.",
                                message_type="warning").exec_()
            return

        try:
            absence_id = self.table.item(self.row_index, 0).text()
            emp_number = self.table.item(self.row_index, 1).text()

            last_name = self.edit_form_fields["اللقب"].text().strip()
            first_name = self.edit_form_fields["الإسم"].text().strip()
            absence_type = self.edit_form_fields["نوع الغياب"].currentText().strip()
            decision_number = self.edit_form_fields["رقم القرار"].text().strip()
            absence_reason = self.edit_form_fields["سبب الغياب"].text().strip()
            absence_reason2 = self.edit_form_fields["إضافة سبب آخر للغياب"].text().strip()

            if absence_type == "مبرر":
                # Validate all fields

                # Validate last and first names (alphabetic only)
                if not last_name.replace(" ", "").isalpha():
                    raise ValueError("اللقب يجب أن يحتوي على حروف فقط.")
                if not first_name.replace(" ", "").isalpha():
                    raise ValueError("الإسم يجب أن يحتوي على حروف فقط.")

                # Validate second reason if provided
                if absence_reason2 and not absence_reason2.replace(" ", "").isalpha():
                    raise ValueError("السبب الإضافي يجب أن يحتوي على حروف فقط.")

                # decision number required and numeric
                if not decision_number:
                    StyledMessageDialog(self, title="تحذير",
                                        message="يجب تعبئة رقم القرار عند اختيار نوع الغياب مبرر.",
                                        message_type="warning").exec_()
                    return
                if not decision_number.isdigit():
                    StyledMessageDialog(self, title="تحذير",
                                        message="رقم القرار يجب أن يحتوي على أرقام فقط.",
                                        message_type="warning").exec_()
                    return

                # decision date must be valid and not in the future
                if not date_decision.isValid():
                    StyledMessageDialog(self, title="تحذير",
                                        message="يجب تعبئة تاريخ القرار عند اختيار نوع الغياب مبرر.",
                                        message_type="warning").exec_()
                    return
                if date_decision > QDate.currentDate():
                    StyledMessageDialog(self, title="تحذير",
                                        message="لا يمكن أن يكون تاريخ القرار بعد اليوم.",
                                        message_type="warning").exec_()
                    return

                # absence reason required and alphabetic
                if not absence_reason:
                    StyledMessageDialog(self, title="تحذير",
                                        message="يجب تعبئة سبب الغياب عند اختيار نوع الغياب مبرر.",
                                        message_type="warning").exec_()
                    return
                if not absence_reason.replace(" ", "").isalpha():
                    StyledMessageDialog(self, title="تحذير",
                                        message="سبب الغياب يجب أن يحتوي على حروف فقط.",
                                        message_type="warning").exec_()
                    return

            elif absence_type in ["غير مبرر", "إجازة مرضية"]:
                # For other types, only names are not validated (optional)
                # Clear these fields as per your request
                decision_number = None
                date_decision = QDate()
                absence_reason = ""
                absence_reason2 = ""

        except Exception as e:
            StyledMessageDialog(self, title="خطأ",
                                message=f"حدث خطأ في قراءة البيانات أو التحقق منها:\n{str(e)}",
                                message_type="warning").exec_()
            return

        # Now call controller update with sanitized data
        status = self.controller.update_absence(
            absence_id=absence_id,
            Type=absence_type,
            DateDebut=date_start.toPyDate(),
            DateFin=date_end.toPyDate(),
            Raison=absence_reason,
            NumeroDecision=decision_number,
            DateDecision=date_decision.toPyDate() if date_decision.isValid() else None,
            Raison2=absence_reason2
        )

        if status == "success":
            # Update table UI
            self.table.setItem(self.row_index, 2, QTableWidgetItem(last_name))
            self.table.setItem(self.row_index, 3, QTableWidgetItem(first_name))
            self.table.setItem(self.row_index, 4, QTableWidgetItem(absence_type))
            self.table.setItem(self.row_index, 5, QTableWidgetItem(decision_number or ""))
            self.table.setItem(self.row_index, 6, QTableWidgetItem(
                date_decision.toString("yyyy-MM-dd") if date_decision.isValid() else ""))
            self.table.setItem(self.row_index, 7, QTableWidgetItem(date_start.toString("yyyy-MM-dd")))
            self.table.setItem(self.row_index, 8, QTableWidgetItem(date_end.toString("yyyy-MM-dd")))
            self.table.setItem(self.row_index, 9, QTableWidgetItem(absence_reason))
            self.table.setItem(self.row_index, 10, QTableWidgetItem(absence_reason2))

            StyledMessageDialog(self, title="نجاح",
                                message="تم تعديل البيانات بنجاح",
                                message_type="info").exec_()
            self.accept()
        else:
            StyledMessageDialog(self, title="فشل",
                                message="لم يتم تعديل البيانات، يرجى المحاولة مرة أخرى.",
                                message_type="warning").exec_()


import re


class AddForm(QDialog):
    """

    """

    def __init__(self, employee_id=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("إضافة غياب")
        self.setFixedSize(600, 800)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")

        # DB and controllers
        self.db_session = db.get_session()
        
        current_user_account_number = None
        if hasattr(parent, 'current_user_data') and parent.current_user_data:
            current_user_account_number = parent.current_user_data.get('account_number')
        self.employee_controller = EmployeeController(self.db_session,current_user_account_number)
        self.controller = AbsenceController(self.db_session,current_user_account_number)
        # Main layout with scroll area
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
       
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setStyleSheet(f"""
                                  background-color: {DARK_BG};
                                  """)

        center_container = QWidget()
        center_layout = QHBoxLayout(center_container)

        form_container_wrapper = QWidget()
        form_container_wrapper.setFixedWidth(500)
        form_wrapper_layout = QVBoxLayout(form_container_wrapper)
        form_wrapper_layout.setContentsMargins(20, 20, 20, 20)
        form_wrapper_layout.setSpacing(15)

        # Title
        self.title_label = QLabel("إضافة غياب")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin: 10px 0;")

        # Form container
        form_container = QWidget()
        form_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")

        container_layout = QVBoxLayout(form_container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.setSpacing(15)

        # Fields definitions
        fields = [
            {"name": "الإسم",  "is_date": False, "widget": "combo"},
            {"name": "اللقب",  "is_date": False, "widget": "combo"},
            {"name": "نوع الغياب",  "is_date": False, "widget": "combo",
             "options": ["مبرر", "غير مبرر", "إجازة مرضية"]},
            {"name": "رقم القرار", "is_date": False},
            {"name": "تاريخ القرار",  "is_date": True},
            {"name": "تاريخ البداية",  "is_date": True},
            {"name": "تاريخ النهاية",  "is_date": True},
            {"name": "سبب الغياب",  "is_date": False},
            {"name": "إضافة سبب آخر للغياب",  "is_date": False},
        ]

        # Load employees for name/lastname combos
        try:
            self.employee_data = self.employee_controller.get_all_employees()  # list of (first, last)

            print("Employees list:", self.employee_data)

            self.employee_map = {first: last for first, last in self.employee_data}
            # Reverse map for last name -> first name
            self.reverse_employee_map = {}
            for first, last in self.employee_data:
                if last not in self.reverse_employee_map:
                    self.reverse_employee_map[last] = first

            employee_names = list(self.employee_map.keys())
            employee_lastnames = list(set(last for _, last in self.employee_data))
        except Exception as e:
            print("Error loading employees:", e)
            employee_names = []
            employee_lastnames = []
            self.employee_map = {}
            self.reverse_employee_map = {}

        self.add_form_fields = {}
        current_date = QDate.currentDate()

        # Build the form fields
        for field in fields:
            layout = QVBoxLayout()
            layout.setSpacing(5)

            label = QLabel(field["name"])
            label.setStyleSheet("font-size: 16px; font-weight: bold;")

            if field["name"] == "الإسم":
                input_field = QComboBox()
                input_field.setEditable(True)
                input_field.addItems(employee_names)
                self.name_combo = input_field
            elif field["name"] == "اللقب":
                input_field = QComboBox()
                input_field.setEditable(True)
                input_field.addItems(employee_lastnames)
                self.lastname_combo = input_field
            elif field.get("widget") == "combo":
                input_field = QComboBox()
                input_field.addItems(field.get("options", []))
                input_field.setEditable(False)
            elif field["is_date"]:
                input_field = QDateEdit()
                input_field.setDisplayFormat("yyyy-MM-dd")
                input_field.setDate(current_date)
                input_field.setCalendarPopup(True)
            else:
                input_field = QLineEdit()
                
                input_field.setAlignment(Qt.AlignRight)
                input_field.setLayoutDirection(Qt.RightToLeft)

            input_field.setStyleSheet(f"""
                {type(input_field).__name__} {{
                    background-color: {DARKER_BG};
                    border: none;
                    border-radius: 5px;
                    padding: 8px;
                    color: {WHITE};
                    font-size: 16px;
                }}
            """)

            layout.addWidget(label)
            layout.addWidget(input_field)
            container_layout.addLayout(layout)

            self.add_form_fields[field["name"]] = input_field

        # Flags to avoid infinite update loops in name/lastname sync
        self.updating_from_name = False
        self.updating_from_lastname = False

        # Connect signals for syncing name and lastname
        self.name_combo.currentTextChanged.connect(self.update_lastname_field)
        self.lastname_combo.currentTextChanged.connect(self.update_name_field)


        # Buttons layout
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)

        save_btn = QPushButton("حفظ")
        save_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
        """)
        save_btn.clicked.connect(self.save_new_entry)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
        """)
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)
        container_layout.addLayout(buttons_layout)

        form_wrapper_layout.addWidget(self.title_label)
        form_wrapper_layout.addWidget(form_container)

        center_layout.addStretch()
        center_layout.addWidget(form_container_wrapper)
        center_layout.addStretch()

        scroll_area.setWidget(center_container)
        main_layout.addWidget(scroll_area)
        

    def update_lastname_field(self, selected_name):
        if self.updating_from_lastname:
            return
        last_name = self.employee_map.get(selected_name)
        if last_name:
            self.updating_from_name = True
            idx = self.lastname_combo.findText(last_name)
            if idx != -1:
                self.lastname_combo.setCurrentIndex(idx)
            else:
                self.lastname_combo.setEditText(last_name)
            self.updating_from_name = False

    def update_name_field(self, selected_lastname):
        if self.updating_from_name:
            return
        first_name = self.reverse_employee_map.get(selected_lastname)
        if first_name:
            self.updating_from_lastname = True
            idx = self.name_combo.findText(first_name)
            if idx != -1:
                self.name_combo.setCurrentIndex(idx)
            else:
                self.name_combo.setEditText(first_name)
            self.updating_from_lastname = False

    def save_new_entry(self):
        def show_error(msg):
            dialog = StyledMessageDialog(self, title="خطأ", message=msg, message_type="warning")
            dialog.exec_()

        # Extract fields and strip whitespace
        name = self.add_form_fields["الإسم"].currentText().strip()
        lastname = self.add_form_fields["اللقب"].currentText().strip()
        print(f"--- Tentative de recherche ---")
        print(f"Prénom récupéré du formulaire: '{name}'")
        print(f"Nom récupéré du formulaire: '{lastname}'")
        employee_id = self.employee_controller.getidbynameandlastname(name=name, lastname=lastname)
        print(f"ID retourné par getidbynameandlastname: {employee_id}")
        # 3. Vérifier si l'employé a été trouvé
        if employee_id is None:
            show_error("Employé non trouvé. Veuillez vérifier que le nom et le prénom sont corrects et existent.")
            return
        absence_type = self.add_form_fields["نوع الغياب"].currentText().strip()
        decision_number_text = self.add_form_fields["رقم القرار"].text().strip()
        reason1 = self.add_form_fields["سبب الغياب"].text().strip()
        reason2 = self.add_form_fields["إضافة سبب آخر للغياب"].text().strip() or None

        # Dates as QDate objects
        decision_date = self.add_form_fields["تاريخ القرار"].date()
        start_date = self.add_form_fields["تاريخ البداية"].date()
        end_date = self.add_form_fields["تاريخ النهاية"].date()


        # Arabic and English letters and spaces validation pattern
        name_pattern = r'^[\u0600-\u06FFa-zA-Z\s]+$'

        # Validate name and lastname
        if not (name and re.match(name_pattern, name)):
            show_error("الاسم يجب أن يحتوي فقط على حروف عربية ، ولا يجب أن يكون فارغًا.")
            return
        if not (lastname and re.match(name_pattern, lastname)):
            show_error("اللقب يجب أن يحتوي فقط على حروف عربية ، ولا يجب أن يكون فارغًا.")
            return

        # Validate mandatory absence type
        if not absence_type:
            show_error("يرجى اختيار نوع الغياب.")
            return

        # Additional validations if absence type is "مبرر"
        if absence_type == "مبرر":
            if not reason1:
                show_error("يرجى إدخال سبب الغياب.")
                return
            if not decision_number_text:
                show_error("يرجى إدخال رقم القرار.")
                return
            if not decision_number_text.isdigit():
                show_error("رقم القرار يجب أن يكون رقماً صحيحاً.")
                return
            decision_number = decision_number_text
            if decision_date > QDate.currentDate():
                show_error("لا يمكن أن يكون تاريخ القرار بعد اليوم.")
                return
        else:
            # Clear fields not needed for non-justified absences
            decision_number = None
            decision_date = None
            reason1 = ""

        # Validate start and end dates
        current_date = QDate.currentDate()
        if start_date > current_date:
            show_error("لا يمكن أن يكون تاريخ البداية بعد اليوم.")
            return
        if end_date < start_date:
            show_error("لا يمكن أن يكون تاريخ النهاية قبل تاريخ البداية.")
            return

        # Convert QDate to Python date
        date_decision = decision_date.toPyDate() if decision_date else None
        date_start = start_date.toPyDate()
        date_end = end_date.toPyDate()

        # Call controller and handle statuses
        try:
            controller = AbsenceController(self.db_session)
            status = controller.save_absence_for_employee(
                Type=absence_type,
                DateDebut=date_start,
                DateFin=date_end,
                Raison=reason1,
                NumeroDecision=decision_number,
                DateDecision=date_decision,
                Raison2=reason2,
                idemploye=employee_id,
                name=name,
                lastname=lastname
            )

            if status == "overlapping_absence":
                show_error("لدى الموظف غياب مسجل لهذه الفترة.")
                return
            elif status != "success":
                show_error("يوجد عدم تطابق في اسم الموظف ولقبه")
                return

        except Exception as e:
            dialog = StyledMessageDialog(
                self,
                title="خطأ",
                message=f"حدث خطأ أثناء حفظ البيانات: {str(e)}",
                message_type="error"
            )
            dialog.exec_()
            return

        # Success dialog and close form
        dialog = StyledMessageDialog(
            self,
            title="تم الحفظ",
            message="تم حفظ بيانات الغياب بنجاح.",
            message_type="information"
        )
        dialog.exec_()
        self.accept()
        
    
if __name__ == '__main__':
    app = QApplication(sys.argv)

    # Set application-wide font for Arabic support
    font = QFont("Arial", 10)
    app.setFont(font)

    # Set right-to-left layout for Arabic
    app.setLayoutDirection(Qt.RightToLeft)

    window = AbsenceManagementSystem()
    window.show()

    sys.exit(app.exec_())