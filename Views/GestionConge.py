import sys
import os
from datetime import datetime, date
from Models import init_db
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QPushButton, QTableWidget, 
                            QTableWidgetItem, QVBoxLayout, QHBoxLayout, QLineEdit, QLabel, 
                            QFrame, QDialog, QMessageBox, QCheckBox, QComboBox, QDateEdit,
                            QScrollArea, QSizePolicy, QStackedWidget, QToolButton, QMenu,
                            QListWidget, QListWidgetItem, QGridLayout, QSpinBox, QHeaderView,QFileDialog)
from PyQt5.QtGui import QIcon, QPixmap, QFont, QColor, QPalette, QBrush
from PyQt5.QtCore import Qt, QSize, QDate, QTimer, QLocale
from PyQt5.QtWidgets import QAbstractItemView

from TablePaginator1 import tablepaginator
# Import from external files
from ui_constants import *  # Import color themes
from sidebar import create_sidebar
from topbar import create_top_bar  # Import sidebar and topbar functions
from custom_dialogs import CustomWarningDialog, CustomInfoDialog, CustomMessageBox  # Import custom dialogs
from Controllers.conge_controller import CongeController
from Controllers.tranche_controller import TrancheController
from Controllers.EmployeController import EmployeeController
from Controllers.BaseController import BaseControllerWithHistory
from Models.Tranche import Tranche
from DatabaseConnection import db
import openpyxl
from weasyprint import HTML, CSS
import html

class AddLeaveDialog(QDialog):
    """
    Dialog for adding a new leave
    """
    def __init__(self, parent=None, employee_id="", employee_name="",session=None):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("إضافة عطلة")
        self.setMinimumSize(500, 400)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")
        self.session = session 
        # Set window flags to allow maximize/minimize
        self.setWindowFlags(
            Qt.Window |              # Regular window
            Qt.WindowCloseButtonHint 
        )
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create a centered container for the form
        center_container = QWidget()
        center_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        center_layout = QHBoxLayout(center_container)
        
        # Form container with fixed width
        form_container_wrapper = QWidget()
        form_container_wrapper.setFixedWidth(500)
        form_wrapper_layout = QVBoxLayout(form_container_wrapper)
        form_wrapper_layout.setContentsMargins(20, 20, 20, 20)
        form_wrapper_layout.setSpacing(15)
        
        # Title
        self.title_label = QLabel(f"إضافة عطلة - {employee_name} ({employee_id})")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 20px; font-weight: bold; margin: 10px 0;")
        
        # Form container
        form_container = QWidget()
        form_container.setStyleSheet(f"background-color: {LIGHT_BG}; border-radius: 10px;")
        
        container_layout = QVBoxLayout(form_container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.setSpacing(15)
        
        # Form fields
        fields = [
            {"name": "رقم القرار", "type": "text"},
            {"name": "تاريخ القرار", "type": "date"},
            {"name": "تاريخ البداية", "type": "date"},
            {"name": "تاريخ النهاية", "type": "date"}
        ]
        
        self.form_fields = {}
        
        for field in fields:
            field_layout = QVBoxLayout()
            field_layout.setSpacing(5)
            
            label = QLabel(field["name"])
            label.setStyleSheet("font-size: 16px; font-weight: bold;")
            
            if field["type"] == "date":
                input_field = QDateEdit()
                input_field.setCalendarPopup(True)
                input_field.setDate(QDate.currentDate())
                
                # Set standard date format
                input_field.setDisplayFormat("yyyy-MM-dd")
                
                input_field.setStyleSheet(f"""
                    QDateEdit {{
                        background-color: {DARKER_BG};
                        border: none;
                        border-radius: 5px;
                        padding: 8px;
                        color: {WHITE};
                        font-size: 16px;
                        text-align: right;
                    }}
                """)
                input_field.setLayoutDirection(Qt.RightToLeft)  # Set text direction to RTL
                
                # Connect date fields to calculate period automatically
                if field["name"] == "تاريخ البداية" or field["name"] == "تاريخ النهاية":
                    input_field.dateChanged.connect(self.calculate_leave_days)
            else:
                input_field = QLineEdit()
               
                input_field.setStyleSheet(f"""
                    QLineEdit {{
                        background-color: {DARKER_BG};
                        border: none;
                        border-radius: 5px;
                        padding: 8px;
                        color: {WHITE};
                        font-size: 16px;
                        text-align: right;
                    }}
                """)
                input_field.setLayoutDirection(Qt.RightToLeft)  # Set text direction to RTL
                input_field.setAlignment(Qt.AlignRight)  # Align text to right for Arabic
            
            field_layout.addWidget(label)
            field_layout.addWidget(input_field)
            container_layout.addLayout(field_layout)
            
            self.form_fields[field["name"]] = input_field
        
        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)
        
        # Save button - حفظ
        save_btn = QPushButton("حفظ")
        save_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #3d8b40;
            }}
        """)
        
        # Cancel button - إلغاء
        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)
        
        # Connect buttons
        save_btn.clicked.connect(self.save_leave)
        cancel_btn.clicked.connect(self.reject)
        
        # Add buttons in right-to-left order for Arabic
        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(save_btn)
        container_layout.addLayout(buttons_layout)
        
        # Add to form wrapper layout
        form_wrapper_layout.addWidget(self.title_label)
        form_wrapper_layout.addWidget(form_container)
        
        # Add form wrapper to center layout
        center_layout.addStretch()
        center_layout.addWidget(form_container_wrapper)
        center_layout.addStretch()
        
        # Add center container to main layout
        main_layout.addWidget(center_container)
        
        # Store calculated days
        self.calculated_days = 0
        
        # Store employee data
        self.employee_id = employee_id
        self.employee_name = employee_name
    
    def calculate_leave_days(self):
        """
        Calculate the leave days automatically based on start and end dates
        """
        # Check if both date fields exist in the form
        if "تاريخ البداية" in self.form_fields and "تاريخ النهاية" in self.form_fields:
            start_date = self.form_fields["تاريخ البداية"].date()
            end_date = self.form_fields["تاريخ النهاية"].date()
            
            # Calculate the difference in days
            days_diff = start_date.daysTo(end_date) + 1  # Include both start and end days
            
            # Store the calculated days for later use
            self.calculated_days = max(0, days_diff)
    
    def validate_form(self):
        """
        Validate the form fields with enhanced date validation
        """
        # Check required fields
        required_fields = ["رقم القرار"]
        
        for field_name in required_fields:
            field = self.form_fields[field_name]
            if isinstance(field, QLineEdit) and not field.text().strip():
                warning = CustomWarningDialog(self, "تحذير", f"الرجاء إدخال {field_name}")
                warning.exec_()
                field.setFocus()
                return False
                
        # Validate date range
        start_date = self.form_fields["تاريخ البداية"].date()
        end_date = self.form_fields["تاريخ النهاية"].date()
        
        if start_date > end_date:
            warning = CustomWarningDialog(self, "تحذير", "تاريخ البداية يجب أن يكون قبل تاريخ النهاية")
            warning.exec_()
            return False
        
        # Validate dates are within current year
        current_year = datetime.now().year
        if start_date.year() != current_year or end_date.year() != current_year:
            warning = CustomWarningDialog(self, "تحذير", f"التواريخ يجب أن تكون ضمن السنة الحالية ({current_year})")
            warning.exec_()
            return False
            
        return True
    
    def save_leave(self):
        """
        Save the leave data and return it to the parent
        """
        if not self.validate_form():
            return
        
        # Get form data
        decision_id = self.form_fields["رقم القرار"].text()
        decision_date = self.form_fields["تاريخ القرار"].date().toString("yyyy-MM-dd")
        start_date = self.form_fields["تاريخ البداية"].date().toString("yyyy-MM-dd")
        end_date = self.form_fields["تاريخ النهاية"].date().toString("yyyy-MM-dd")
        
        # Create data dictionary to return
        leave_data = {
            "decision_id": decision_id,
            "decision_date": decision_date,
            "start_date": start_date,
            "end_date": end_date,
            "days": self.calculated_days
        }
    # Enregistrer dans l'historique (si log_history est disponible)
        if hasattr(self, "log_history"):
            self.log_history(
             event="إنشاء إجازة سنوية جديدة",
             details=f"تم إنشاء إجازة سنوية للموظف: {self.employee_name} - السنة: {start_date[:4]} - الأيام المخصصة: {self.calculated_days} يوم",
             conge_id=None,  # Remplace par l'ID réel si disponible
             employee_id=self.employee_id
        )       
        # Set the result and accept the dialog
        self.leave_data = leave_data
        self.accept()
    

class EditTrancheDialog(QDialog):
    """
    Dialog for editing a tranche
    """
    def __init__(self, parent=None, tranche_data=None, tranche_id=None):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("تعديل الشطر")
        self.setMinimumSize(500, 400)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")
        
        # Set window flags to allow maximize/minimize
        self.setWindowFlags(
            Qt.Window |              # Regular window
            Qt.WindowCloseButtonHint 
        )
        
        # Store tranche data and ID
        self.tranche_data = tranche_data or {}
        self.tranche_id = tranche_id
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create a centered container for the form
        center_container = QWidget()
        center_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        center_layout = QHBoxLayout(center_container)
        
        # Form container with fixed width
        form_container_wrapper = QWidget()
        form_container_wrapper.setFixedWidth(500)
        form_wrapper_layout = QVBoxLayout(form_container_wrapper)
        form_wrapper_layout.setContentsMargins(20, 20, 20, 20)
        form_wrapper_layout.setSpacing(15)
        
        # Title
        self.title_label = QLabel(f"تعديل الشطر")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 20px; font-weight: bold; margin: 10px 0;")
        
        # Form container
        form_container = QWidget()
        form_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")
        
        container_layout = QVBoxLayout(form_container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.setSpacing(15)
        
        # Form fields
        fields = [
            {"name": "رقم القرار", "type": "text"},
            {"name": "تاريخ القرار", "type": "date"},
            {"name": "تاريخ البداية", "type": "date"},
            {"name": "تاريخ النهاية", "type": "date"}
        ]
        
        self.form_fields = {}
        
        for field in fields:
            field_layout = QVBoxLayout()
            field_layout.setSpacing(5)
            
            label = QLabel(field["name"])
            label.setStyleSheet("font-size: 16px; font-weight: bold;")
            
            if field["type"] == "date":
                input_field = QDateEdit()
                input_field.setCalendarPopup(True)
                input_field.setDate(QDate.currentDate())
                
                # Set standard date format
                input_field.setDisplayFormat("yyyy-MM-dd")
                
                input_field.setStyleSheet(f"""
                    QDateEdit {{
                        background-color: {DARKER_BG};
                        border: none;
                        border-radius: 5px;
                        padding: 8px;
                        color: {WHITE};
                        font-size: 16px;
                        text-align: right;
                    }}
                """)
                input_field.setLayoutDirection(Qt.RightToLeft)  # Set text direction to RTL
                
                # Connect date fields to calculate period automatically
                if field["name"] == "تاريخ البداية" or field["name"] == "تاريخ النهاية":
                    input_field.dateChanged.connect(self.calculate_tranche_days)
            else:
                input_field = QLineEdit()
                
                input_field.setStyleSheet(f"""
                    QLineEdit {{
                        background-color: {DARKER_BG};
                        border: none;
                        border-radius: 5px;
                        padding: 8px;
                        color: {WHITE};
                        font-size: 16px;
                        text-align: right;
                    }}
                """)
                input_field.setLayoutDirection(Qt.RightToLeft)  # Set text direction to RTL
                input_field.setAlignment(Qt.AlignRight)  # Align text to right for Arabic
            
            field_layout.addWidget(label)
            field_layout.addWidget(input_field)
            container_layout.addLayout(field_layout)
            
            self.form_fields[field["name"]] = input_field
        
        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)
        
        # Save button - حفظ
        save_btn = QPushButton("حفظ")
        save_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {GREEN};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #3d8b40;
            }}
        """)
        
        # Cancel button - إلغاء
        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)
        
        # Connect buttons
        save_btn.clicked.connect(self.save_tranche)
        cancel_btn.clicked.connect(self.reject)
        
        # Add buttons in right-to-left order for Arabic
        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)
        
        container_layout.addLayout(buttons_layout)
        
        # Add to form wrapper layout
        form_wrapper_layout.addWidget(self.title_label)
        form_wrapper_layout.addWidget(form_container)
        
        # Add form wrapper to center layout
        center_layout.addStretch()
        center_layout.addWidget(form_container_wrapper)
        center_layout.addStretch()
        
        # Add center container to main layout
        main_layout.addWidget(center_container)
        
        # Store calculated days
        self.calculated_tranche_days = 0
        
        # Fill form with tranche data
        self.fill_form()
    
    def fill_form(self):
        """
        Fill the form with tranche data
        """
        if "رقم القرار" in self.form_fields:
            self.form_fields["رقم القرار"].setText(str(self.tranche_data.get("decision_id", "")))
            
        if "تاريخ القرار" in self.form_fields:
            decision_date = QDate.fromString(self.tranche_data.get("decision_date", ""), "yyyy-MM-dd")
            if decision_date.isValid():
                self.form_fields["تاريخ القرار"].setDate(decision_date)
            else:
                self.form_fields["تاريخ القرار"].setDate(QDate.currentDate())
                
        if "تاريخ البداية" in self.form_fields:
            start_date = QDate.fromString(self.tranche_data.get("start_date", ""), "yyyy-MM-dd")
            if start_date.isValid():
                self.form_fields["تاريخ البداية"].setDate(start_date)
            else:
                self.form_fields["تاريخ البداية"].setDate(QDate.currentDate())
                
        if "تاريخ النهاية" in self.form_fields:
            end_date = QDate.fromString(self.tranche_data.get("end_date", ""), "yyyy-MM-dd")
            if end_date.isValid():
                self.form_fields["تاريخ النهاية"].setDate(end_date)
            else:
                self.form_fields["تاريخ النهاية"].setDate(QDate.currentDate())
        
        # Calculate days
        self.calculate_tranche_days()
    
    def calculate_tranche_days(self):
        """
        Calculate the tranche days automatically based on start and end dates
        """
        # Check if both date fields exist in the form
        if "تاريخ البداية" in self.form_fields and "تاريخ النهاية" in self.form_fields:
            start_date = self.form_fields["تاريخ البداية"].date()
            end_date = self.form_fields["تاريخ النهاية"].date()
            
            # Calculate the difference in days
            days_diff = start_date.daysTo(end_date) + 1  # Include both start and end days
            
            # Store the calculated days for later use
            self.calculated_tranche_days = max(0, days_diff)
    
    def validate_form(self):
        """
        Validate the form fields with enhanced validation
        """
        # Check required fields
        required_fields = ["رقم القرار"]
        
        for field_name in required_fields:
            field = self.form_fields[field_name]
            if isinstance(field, QLineEdit) and not field.text().strip():
                warning = CustomWarningDialog(self, "تحذير", f"الرجاء إدخال {field_name}")
                warning.exec_()
                field.setFocus()
                return False
                
        # Validate date range
        start_date = self.form_fields["تاريخ البداية"].date()
        end_date = self.form_fields["تاريخ النهاية"].date()
        
        if start_date > end_date:
            warning = CustomWarningDialog(self, "تحذير", "تاريخ البداية يجب أن يكون قبل تاريخ النهاية")
            warning.exec_()
            return False
        
        # Validate dates are within current year
        current_year = datetime.now().year
        if start_date.year() != current_year or end_date.year() != current_year:
            warning = CustomWarningDialog(self, "تحذير", f"التواريخ يجب أن تكون ضمن السنة الحالية ({current_year})")
            warning.exec_()
            return False
            
        return True
    
    def save_tranche(self):
        """
        Save the tranche data and return it to the parent
        """
        if not self.validate_form():
            return
        
        # Get form data
        decision_id = self.form_fields["رقم القرار"].text()
        decision_date = self.form_fields["تاريخ القرار"].date().toString("yyyy-MM-dd")
        start_date = self.form_fields["تاريخ البداية"].date().toString("yyyy-MM-dd")
        end_date = self.form_fields["تاريخ النهاية"].date().toString("yyyy-MM-dd")
        
        # Create data dictionary to return
        tranche_data = {
            "id": self.tranche_id,
            "decision_id": decision_id,
            "decision_date": decision_date,
            "start_date": start_date,
            "end_date": end_date,
            "days": self.calculated_tranche_days
        }
        
        # Set the result and accept the dialog
        self.tranche_data = tranche_data
        self.accept()

class LeaveDetailsDialog(QDialog):
    """
    Dialog for showing leave details
    """
    def __init__(self, parent=None, leave_id="", employee_data=None, tranches=None):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("تفاصيل العطلة")
        self.setMinimumSize(900, 600)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")


        # Set window flags to allow maximize/minimize
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)
        
        # Store data
        self.leave_id = leave_id
        self.employee_data = employee_data or {}
        self.tranches = tranches or []
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create a centered container for the content
        center_container = QWidget()
        center_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        center_layout = QVBoxLayout(center_container)
        center_layout.setContentsMargins(20, 20, 20, 20)
        center_layout.setSpacing(15)

        # Add top bar from sidebar.py
        self.top_bar_widget, self.export_button_topbar, self.print_button_topbar = create_top_bar(
            self,           # parent
            center_layout,  # layout où ajouter la topbar (assurez-vous que center_layout est défini avant)
            None,           # sidebar_toggle
            excel_export_action_callback=self.export_data_to_excel,
            pdf_print_action_callback=self.print_data_to_pdf
        )

        # Title
        title_label = QLabel("جدول تفاصيل العطلة")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        
        # Tranche details table
        tranche_container = QWidget()
        tranche_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")
        tranche_layout = QVBoxLayout(tranche_container)
        tranche_layout.setContentsMargins(15, 15, 15, 15)
        
        self.tranche_table = QTableWidget()
        self.tranche_table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.tranche_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                border: none;
                gridline-color: {LIGHT_BG};
            }}
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid {LIGHT_BG};
            }}
            QTableWidget::item:selected {{
                background-color: {ORANGE};
                color: {WHITE};
            }}
            QHeaderView::section {{
                background-color: {DARKER_BG};
                color: {WHITE};
                padding: 8px;
                border: none;
                font-weight: bold;
                text-align: center;
            }}
            QScrollBar {{
                width: 0px;
                height: 0px;
            }}
            QscrollBar:horizontal {{
                width: 0px;
                margin: 0px;
                }}
        """)
        
        # Set up columns for tranche table
        tranche_columns = [
            "رقم العطلة", "رقم الموظف", "لقب الموظف", "اسم الموظف", 
            "رقم الشطر", "رقم القرار", "تاريخ القرار", "تاريخ البداية", "تاريخ النهاية"
        ]
        
        self.tranche_table.setColumnCount(len(tranche_columns))
        self.tranche_table.setHorizontalHeaderLabels(tranche_columns)
        self.tranche_table.horizontalHeader().setStretchLastSection(True)
        self.tranche_table.verticalHeader().setVisible(False)
        self.tranche_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # Set text alignment for all columns to center-aligned
        for i in range(len(tranche_columns)):
            self.tranche_table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignCenter)
        
        # Connect double-click signal to edit tranche
        
        
        tranche_layout.addWidget(self.tranche_table)
        
        # Allow moving columns
        self.tranche_table.horizontalHeader().setSectionsMovable(True)
        # Set column resizing fixed
        self.tranche_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)

        # Action buttons at the bottom
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 10, 0, 0)
        buttons_layout.setSpacing(15)
        
        edit_btn = QPushButton("تعديل")
        delete_btn = QPushButton("حذف")
        
        for btn in [edit_btn, delete_btn]:
            btn.setFixedSize(120, 40)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
            """)
        
        # Make delete button red
        delete_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {RED};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                font-size: 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)
        
        # Connect buttons
        edit_btn.clicked.connect(self.edit_selected_tranche)
        delete_btn.clicked.connect(self.delete_selected_tranche)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(edit_btn)
        buttons_layout.addWidget(delete_btn)
        buttons_layout.addStretch()
        
        # Add all widgets to the center layout
        center_layout.addWidget(title_label)
        center_layout.addWidget(tranche_container)
        center_layout.addWidget(buttons_widget)
        
        # Add center container to main layout
        main_layout.addWidget(center_container)
        
        # Populate table with tranche data
        self.populate_table()

    
    
    def _get_table_data_as_lists_tranches(self): # Nom spécifique
        """Helper function to extract data from self.tranche_table."""
        if not hasattr(self, 'tranche_table') or not self.tranche_table or self.tranche_table.rowCount() == 0:
            return None, None
        headers = [self.tranche_table.horizontalHeaderItem(i).text() for i in range(self.tranche_table.columnCount())]
        all_row_data = []
        for row in range(self.tranche_table.rowCount()):
            # Pas de isRowHidden ici car cette table n'a pas de paginateur dédié visiblement
            row_data = [self.tranche_table.item(row, col).text() if self.tranche_table.item(row, col) else "" 
                        for col in range(self.tranche_table.columnCount())]
            all_row_data.append(row_data)
        return headers, all_row_data

    def export_data_to_excel(self):
        print(f"{self.__class__.__name__}: export_data_to_excel (détails tranches) called")
        headers, data = self._get_table_data_as_lists_tranches()
        if not data:
            QMessageBox.information(self, "لا بيانات للتصدير", "لا توجد تفاصيل أشطر للعرض.")
            return
        filePath, _ = QFileDialog.getSaveFileName(self, "تصدير تفاصيل الأشطر إلى Excel", 
                                                  os.path.expanduser(f"~/Documents/Details_Tranches_Conge_{self.leave_id}.xlsx"),
                                                  "Excel Workbook (*.xlsx);;All Files (*)")
        if not filePath: return
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"أشطر العطلة {self.leave_id}"
            sheet.sheet_view.rightToLeft = True
            sheet.append(headers)
            for row_values in data: sheet.append(row_values)
            for col_idx, column_cells in enumerate(sheet.columns):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = length + 5
            workbook.save(filePath)
            QMessageBox.information(self, "نجاح التصدير", f"تم تصدير بيانات الأشطر بنجاح إلى:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ: {e}")

    def print_data_to_pdf(self):
        print(f"{self.__class__.__name__}: print_data_to_pdf (détails tranches) called")
        headers, data = self._get_table_data_as_lists_tranches()
        if not data:
            QMessageBox.information(self, "لا بيانات للطباعة", "لا توجد تفاصيل أشطر للعرض.")
            return
        filePath, _ = QFileDialog.getSaveFileName(self, "طباعة تفاصيل الأشطر إلى PDF", 
                                                  os.path.expanduser(f"~/Documents/Details_Tranches_Conge_{self.leave_id}.pdf"),
                                                  "PDF Document (*.pdf);;All Files (*)")
        if not filePath: return
        html_content = "<html><head><meta charset='UTF-8'><style>"
        html_content += "body { direction: rtl; font-family: 'Arial', 'DejaVu Sans', sans-serif; font-size: 9pt; } table { width: 100%; border-collapse: collapse; margin-top: 10px; } th, td { border: 1px solid #333; padding: 4px; text-align: right; word-wrap: break-word; } th { background-color: #f0f0f0; font-weight: bold; } caption { font-size: 1.1em; font-weight: bold; margin-bottom: 8px; text-align: center; }"
        html_content += "</style></head><body>"
        html_content += f"<table><caption>تفاصيل أشطر العطلة رقم {self.leave_id} للموظف {self.employee_data.get('first_name', '')} {self.employee_data.get('last_name', '')}</caption><thead><tr>"
        for header in headers: html_content += f"<th>{header}</th>"
        html_content += "</tr></thead><tbody>"
        for row_values in data:
            html_content += "<tr>"
            for cell_value in row_values: html_content += f"<td>{html.escape(str(cell_value))}</td>"
            html_content += "</tr>"
        html_content += "</tbody></table></body></html>"
        try:
            css_style = CSS(string='@page { size: A4 landscape; margin: 1cm; }')
            HTML(string=html_content).write_pdf(filePath, stylesheets=[css_style])
            QMessageBox.information(self, "نجاح الطباعة", f"تم إنشاء ملف PDF بنجاح:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في الطباعة", f"حدث خطأ: {e}")

    
    
    
    def populate_table(self):
        """
        Populate the table with tranche data - FIXED to show decision number
        """
        # Clear existing rows
        self.tranche_table.setRowCount(0)
        
        # Add rows for each tranche
        for i, tranche in enumerate(self.tranches):
            # Add a new row
            row_idx = self.tranche_table.rowCount()
            self.tranche_table.insertRow(row_idx)
            
            # Set data for each column - NOW INCLUDING DECISION NUMBER
            self.tranche_table.setItem(row_idx, 0, QTableWidgetItem(self.leave_id))
            self.tranche_table.setItem(row_idx, 1, QTableWidgetItem(self.employee_data.get("employee_id", "")))
            self.tranche_table.setItem(row_idx, 2, QTableWidgetItem(self.employee_data.get("last_name", "")))
            self.tranche_table.setItem(row_idx, 3, QTableWidgetItem(self.employee_data.get("first_name", "")))
            self.tranche_table.setItem(row_idx, 4, QTableWidgetItem(f"الشطر {i+1}"))
            self.tranche_table.setItem(row_idx, 5, QTableWidgetItem(str(tranche.get("decision_id", ""))))  # FIXED: Decision number now displayed
            self.tranche_table.setItem(row_idx, 6, QTableWidgetItem(tranche.get("decision_date", "")))
            self.tranche_table.setItem(row_idx, 7, QTableWidgetItem(tranche.get("start_date", "")))
            self.tranche_table.setItem(row_idx, 8, QTableWidgetItem(tranche.get("end_date", "")))
            
            # Center align all items in the row
            for col in range(self.tranche_table.columnCount()):
                item = self.tranche_table.item(row_idx, col)
                if item:
                    item.setTextAlignment(Qt.AlignCenter)
    
    def edit_tranche(self, row, column):
        """
        Edit a tranche when double-clicked
        """
        self.edit_selected_tranche(row)
    
    def edit_selected_tranche(self, row=None):
        """
        Edit the selected tranche - FIXED to use tranche ID
        """
        row = self.tranche_table.currentRow()
        if row < 0:  # No selection at all
            warning = CustomWarningDialog(self, "تحذير", "الرجاء اختيار شطر للتعديل")
            warning.exec_()
            return
        # Get tranche data using the actual tranche from the list
        if row < len(self.tranches):
            tranche_data = self.tranches[row]
            tranche_id = tranche_data.get("id")
            
            # Open edit dialog
            edit_dialog = EditTrancheDialog(self, tranche_data, tranche_id)
            if edit_dialog.exec_() == QDialog.Accepted:
                # Update tranche data
                updated_data = edit_dialog.tranche_data
                
                # Update table
                self.tranche_table.setItem(row, 5, QTableWidgetItem(str(updated_data.get("decision_id", ""))))
                self.tranche_table.setItem(row, 6, QTableWidgetItem(updated_data.get("decision_date", "")))
                self.tranche_table.setItem(row, 7, QTableWidgetItem(updated_data.get("start_date", "")))
                self.tranche_table.setItem(row, 8, QTableWidgetItem(updated_data.get("end_date", "")))
                
                # Center align all updated items
                for col in [5, 6, 7, 8]:
                    item = self.tranche_table.item(row, col)
                    if item:
                        item.setTextAlignment(Qt.AlignCenter)
                
                # Update tranches list
                self.tranches[row].update(updated_data)
                
                # Notify parent of the change
                self.parent.update_tranche_by_id(tranche_id, updated_data)


    
    def delete_selected_tranche(self):
        """
        Delete the selected tranche - FIXED to use tranche ID
        """
        selected_rows = self.tranche_table.selectionModel().selectedRows()
        if not selected_rows:
            warning = CustomWarningDialog(self, "تحذير", "الرجاء اختيار شطر للحذف")
            warning.exec_()
            return
        
        row = selected_rows[0].row()
        
        # Show confirmation dialog
        dialog = CustomMessageBox(
            self,
            "حذف الشطر",
            "هل أنت متأكد أنك تريد حذف هذا الشطر؟"
        )
        
        if dialog.exec_() == QDialog.Accepted:
            try:
                # Get the actual tranche ID from the stored tranches data
                if row < len(self.tranches):
                    tranche_to_delete = self.tranches[row]
                    tranche_id = tranche_to_delete.get("id")
                    
                    if tranche_id:
                        # Delete from database using tranche ID
                        self.parent.delete_tranche_by_id(tranche_id)
                        
                        # Remove the row from the table
                        self.tranche_table.removeRow(row)
                        
                        # Remove from tranches list
                        del self.tranches[row]
                        
                        # Update row numbers for remaining tranches
                        self.update_tranche_numbers()
                    else:
                        CustomWarningDialog(self, "خطأ", "لا يمكن العثور على معرف الشطر").exec_()
                else:
                    CustomWarningDialog(self, "خطأ", "فهرس الشطر غير صحيح").exec_()
                    
            except Exception as e:
                CustomWarningDialog(self, "خطأ", f"خطأ في حذف الشطر: {str(e)}").exec_()

    def update_tranche_numbers(self):
        """
        Update tranche numbers after deletion to maintain correct sequence
        """
        for i in range(self.tranche_table.rowCount()):
            # Update the tranche number column (column 4)
            tranche_number_item = QTableWidgetItem(f"الشطر {i+1}")
            tranche_number_item.setTextAlignment(Qt.AlignCenter)
            self.tranche_table.setItem(i, 4, tranche_number_item)



class PreviousYearsDialog(QDialog):
    """
    Dialog for showing previous years data with pagination and filtering
    """
    def __init__(self, parent=None, previous_years_data=None):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("عطلات السنوات السابقة")
        self.setMinimumSize(1000, 600)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")
        
        # Set window flags to allow maximize/minimize
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)
        
        # Store data
        self.previous_years_data = previous_years_data or []
        self.original_data = previous_years_data or []  # Keep original data for filtering
        
        # Initialize filter state
        self.filter_dialog = None
        self.selected_filter_columns = []
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create a centered container for the content
        center_container = QWidget()
        center_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        center_layout = QVBoxLayout(center_container)
        center_layout.setContentsMargins(20, 20, 20, 20)
        center_layout.setSpacing(15)

        self.top_bar_widget, self.export_button_topbar, self.print_button_topbar = create_top_bar(
            self,           # parent
            center_layout,  # layout où ajouter la topbar (assurez-vous que center_layout est défini avant)
            None,           # sidebar_toggle
            excel_export_action_callback=self.export_data_to_excel,
            pdf_print_action_callback=self.print_data_to_pdf
        )
        
        # Add action bar with filter and refresh buttons
        self.create_action_bar(center_layout)
        
        # Title
        title_label = QLabel("جدول تفاصيل عطلات السنوات السابقة")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        
        # Table container
        table_container = QWidget()
        table_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(15, 15, 15, 15)
        
        # Previous years table
        self.previous_years_table = QTableWidget()
        self.previous_years_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                border: none;
                gridline-color: {LIGHT_BG};
            }}
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid {LIGHT_BG};
            }}
            QHeaderView::section {{
                background-color: {DARKER_BG};
                color: {WHITE};
                padding: 8px;
                border: none;
                font-weight: bold;
                text-align: center;
            }}
            QScrollBar {{
                width: 0px;
                height: 0px;
            }}
            QscrollBar:horizontal {{
                width: 0px;
                margin: 0px;
                }}
        """)
        
        # Set up columns for previous years table
        columns = [
            "رقم العطلة", "السنة", "رقم الموظف", "لقب الموظف", "اسم الموظف", 
            "عدد الأيام المستحقة", "الشطر 1", "الشطر 2", "الشطر 3", "الشطر 4", "الشطر 5",
            "عدد الأيام المستهلكة", "عدد الأيام المتبقية"
        ]
        
        self.previous_years_table.setColumnCount(len(columns))
        self.previous_years_table.setHorizontalHeaderLabels(columns)
        self.previous_years_table.horizontalHeader().setStretchLastSection(True)
        self.previous_years_table.verticalHeader().setVisible(False)
        
        # Set text alignment for all columns to center-aligned
        for i in range(len(columns)):
            self.previous_years_table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignCenter)
        
        # Allow moving columns
        self.previous_years_table.horizontalHeader().setSectionsMovable(True)
        # Set column resizing fixed
        self.previous_years_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        
        # Add table to container
        table_layout.addWidget(self.previous_years_table)
        
        # Add pagination to the table
        self.paginator = tablepaginator(self.previous_years_table, rows_per_page=10)
        table_layout.addWidget(self.paginator)
        
        # Add all widgets to the center layout
        center_layout.addWidget(title_label)
        center_layout.addWidget(table_container)
        
        # Add center container to main layout
        main_layout.addWidget(center_container)
        
        # Load initial data
        self.load_previous_years_data()

    def create_action_bar(self, center_layout):
        """
        Create action bar with filter and refresh buttons
        """
        action_bar = QWidget()
        action_layout = QHBoxLayout(action_bar)
        action_layout.setContentsMargins(0, 0, 0, 10)
        action_layout.setSpacing(10)

        # Filter button
        filter_btn = QPushButton("ترشيح")
        filter_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {MEDIUM_BG};
                border: none;
                border-radius: 8px;
                color: {WHITE};
                font-weight: bold;
                font-size: 16px;
                padding: 8px 15px;
                min-width: 80px;
            }}
            QPushButton:hover {{
                background-color: {LIGHT_BG};
            }}
        """)
        
        # Load and set filter icon
        filter_icon = QPixmap("pics/filter.png")
        if not filter_icon.isNull():
            filter_icon = filter_icon.scaled(16, 16, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            filter_btn.setIcon(QIcon(filter_icon))
            filter_btn.setIconSize(QSize(16, 16))
        
        filter_btn.setToolTip("ترشيح")
        filter_btn.clicked.connect(self.show_filter_dialog)

        # Refresh button
        refresh_btn = QPushButton("⟳")
        refresh_btn.setFixedSize(40, 40)
        refresh_btn.setToolTip("تحديث البيانات")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 17px;
                color: #ffffff;
                font-weight: bold;
                font-size: 25px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_previous_years_data)
        
        # Add to layout
        action_layout.addWidget(filter_btn, alignment=Qt.AlignLeft)
        action_layout.addStretch()
        action_layout.addWidget(refresh_btn, alignment=Qt.AlignRight)
        
        # Insert into center layout
        center_layout.addWidget(action_bar)

    def show_filter_dialog(self):
        """
        Show the filter dialog with checkboxes for column selection
        """
        # Create filter dialog if it doesn't exist
        if not self.filter_dialog:
            self.filter_dialog = QDialog(self)
            self.filter_dialog.setWindowTitle("ترشيح")
            self.filter_dialog.setFixedWidth(350)
            self.filter_dialog.setStyleSheet(f"background-color: {MEDIUM_BG}; color: {WHITE};")
            
            dialog_layout = QVBoxLayout(self.filter_dialog)
            dialog_layout.setContentsMargins(20, 20, 20, 20)
            
            # Title
            title_label = QLabel("اختر الأعمدة للترشيح:")
            title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 15px;")
            dialog_layout.addWidget(title_label)
            
            # Create checkboxes for each column
            columns = [
                "رقم العطلة", "السنة", "رقم الموظف", "لقب الموظف", "اسم الموظف", 
                "عدد الأيام المستحقة", "الشطر 1", "الشطر 2", "الشطر 3", "الشطر 4", "الشطر 5",
                "عدد الأيام المستهلكة", "عدد الأيام المتبقية"
            ]
            
            self.filter_checkboxes = {}
            
            for column in columns:
                checkbox = QCheckBox(column)
                checkbox.setStyleSheet(f"""
                    QCheckBox {{
                        font-size: 16px;
                        padding: 8px 0;
                    }}
                    QCheckBox::indicator {{
                        width: 18px;
                        height: 18px;
                    }}
                    QCheckBox::indicator:checked {{
                        background-color: {ORANGE};
                        border: 2px solid {WHITE};
                    }}
                """)
                dialog_layout.addWidget(checkbox)
                self.filter_checkboxes[column] = checkbox
            
            # Filter value input
            filter_value_layout = QHBoxLayout()
            filter_value_layout.setSpacing(10)
            
            filter_value_label = QLabel("قيمة الترشيح:")
            filter_value_label.setStyleSheet("font-size: 16px;")
            
            self.filter_value_input = QLineEdit()
            self.filter_value_input.setPlaceholderText("أدخل قيمة للترشيح...")
            self.filter_value_input.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    padding: 10px;
                    color: {BLACK};
                    font-size: 16px;
                    text-align: right;
                }}
            """)
            self.filter_value_input.setLayoutDirection(Qt.RightToLeft)
            
            filter_value_layout.addWidget(filter_value_label)
            filter_value_layout.addWidget(self.filter_value_input)
            
            dialog_layout.addSpacing(15)
            dialog_layout.addLayout(filter_value_layout)
            
            # Buttons
            buttons_layout = QHBoxLayout()
            buttons_layout.setSpacing(15)
            
            apply_btn = QPushButton("تطبيق")
            apply_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    padding: 10px;
                    font-size: 16px;
                    font-weight: bold;
                    min-width: 100px;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
            """)
            apply_btn.clicked.connect(self.apply_filter)
            
            cancel_btn = QPushButton("إلغاء")
            cancel_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {MEDIUM_BG};
                    color: {WHITE};
                    border: 1px solid {LIGHT_BG};
                    border-radius: 5px;
                    padding: 10px;
                    font-size: 16px;
                    min-width: 100px;
                }}
                QPushButton:hover {{
                    background-color: {LIGHT_BG};
                }}
            """)
            cancel_btn.clicked.connect(self.filter_dialog.reject)
            
            # Add buttons in right-to-left order for Arabic
            buttons_layout.addWidget(cancel_btn)
            buttons_layout.addWidget(apply_btn)
            
            dialog_layout.addSpacing(15)
            dialog_layout.addLayout(buttons_layout)
        
        # Show the dialog
        self.filter_dialog.exec_()

    def apply_filter(self):
        """
        Apply the filter to the table based on selected columns and input value
        """
        # Get selected columns
        self.selected_filter_columns = []
        for column, checkbox in self.filter_checkboxes.items():
            if checkbox.isChecked():
                self.selected_filter_columns.append(column)
        
        # Get filter value
        filter_value = self.filter_value_input.text().strip().lower()
        
        if not filter_value or not self.selected_filter_columns:
            # Reset filter if no value or no columns selected - show all data
            self.previous_years_data = self.original_data.copy()
            self.load_previous_years_data()
            self.filter_dialog.accept()
            return
        
        # Get column indices for selected columns
        column_indices = []
        for col_name in self.selected_filter_columns:
            for col_idx in range(self.previous_years_table.columnCount()):
                if self.previous_years_table.horizontalHeaderItem(col_idx).text() == col_name:
                    column_indices.append(col_idx)
        
        # Filter the data
        filtered_data = []
        for row_data in self.original_data:
            row_visible = False
            
            # Check if any selected column contains the filter value
            for col_idx in column_indices:
                if col_idx < len(row_data):
                    cell_value = str(row_data[col_idx]).lower()
                    if filter_value in cell_value:
                        row_visible = True
                        break
            
            if row_visible:
                filtered_data.append(row_data)
        
        # Update the displayed data
        self.previous_years_data = filtered_data
        self.load_previous_years_data()
        
        # Close the dialog
        self.filter_dialog.accept()


    def _get_table_data_as_lists_previous(self): # Nom spécifique
        """Helper function to extract data ONLY from VISIBLE rows in self.previous_years_table."""
        if not hasattr(self, 'previous_years_table') or not self.previous_years_table or self.previous_years_table.rowCount() == 0:
            return None, None
        headers = [self.previous_years_table.horizontalHeaderItem(i).text() for i in range(self.previous_years_table.columnCount())]
        visible_row_data = []
        for row in range(self.previous_years_table.rowCount()):
            if not self.previous_years_table.isRowHidden(row): # Respecte la pagination
                row_data = [self.previous_years_table.item(row, col).text() if self.previous_years_table.item(row, col) else "" 
                            for col in range(self.previous_years_table.columnCount())]
                visible_row_data.append(row_data)
        return headers, visible_row_data

    def export_data_to_excel(self):
        print(f"{self.__class__.__name__}: export_data_to_excel (congés années précédentes) called")
        headers, data = self._get_table_data_as_lists_previous()
        if not data:
            QMessageBox.information(self, "لا بيانات للتصدير", "لا توجد عطلات سابقة للعرض.")
            return
        filePath, _ = QFileDialog.getSaveFileName(self, "تصدير عطلات سابقة إلى Excel", 
                                                  os.path.expanduser(f"~/Documents/Conges_Annees_Precedentes.xlsx"),
                                                  "Excel Workbook (*.xlsx);;All Files (*)")
        if not filePath: return
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "عطلات سنوات سابقة"
            sheet.sheet_view.rightToLeft = True
            sheet.append(headers)
            for row_values in data: sheet.append(row_values)
            for col_idx, column_cells in enumerate(sheet.columns):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = length + 5
            workbook.save(filePath)
            QMessageBox.information(self, "نجاح التصدير", f"تم تصدير البيانات بنجاح إلى:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ: {e}")

    def print_data_to_pdf(self):
        print(f"{self.__class__.__name__}: print_data_to_pdf (congés années précédentes) called")
        headers, data = self._get_table_data_as_lists_previous()
        if not data:
            QMessageBox.information(self, "لا بيانات للطباعة", "لا توجد عطلات سابقة للعرض.")
            return
        filePath, _ = QFileDialog.getSaveFileName(self, "طباعة عطلات سابقة إلى PDF", 
                                                  os.path.expanduser(f"~/Documents/Conges_Annees_Precedentes.pdf"),
                                                  "PDF Document (*.pdf);;All Files (*)")
        if not filePath: return
        html_content = "<html><head><meta charset='UTF-8'><style>"
        html_content += "body { direction: rtl; font-family: 'Arial', 'DejaVu Sans', sans-serif; font-size: 8pt; } table { width: 100%; border-collapse: collapse; margin-top: 10px; } th, td { border: 1px solid #333; padding: 3px; text-align: right; word-wrap: break-word; } th { background-color: #f0f0f0; font-weight: bold; } caption { font-size: 1.0em; font-weight: bold; margin-bottom: 7px; text-align: center; }"
        html_content += "</style></head><body>"
        html_content += f"<table><caption>تفاصيل عطلات السنوات السابقة</caption><thead><tr>"
        for header in headers: html_content += f"<th>{header}</th>"
        html_content += "</tr></thead><tbody>"
        for row_values in data:
            html_content += "<tr>"
            for cell_value in row_values: html_content += f"<td>{html.escape(str(cell_value))}</td>"
            html_content += "</tr>"
        html_content += "</tbody></table></body></html>"
        try:
            css_style = CSS(string='@page { size: A3 landscape; margin: 0.7cm; }') # A3 car beaucoup de colonnes
            HTML(string=html_content).write_pdf(filePath, stylesheets=[css_style])
            QMessageBox.information(self, "نجاح الطباعة", f"تم إنشاء ملف PDF بنجاح:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في الطباعة", f"حدث خطأ: {e}")

    def load_previous_years_data(self):
        """
        Load previous years data into the table with pagination
        """
        # Clear existing rows
        self.previous_years_table.setRowCount(0)
        
        # Add data to table
        for row, data in enumerate(self.previous_years_data):
            self.previous_years_table.insertRow(row)
            for col, value in enumerate(data):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.previous_years_table.setItem(row, col, item)
        
        # Update paginator with new data
        self.paginator.update_total_rows()
        self.paginator.update_page(1)  # Always start from page 1

    def refresh_previous_years_data(self):
        """
        Refresh previous years data and reset pagination to first page
        """
        try:
            # Get fresh data from parent
            if self.parent:
                # Get current year
                current_year = datetime.now().year
                
                # Get previous years data from database
                previous_conges = self.parent.conge_controller.get_previous_years_conges(current_year)
                
                # Format data for display
                refreshed_data = []
                
                for conge in previous_conges:
                    # Get tranches for this conge
                    tranches = self.parent.tranche_controller.get_tranches_by_conge(conge.idConge)
                    
                    # Format tranche data
                    tranche_values = ["0", "0", "0", "0", "0"]  # Default 5 empty tranches
                    
                    for i, tranche in enumerate(tranches[:5]):  # Limit to 5 tranches
                        days = (tranche.DateFin - tranche.DateDebut).days + 1
                        tranche_values[i] = str(days)
                    
                    # Add row data
                    row_data = [
                        str(conge.idConge),
                        str(conge.Annee),
                        str(conge.idemploye),
                        conge.employe.Nom if conge.employe else "",
                        conge.employe.Prenom if conge.employe else "",
                        str(conge.NbrJoursAlloues),
                        tranche_values[0],
                        tranche_values[1],
                        tranche_values[2],
                        tranche_values[3],
                        tranche_values[4],
                        str(conge.NbrJoursPris),
                        str(conge.NbrJoursRestants)
                    ]
                    
                    refreshed_data.append(row_data)
                
                # Update both original and displayed data
                self.original_data = refreshed_data
                self.previous_years_data = refreshed_data.copy()
                
                # Reset any applied filters
                self.selected_filter_columns = []
                if hasattr(self, 'filter_value_input'):
                    self.filter_value_input.clear()
                if hasattr(self, 'filter_checkboxes'):
                    for checkbox in self.filter_checkboxes.values():
                        checkbox.setChecked(False)
                
                # Reload the table with fresh data
                self.load_previous_years_data()
                
        except Exception as e:
            CustomWarningDialog(self, "خطأ", f"خطأ في تحديث البيانات: {str(e)}").exec_()

class LeaveManagementSystem(QMainWindow):
    def __init__(self, current_user_data=None, session=None):
        """
        Initialize the main application window and set up the UI components
        """
        super().__init__()
        self.setWindowTitle("نظام إدارة الموارد البشرية - إدارة الإجازات")
        self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet(f"background-color: {DARK_BG}; color: {WHITE};")
        self.current_user_data = current_user_data or {}
        current_user_account_number =  None
        if current_user_data:
          current_user_account_number = current_user_data.get('account_number')        

    
        
        # MODIFICATION: Utiliser la session partagée ou créer une nouvelle si nécessaire
        if session:
            self.session = session
            print("DEBUG - Utilisation de la session partagée dans gestionConges")
        else:
            print("DEBUG - Création d'une nouvelle session dans gestionconges")
            self.session = init_db('mysql+pymysql://hr:hr@localhost/HR')
        # Initialize controllers
        self.conge_controller = CongeController(session,current_user_account_number)
        self.tranche_controller = TrancheController(session,current_user_account_number)
        
        # Set up the central widget
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        # Main layout
        self.main_layout = QHBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)
        
        # Create main content area
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(0)
        
        # Create main page
        self.main_page = QWidget()
        self.main_page_layout = QVBoxLayout(self.main_page)
        self.main_page_layout.setContentsMargins(15, 15, 15, 15)
        
        # Create action bar with Filter button and Add button
        self.create_action_bar()
        
        # Add title for the main table
        self.create_table_title()
        
        # Create table and buttons
        self.create_table()
        self.create_action_buttons()
        
        # Add main page to content layout
        self.content_layout.addWidget(self.main_page)
        
        # Add content to main layout
        self.main_layout.addWidget(self.content_widget)
        
        # Load data from database
        self.force_refresh_data()
        
        # Initialize filter state
        self.filter_dialog = None
        self.selected_filter_columns = []
        
        # Get current date for validation
        self.today = QDate.currentDate()
        
        # Current selected tranche for editing
        self.current_tranche_index = -1
        self.current_leave_id = -1
        
        # Store tranches data for each leave
        self.leave_tranches = {}
        
        # History data
        self.history_data = [
            ["2025-04-10 14:30", "أحمد محمد", "إضافة", "إضافة عطلة جديدة للموظف رقم 1001"],
            ["2025-04-09 10:15", "سارة أحمد", "تعديل", "تعديل عطلة للموظف رقم 1002"],
            ["2025-04-08 16:45", "محمد علي", "حذف", "حذف عطلة للموظف رقم 1003"],
            ["2025-04-07 09:30", "فاطمة محمد", "إضافة", "إضافة عطلة جديدة للموظف رقم 1004"],
            ["2025-04-06 11:20", "أحمد محمد", "تعديل", "تعديل عطلة للموظف رقم 1005"]
        ]

    def force_refresh_data(self):
        """
        Force refresh data including creating annual leaves for new employees
        """
        try:
            current_year = datetime.now().year
            # Always ensure annual leaves are created for all employees
            self.conge_controller.create_annual_leaves_for_all_employees(current_year)
            # Load the updated data
            self.load_data_from_database()
        except Exception as e:
            print(f"Error in force refresh: {e}")
            # Fallback to regular load
            self.load_data_from_database()

    def create_action_bar(self):
        """
        Create the action bar with Filter button and Refresh button (moved Add button to bottom)
        """
        action_bar = QWidget()
        action_layout = QHBoxLayout(action_bar)
        action_layout.setContentsMargins(0, 0, 0, 20)  # Add margin at bottom
        action_layout.setSpacing(10)

        # Refresh button as requested
        refresh_btn = QPushButton("⟳")
        refresh_btn.setFixedSize(40, 40)
        refresh_btn.setToolTip("تحديث الجدول")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 17px;
                color: #ffffff;
                font-weight: bold;
                font-size: 25px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_data)

        # Filter button with integrated icon and text
        filter_btn = QPushButton("ترشيح")
        filter_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {MEDIUM_BG};
                border: none;
                border-radius: 8px;
                color: {WHITE};
                font-weight: bold;
                font-size: 16px;  /* Increased font size */
                padding: 8px 15px;
                min-width: 80px;
            }}
            QPushButton:hover {{
                background-color: {LIGHT_BG};
            }}
        """)
        
        # Load and set filter icon
        filter_icon = QPixmap("pics/filter.png")
        if not filter_icon.isNull():
            filter_icon = filter_icon.scaled(16, 16, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            filter_btn.setIcon(QIcon(filter_icon))
            filter_btn.setIconSize(QSize(16, 16))
        
        filter_btn.setToolTip("ترشيح")
        filter_btn.clicked.connect(self.show_filter_dialog)
        
        # Add widgets to layout - right to left for Arabic
        action_layout.addWidget(filter_btn, alignment=Qt.AlignLeft)
        action_layout.addStretch()
        action_layout.addWidget(refresh_btn, alignment=Qt.AlignRight)
        
        self.main_page_layout.addWidget(action_bar)

    
    def _get_table_data_as_lists(self):
        """Helper function to extract data ONLY from VISIBLE rows in self.table."""
        if not hasattr(self, 'table') or not self.table or self.table.rowCount() == 0:
            return None, None
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        visible_row_data = []
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row): # Respecte la pagination
                row_data = [self.table.item(row, col).text() if self.table.item(row, col) else "" 
                            for col in range(self.table.columnCount())]
                visible_row_data.append(row_data)
        return headers, visible_row_data

    def export_data_to_excel(self):
        print(f"{self.__class__.__name__}: export_data_to_excel (congés année courante) called")
        headers, data = self._get_table_data_as_lists()
        if not data:
            QMessageBox.information(self, "لا بيانات للتصدير", "لا توجد عطلات للعرض في الجدول الحالي.")
            return
        filePath, _ = QFileDialog.getSaveFileName(self, "تصدير عطلات السنة الحالية إلى Excel", 
                                                  os.path.expanduser(f"~/Documents/Conges_Annee_Courante.xlsx"),
                                                  "Excel Workbook (*.xlsx);;All Files (*)")
        if not filePath: return
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "عطلات السنة الحالية"
            sheet.sheet_view.rightToLeft = True
            sheet.append(headers)
            for row_values in data: sheet.append(row_values)
            for col_idx, column_cells in enumerate(sheet.columns):
                length = max(len(str(cell.value) or "") for cell in column_cells)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = length + 5
            workbook.save(filePath)
            QMessageBox.information(self, "نجاح التصدير", f"تم تصدير البيانات بنجاح إلى:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ: {e}")

    def print_data_to_pdf(self):
        print(f"{self.__class__.__name__}: print_data_to_pdf (congés année courante) called")
        headers, data = self._get_table_data_as_lists()
        if not data:
            QMessageBox.information(self, "لا بيانات للطباعة", "لا توجد عطلات للعرض في الجدول الحالي.")
            return
        filePath, _ = QFileDialog.getSaveFileName(self, "طباعة عطلات السنة الحالية إلى PDF", 
                                                  os.path.expanduser(f"~/Documents/Conges_Annee_Courante.pdf"),
                                                  "PDF Document (*.pdf);;All Files (*)")
        if not filePath: return
        html_content = "<html><head><meta charset='UTF-8'><style>"
        html_content += "body { direction: rtl; font-family: 'Arial', 'DejaVu Sans', sans-serif; font-size: 8pt; } table { width: 100%; border-collapse: collapse; margin-top: 10px; } th, td { border: 1px solid #333; padding: 3px; text-align: right; word-wrap: break-word; } th { background-color: #f0f0f0; font-weight: bold; } caption { font-size: 1.0em; font-weight: bold; margin-bottom: 7px; text-align: center; }"
        html_content += "</style></head><body>"
        html_content += f"<table><caption>تفاصيل جدول الإجازات (السنة الحالية)</caption><thead><tr>"
        for header in headers: html_content += f"<th>{header}</th>"
        html_content += "</tr></thead><tbody>"
        for row_values in data:
            html_content += "<tr>"
            for cell_value in row_values: html_content += f"<td>{html.escape(str(cell_value))}</td>"
            html_content += "</tr>"
        html_content += "</tbody></table></body></html>"
        try:
            css_style = CSS(string='@page { size: A3 landscape; margin: 0.7cm; }') # A3 car beaucoup de colonnes
            HTML(string=html_content).write_pdf(filePath, stylesheets=[css_style])
            QMessageBox.information(self, "نجاح الطباعة", f"تم إنشاء ملف PDF بنجاح:\n{filePath}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في الطباعة", f"حدث خطأ: {e}")
            
    
    
    def refresh_data(self):
        """Refresh the table data from database and reset pagination to first page"""
        # Force refresh to ensure new employees are included
        self.force_refresh_data()
        
        # Reset any applied filters
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)
        self.selected_filter_columns = []
        
        # Reset pagination to first page with 10 rows per page
        if hasattr(self, 'paginator'):
            self.paginator.update_total_rows()
            self.paginator.update_page(1)  # Go back to first page
    def update_data(self):
        self.refresh_data()
    def create_table_title(self):
        """
        Create a title for the main table
        """
        title_container = QWidget()
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 10)
        
        # Create title label
        title_label = QLabel("تفاصيل جدول الإجازات")
        title_label.setStyleSheet(f"""
            color: {WHITE};
            font-size: 18px;
            font-weight: bold;
        """)
        title_label.setLayoutDirection(Qt.RightToLeft)
        title_layout.addWidget(title_label)
        
        self.main_page_layout.addWidget(title_container)

    def show_filter_dialog(self):
        """
        Show the filter dialog with checkboxes for column selection
        """
        # Create filter dialog if it doesn't exist
        if not self.filter_dialog:
            self.filter_dialog = QDialog(self)
            self.filter_dialog.setWindowTitle("ترشيح")
            self.filter_dialog.setFixedWidth(350)
            self.filter_dialog.setStyleSheet(f"background-color: {MEDIUM_BG}; color: {WHITE};")
            
            dialog_layout = QVBoxLayout(self.filter_dialog)
            dialog_layout.setContentsMargins(20, 20, 20, 20)
            
            # Title
            title_label = QLabel("اختر الأعمدة للترشيح:")
            title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 15px;")
            dialog_layout.addWidget(title_label)
            
            # Create checkboxes for each column
            columns = [
                "السنة", "رقم الموظف", "الإسم", "اللقب", "عدد الأيام المستحقة", 
                "الشطر 1", "الشطر 2", "الشطر 3", "الشطر 4", "الشطر 5",
                "عدد الأيام المستهلكة", "عدد الأيام المتبقية"
            ]
            
            self.filter_checkboxes = {}
            
            for column in columns:
                checkbox = QCheckBox(column)
                checkbox.setStyleSheet(f"""
                    QCheckBox {{
                        font-size: 16px;
                        padding: 8px 0;  /* Increased spacing between items */
                    }}
                    QCheckBox::indicator {{
                        width: 18px;
                        height: 18px;
                    }}
                    QCheckBox::indicator:checked {{
                        background-color: {ORANGE};
                        border: 2px solid {WHITE};
                    }}
                """)
                dialog_layout.addWidget(checkbox)
                self.filter_checkboxes[column] = checkbox
            
            # Filter value input
            filter_value_layout = QHBoxLayout()
            filter_value_layout.setSpacing(10)
            
            filter_value_label = QLabel("قيمة الترشيح:")
            filter_value_label.setStyleSheet("font-size: 16px;")
            
            self.filter_value_input = QLineEdit()
            self.filter_value_input.setPlaceholderText("أدخل قيمة للترشيح...")
            self.filter_value_input.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    padding: 10px;
                    color: {BLACK};
                    font-size: 16px;
                    text-align: right;
                }}
            """)
            self.filter_value_input.setLayoutDirection(Qt.RightToLeft)  # Set text direction to RTL
            
            filter_value_layout.addWidget(filter_value_label)
            filter_value_layout.addWidget(self.filter_value_input)
            
            dialog_layout.addSpacing(15)
            dialog_layout.addLayout(filter_value_layout)
            
            # Buttons
            buttons_layout = QHBoxLayout()
            buttons_layout.setSpacing(15)
            
            apply_btn = QPushButton("تطبيق")
            apply_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    padding: 10px;
                    font-size: 16px;
                    font-weight: bold;
                    min-width: 100px;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
            """)
            apply_btn.clicked.connect(self.apply_filter)
            
            cancel_btn = QPushButton("إلغاء")
            cancel_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {MEDIUM_BG};
                    color: {WHITE};
                    border: 1px solid {LIGHT_BG};
                    border-radius: 5px;
                    padding: 10px;
                    font-size: 16px;
                    min-width: 100px;
                }}
                QPushButton:hover {{
                    background-color: {LIGHT_BG};
                }}
            """)
            cancel_btn.clicked.connect(self.filter_dialog.reject)
            
            # Add buttons in right-to-left order for Arabic
            buttons_layout.addWidget(cancel_btn)
            buttons_layout.addWidget(apply_btn)
            
            dialog_layout.addSpacing(15)
            dialog_layout.addLayout(buttons_layout)
        
        # Show the dialog
        self.filter_dialog.exec_()

    def apply_filter(self):
        """
        Apply the filter to the table based on selected columns and input value
        """
        # Get selected columns
        self.selected_filter_columns = []
        for column, checkbox in self.filter_checkboxes.items():
            if checkbox.isChecked():
                self.selected_filter_columns.append(column)
        
        # Get filter value
        filter_value = self.filter_value_input.text().strip().lower()
        
        if not filter_value or not self.selected_filter_columns:
            # Reset filter if no value or no columns selected
            for row in range(self.table.rowCount()):
                self.table.setRowHidden(row, False)
            self.filter_dialog.accept()
            return
        
        # Get column indices for selected columns
        column_indices = []
        for col_name in self.selected_filter_columns:
            for col_idx in range(self.table.columnCount()):
                if self.table.horizontalHeaderItem(col_idx).text() == col_name:
                    column_indices.append(col_idx)
        
        # Apply filter
        for row in range(self.table.rowCount()):
            row_visible = False
            
            # Check if any selected column contains the filter value
            for col_idx in column_indices:
                cell_value = self.table.item(row, col_idx).text().lower()
                if filter_value in cell_value:
                    row_visible = True
                    break
            
            self.table.setRowHidden(row, not row_visible)
        
        # Close the dialog
        self.filter_dialog.accept()

    def create_table(self):
        """
        Create the main data table with all columns
        """
        # Table container
        table_container = QWidget()
        table_container.setStyleSheet(f"background-color: {MEDIUM_BG}; border-radius: 10px;")
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(15, 15, 15, 15)  # Increased margins
        
        # Create table
        self.table = QTableWidget()
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                border: none;
                gridline-color: {LIGHT_BG};
                font-size: 16px;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {LIGHT_BG};
                color: {WHITE};
                border: none;
                font-weight: bold;
                font-size: 16px;
                text-align: right;
            }}
            QTableWidget::item:selected {{
                background-color: {ORANGE};
                color: {WHITE};
            }}
            QHeaderView::section {{
                background-color: {DARKER_BG};
                color: {WHITE};
                padding: 10px;
                border: none;
                font-weight: bold;
                font-size: 16px;
                text-align: right;
            }}
            QScrollBar:vertical {{
                background: {MEDIUM_BG};
                width: 0px;
                margin: 0px;
            }}
            
           """)
    
        
        # Set up columns - UPDATED ORDER with 5th tranche
        columns = [
            "رقم العطلة", "السنة", "رقم الموظف", "لقب الموظف", "اسم الموظف", 
            "عدد الأيام المستحقة", "الشطر 1", "الشطر 2", "الشطر 3", "الشطر 4", "الشطر 5",
            "عدد الأيام المستهلكة", "عدد الأيام المتبقية"
        ]
        
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # Make entitled days column editable, others read-only
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        # Set column widths
        for i in range(len(columns)):
            self.table.setColumnWidth(i, 120)
        
        # Set text alignment for all columns to center-aligned
        for i in range(len(columns)):
            self.table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignCenter)
        
        # To move the columns
        self.table.horizontalHeader().setSectionsMovable(True)
        #Set column resizing fixed
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)

        
        # Connect double-click signal to edit entitled days
        self.table.cellDoubleClicked.connect(self.handle_cell_double_click)
        
        # Add table to layout
        table_layout.addWidget(self.table)
        self.main_page_layout.addWidget(table_container)
        #Ajouter le paginator juste en dessous du tableau
        self.paginator = tablepaginator(self.table, rows_per_page=10)
        table_layout.addWidget(self.paginator)

    def handle_cell_double_click(self, row, column):
        """
        Handle double-click on table cells
        Only allow editing of the entitled days column (index 5)
        """
        if column == 5:  # عدد الأيام المستحقة column
            # Create a temporary line edit for in-place editing
            current_value = self.table.item(row, column).text()
            line_edit = QLineEdit(current_value)
            line_edit.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {WHITE};
                    color: {BLACK};
                    border: none;
                    padding: 2px;
                }}
            """)
            
            # Connect editing finished signal
            line_edit.editingFinished.connect(lambda: self.finish_entitled_days_edit(row, column, line_edit))
            
            # Set the line edit as the cell widget
            self.table.setCellWidget(row, column, line_edit)
            line_edit.setFocus()
            line_edit.selectAll()

    def finish_entitled_days_edit(self, row, column, line_edit):
        """
        Finish editing entitled days and update related fields
        """
        try:
            # Get the new value
            new_value = line_edit.text()
            
            # Validate that it's a number
            if not new_value.isdigit():
                warning = CustomWarningDialog(self, "تحذير", "يجب أن يكون عدد الأيام المستحقة رقمًا")
                warning.exec_()
                new_value = "0"
            
            # Update the cell
            self.table.removeCellWidget(row, column)
            self.table.setItem(row, column, QTableWidgetItem(new_value))
            
            # Center align the new item
            item = self.table.item(row, column)
            if item:
                item.setTextAlignment(Qt.AlignCenter)
            
            # Get conge ID
            conge_id = int(self.table.item(row, 0).text())
            
            # Update in database
            self.conge_controller.update_allocated_days(conge_id, int(new_value))
            
            # Update remaining days
            self.update_leave_days(row)
            
            # Log the action
            employee_id = self.table.item(row, 2).text()
            self.log_action("تعديل", f"تعديل عدد الأيام المستحقة للموظف رقم {employee_id}")
            
        except Exception as e:
            CustomWarningDialog(self, "خطأ", str(e)).exec_()

    def update_leave_days(self, row):
        """
        Update consumed and remaining days based on tranches
        """
        # Get entitled days
        entitled_item = self.table.item(row, 5)
        entitled_days = int(entitled_item.text()) if entitled_item and entitled_item.text().isdigit() else 0
        
        # Calculate consumed days from tranches
        consumed_days = 0
        for i in range(5):  # Now 5 tranches
            tranche_col = 6 + i  # Tranche columns start at index 6
            tranche_item = self.table.item(row, tranche_col)
            if tranche_item and tranche_item.text().isdigit():
                consumed_days += int(tranche_item.text())
        
        # Update consumed days (column 11)
        consumed_item = QTableWidgetItem(str(consumed_days))
        consumed_item.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(row, 11, consumed_item)
        
        # Update remaining days (column 12)
        remaining_days = entitled_days - consumed_days
        remaining_item = QTableWidgetItem(str(remaining_days))
        remaining_item.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(row, 12, remaining_item)

    def create_action_buttons(self):
        """
        Create action buttons for view details, previous years, and activity log
        All buttons have the same size
        """
        # Buttons container
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 20, 0, 0)  # Increased top margin
        buttons_layout.setSpacing(25)  # Increased spacing
        
        # Create buttons - REMOVED DELETE BUTTON as requested
        add_btn = QPushButton("إضافة")
        view_details_btn = QPushButton("عرض التفاصيل")
        previous_years_btn = QPushButton("عرض تفاصيل السنوات السابقة")
        history_btn = QPushButton("سجل الأنشطة")
        
        # Style buttons - all with the same fixed size
        for btn in [add_btn, view_details_btn, previous_years_btn, history_btn]:
            btn.setFixedSize(160, 45)  # Same fixed size for all buttons
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
                QPushButton:disabled {{
                    background-color: #a0a0a0;
                    color: #e0e0e0;
                }}
            """)
        
        # Connect buttons to actions
        add_btn.clicked.connect(self.show_add_leave_form)
        view_details_btn.clicked.connect(self.show_leave_details)
        previous_years_btn.clicked.connect(self.show_previous_years)
        history_btn.clicked.connect(self.show_history)
        
        # Add buttons to layout - right to left for Arabic
        buttons_layout.addWidget(add_btn)
        buttons_layout.addWidget(view_details_btn)
        buttons_layout.addWidget(previous_years_btn)
        buttons_layout.addWidget(history_btn)
        
        self.main_page_layout.addWidget(buttons_widget)

    def validate_tranche_dates(self, start_date, end_date, leave_id, exclude_tranche_id=None):
        """
        Validate tranche dates for overlaps and year constraints
        """
        current_year = datetime.now().year
        
        # Convert string dates to datetime objects if needed
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        
        # Check if dates are within current year
        if start_date.year != current_year or end_date.year != current_year:
            raise ValueError(f"التواريخ يجب أن تكون ضمن السنة الحالية ({current_year})")
        
        # Check if start date is before end date
        if start_date > end_date:
            raise ValueError("تاريخ البداية يجب أن يكون قبل تاريخ النهاية")
        
        # Check for overlaps with existing tranches in database
        try:
            conge_id = int(leave_id)
            existing_tranches = self.tranche_controller.get_tranches_by_conge(conge_id)
            
            for tranche in existing_tranches:
                # Skip the tranche being edited
                if exclude_tranche_id is not None and tranche.idTranche == exclude_tranche_id:
                    continue
                    
                existing_start = tranche.DateDebut
                existing_end = tranche.DateFin
                
                # Check for overlap
                if (start_date <= existing_end and end_date >= existing_start):
                    raise ValueError(f"التواريخ تتداخل مع شطر موجود (القرار رقم: {tranche.NumeroDecision})")
        except Exception as e:
            if "التواريخ تتداخل" in str(e):
                raise e
            # If there's an error getting tranches, continue without overlap check
            pass
        
        return True

    def show_add_leave_form(self):
        """
        Show the add leave form dialog with enhanced validation
        """
        # Check if a row is selected
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            warning = CustomWarningDialog(self, "تحذير", "الرجاء اختيار موظف لإضافة عطلة")
            warning.exec_()
            return
            
        # Get selected row data
        row = selected_rows[0].row()
        
        # Get employee data
        employee_id = self.table.item(row, 2).text()
        employee_name = f"{self.table.item(row, 4).text()} {self.table.item(row, 3).text()}"
        
        # Check if employee has already used all 5 tranches
        leave_id = self.table.item(row, 0).text()
        tranches_used = 0
        for i in range(5):
            tranche_col = 6 + i
            tranche_value = self.table.item(row, tranche_col).text()
            if tranche_value != "0":
                tranches_used += 1
        
        if tranches_used >= 5:
            warning = CustomWarningDialog(self, "تحذير", "لا يمكن إضافة عطلة جديدة. تم استخدام جميع الأشطر الخمسة المتاحة.")
            warning.exec_()
            return
        
        # Store the selected row for later use
        self.selected_row = row
        
        # Create and show add leave dialog
        add_dialog = AddLeaveDialog(self, employee_id, employee_name,session=self.session)
        if add_dialog.exec_() == QDialog.Accepted:
            try:
                # Get leave data
                leave_data = add_dialog.leave_data
                
                # Validate dates before saving
                self.validate_tranche_dates(
                    leave_data["start_date"], 
                    leave_data["end_date"], 
                    leave_id
                )
                
                # Get conge ID
                conge_id = int(leave_id)
                
                # Convert dates to datetime objects
                decision_date = datetime.strptime(leave_data["decision_date"], "%Y-%m-%d").date()
                start_date = datetime.strptime(leave_data["start_date"], "%Y-%m-%d").date()
                end_date = datetime.strptime(leave_data["end_date"], "%Y-%m-%d").date()
                
                # Create tranche in database
                result = self.tranche_controller.create_tranche(
                    conge_id=conge_id,
                    numero_decision=int(leave_data["decision_id"]),
                    date_decision=decision_date,
                    date_debut=start_date,
                    date_fin=end_date
                )
                self.tranche_controller.log_history(
                event="إضافة عطلة",
                details=f"تمت إضافة عطلة جديدة للموظف رقم {employee_id} للسنة {start_date.year} من {start_date} إلى {end_date} (قرار رقم: {leave_data['decision_id']})",
                gestion="إدارة الإجازات",
                conge_id=conge_id,
                employee_id=employee_id
               )
                
                # Refresh the tranche display
                self.refresh_tranche_display(row, leave_id)
                
                # Update consumed and remaining days
                self.update_leave_days(row)
                
                # Log the action
                self.log_action("إضافة", f"إضافة عطلة جديدة للموظف رقم {employee_id}")
                
            except ValueError as ve:
                CustomWarningDialog(self, "تحذير", str(ve)).exec_()
            except Exception as e:
                CustomWarningDialog(self, "خطأ", f"خطأ في إضافة العطلة: {str(e)}").exec_()

    def show_history(self):
        """
        Show the activity history dialog.
        """
        try:

            from Views.Historique import HistoryDialog  # adapte le chemin si besoin
            print(f"DEBUG - current_user_data dans show_history: {self.current_user_data}")
            history_dialog = HistoryDialog(
                parent=self,
                current_user_data=self.current_user_data,
                session=self.session,
                module_name="إدارة الإجازات",
                gestion_filter= "إدارة الإجازات"
            )
            history_dialog.show()
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء عرض السجل:\n{str(e)}")
    
    
    def show_leave_details(self):

        """
        Show the leave details dialog
    
      """
        # DEBUG: Vérifications complètes
        print(f"DEBUG - Attributs de self: {dir(self)}")
        print(f"DEBUG - hasattr current_user_data: {hasattr(self, 'current_user_data')}")
        if hasattr(self, 'current_user_data'):
          print(f"DEBUG - current_user_data: {self.current_user_data}")
        else:
          print("DEBUG - current_user_data n'existe pas!")
     
        # Check if a row is selected
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            warning = CustomWarningDialog(self, "تحذير", "الرجاء اختيار عطلة لعرض التفاصيل")
            warning.exec_()
            return
        
        # Get selected row data
        row = selected_rows[0].row()
        
        # Get leave ID
        leave_id = self.table.item(row, 0).text()
        
        # Get employee data
        employee_data = {
            "employee_id": self.table.item(row, 2).text(),
            "last_name": self.table.item(row, 3).text(),
            "first_name": self.table.item(row, 4).text()
        }
        
        try:
            # Get tranches data from database
            conge_with_tranches = self.conge_controller.get_conge_with_tranches(int(leave_id))
            tranches = conge_with_tranches["tranches"]
            
            # Store tranches data
            self.leave_tranches[leave_id] = tranches
            
            # Create and show leave details dialog
            details_dialog = LeaveDetailsDialog(self, leave_id, employee_data, tranches)
            details_dialog.exec_()
        except Exception as e:
            CustomWarningDialog(self, "خطأ", str(e)).exec_()

    def show_previous_years(self):
        """
        Show the previous years dialog with pagination
        """
        try:
            # Get current year
            current_year = datetime.now().year
            
            # Get previous years data from database
            previous_conges = self.conge_controller.get_previous_years_conges(current_year)
            
            # Format data for display
            previous_years_data = []
            
            for conge in previous_conges:
                # Get tranches for this conge
                tranches = self.tranche_controller.get_tranches_by_conge(conge.idConge)
                
                # Format tranche data
                tranche_values = ["0", "0", "0", "0", "0"]  # Default 5 empty tranches
                
                for i, tranche in enumerate(tranches[:5]):  # Limit to 5 tranches
                    days = (tranche.DateFin - tranche.DateDebut).days + 1
                    tranche_values[i] = str(days)
                
                # Add row data
                row_data = [
                    str(conge.idConge),
                    str(conge.Annee),
                    str(conge.idemploye),
                    conge.employe.Prenom if conge.employe else "",
                    conge.employe.Nom if conge.employe else "",
                    str(conge.NbrJoursAlloues),
                    tranche_values[0],
                    tranche_values[1],
                    tranche_values[2],
                    tranche_values[3],
                    tranche_values[4],
                    str(conge.NbrJoursPris),
                    str(conge.NbrJoursRestants)
                ]
                
                previous_years_data.append(row_data)
            
            # Create and show previous years dialog
            previous_years_dialog = PreviousYearsDialog(self, previous_years_data)
            previous_years_dialog.exec_()
        except Exception as e:
            CustomWarningDialog(self, "خطأ", str(e)).exec_()


    def update_tranche_by_id(self, tranche_id, tranche_data):
   
        try:
            # Get the selected leave ID from the main table
            selected_rows = self.table.selectionModel().selectedRows()
            if not selected_rows:
                CustomWarningDialog(self, "تحذير", "الرجاء تحديد عطلة قبل التعديل").exec_()
                return
            
            row = selected_rows[0].row()
            leave_id = int(self.table.item(row, 0).text())  # ✅ use local variable

            # Validate dates before updating
            self.validate_tranche_dates(
                tranche_data["start_date"], 
                tranche_data["end_date"], 
                leave_id,
                exclude_tranche_id=tranche_id
            )

            # Convert dates to datetime objects
            decision_date = datetime.strptime(tranche_data["decision_date"], "%Y-%m-%d").date()
            start_date = datetime.strptime(tranche_data["start_date"], "%Y-%m-%d").date()
            end_date = datetime.strptime(tranche_data["end_date"], "%Y-%m-%d").date()
            
            # Update tranche in database
            result = self.tranche_controller.update_tranche(
                tranche_id=tranche_id,
                numero_decision=int(tranche_data["decision_id"]),
                date_decision=decision_date,
                date_debut=start_date,
                date_fin=end_date
            )

            # ✅ Force reload of updated data from DB
            self.refresh_tranche_display(row, leave_id)
            self.update_leave_days(row)

            # Log the action
            self.log_action("تعديل", f"تعديل شطر للعطلة رقم {leave_id}")

        except ValueError as ve:
            CustomWarningDialog(self, "تحذير", str(ve)).exec_()
        except Exception as e:
            CustomWarningDialog(self, "خطأ", f"خطأ في تعديل الشطر: {str(e)}").exec_()


    def delete_tranche_by_id(self, tranche_id):
        """
        Delete a tranche using its database ID instead of table index
        """
        try:
            # Delete tranche from database using the controller
            result = self.tranche_controller.delete_tranche(tranche_id)
            
            # Get the selected row in the main table
            selected_rows = self.table.selectionModel().selectedRows()
            if selected_rows:
                row = selected_rows[0].row()
                leave_id = self.table.item(row, 0).text()
                
                # Refresh the tranche display
                self.refresh_tranche_display(row, leave_id)
                
                # Update consumed and remaining days
                self.update_leave_days(row)
                
                # Log the action
                self.log_action("حذف", f"حذف شطر للعطلة رقم {leave_id}")
                
        except Exception as e:
            CustomWarningDialog(self, "خطأ", f"خطأ في حذف الشطر: {str(e)}").exec_()

    def refresh_tranche_display(self, row, leave_id):
        """
        Refresh the tranche display in the main table after deletion or update
        """
        try:
            # Get updated tranches from database
            conge_with_tranches = self.conge_controller.get_conge_with_tranches(int(leave_id))
            tranches = conge_with_tranches["tranches"]
            
            # Reset all tranche columns to 0
            for i in range(5):
                tranche_col = 6 + i
                empty_item = QTableWidgetItem("0")
                empty_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, tranche_col, empty_item)
            
            # Fill with actual tranche data
            for i, tranche in enumerate(tranches[:5]):  # Limit to 5 tranches
                tranche_col = 6 + i
                days = tranche.get("days", 0)
                tranche_item = QTableWidgetItem(str(days))
                tranche_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, tranche_col, tranche_item)
            
            # Update stored tranches data
            self.leave_tranches[leave_id] = tranches
            
        except Exception as e:
            print(f"Error refreshing tranche display: {e}")

    def log_action(self, action_type, details):
        """
        Log user actions for auditing purposes
        """
        # Create timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        user = "اسم المستخدم"  # This would come from the logged-in user
        
        # Add to history data
        self.history_data.insert(0, [timestamp, user, action_type, details])
        
        # Print to console for debugging
        print(f"Action logged: {timestamp} | {user} | {action_type} | {details}")

    def load_data_from_database(self):
        session = self.session
        try:
            current_year = datetime.now().year
            conges = self.conge_controller.get_all_conges(current_year)
            employe_controller = EmployeeController(self.session)

            # 🚨 Get list of definitively departed employee IDs
            departed_ids = set(depart.idemploye for depart in employe_controller.get_final_departures())

            self.table.setRowCount(0)

            for conge in conges:
                if conge.idemploye in departed_ids:
                    continue  # Skip archived employees

                tranches = self.tranche_controller.get_tranches_by_conge(conge.idConge)
                tranche_values = ["0", "0", "0", "0", "0"]

                for i, tranche in enumerate(tranches[:5]):
                    days = (tranche.DateFin - tranche.DateDebut).days + 1
                    tranche_values[i] = str(days)

                row_idx = self.table.rowCount()
                self.table.insertRow(row_idx)

                self.table.setItem(row_idx, 0, QTableWidgetItem(str(conge.idConge)))
                self.table.setItem(row_idx, 1, QTableWidgetItem(str(conge.Annee)))
                self.table.setItem(row_idx, 2, QTableWidgetItem(str(conge.idemploye)))
                self.table.setItem(row_idx, 3, QTableWidgetItem(conge.employe.Nom if conge.employe else ""))
                self.table.setItem(row_idx, 4, QTableWidgetItem(conge.employe.Prenom if conge.employe else ""))
                self.table.setItem(row_idx, 5, QTableWidgetItem(str(conge.NbrJoursAlloues)))
                self.table.setItem(row_idx, 6, QTableWidgetItem(tranche_values[0]))
                self.table.setItem(row_idx, 7, QTableWidgetItem(tranche_values[1]))
                self.table.setItem(row_idx, 8, QTableWidgetItem(tranche_values[2]))
                self.table.setItem(row_idx, 9, QTableWidgetItem(tranche_values[3]))
                self.table.setItem(row_idx, 10, QTableWidgetItem(tranche_values[4]))
                self.table.setItem(row_idx, 11, QTableWidgetItem(str(conge.NbrJoursPris)))
                self.table.setItem(row_idx, 12, QTableWidgetItem(str(conge.NbrJoursRestants)))

                for col in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col)
                    if item:
                        item.setTextAlignment(Qt.AlignCenter)

            self.paginator.update_total_rows()
            self.paginator.update_page(1)

        except Exception as e:
            CustomWarningDialog(self, "خطأ", f"خطأ في تحميل البيانات: {str(e)}").exec_()


    def migrate_previous_year_data(self):
        """
        Migrate previous year's data to historical table (called at year change)
        """
        try:
            current_year = datetime.now().year
            previous_year = current_year - 1
            
            # This would typically move data to a historical table
            # For now, we'll just ensure the data exists in the previous years view
            previous_conges = self.conge_controller.get_all_conges(previous_year)
            
            # Log the migration
            self.log_action("ترحيل", f"ترحيل بيانات السنة {previous_year} إلى الأرشيف")
            
            return len(previous_conges)
            
        except Exception as e:
            CustomWarningDialog(self, "خطأ", f"خطأ في ترحيل البيانات: {str(e)}").exec_()
            return 0


if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # Set application-wide font for Arabic support
    font = QFont("Arial", 10)
    app.setFont(font)
    
    # Set right-to-left layout for Arabic
    app.setLayoutDirection(Qt.RightToLeft)
    
    window = LeaveManagementSystem()
    window.show()
    
    sys.exit(app.exec_())
