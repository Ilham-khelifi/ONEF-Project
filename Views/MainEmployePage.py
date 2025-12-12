import sys
import os
import tempfile
import traceback
import warnings
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
# Suppress the DeprecationWarning about sipPyTypeDict
warnings.filterwarnings("ignore", category=DeprecationWarning, module="PyQt5.sip")

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QPushButton, QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
                             QHeaderView, QFrame, QToolButton, QScrollArea,QSizePolicy)
from PyQt5.QtGui import QIcon, QPixmap, QFont
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QMessageBox, QFileDialog
import openpyxl
from openpyxl.styles import Font, Alignment
from jinja2 import Environment, FileSystemLoader
from weasyprint import HTML as WeasyHTML 
from weasyprint.text.fonts import FontConfiguration
from datetime import datetime
# Import the sidebar function from the sidebar module
from sidebar import create_sidebar
from ui_constants import *

from AddEmploye import AddEmployeeWindow
from EmployeDetails import EmployeeDetailsWindow
from Filter import FilterWindow
from FilterTableColumns import FilterTableColumnsWindow
from Controllers.EmployeController import EmployeeController  # Updated import
from DatabaseConnection import db
from temporaryDepartureForm import TemporaryDepartureForm
from finalDepartureForm import FinalDepartureForm
from Models.Employe import Employe
from Models.Depart import Depart
from Models.DepartTemporaire import DepartTemporaire
from statistics import StatisticsSidebar

class MainEmployeeWindow(QMainWindow):
    def __init__(self, session, current_user_data=None, external_sidebar=None, external_sidebar_toggle=None, logout_btn=None):
        super().__init__()
        if session is not None:
            self.session = session
        else:
            # Fallback si aucune session n'est fournie (pour les tests)
            from DatabaseConnection import db
            self.session = db.get_session()
            
        # Store current user data
        self.current_user_data = current_user_data or {}
        current_user_account_number = self.current_user_data.get('account_number')
        
        # Initialize controller with session and current user account number
        self.controller = EmployeeController(self.session, current_user_account_number)
        
        self.setWindowTitle("نظام إدارة الموظفين")
        self.setGeometry(100, 100, 1400, 800)
        self.sidebar_visible = True
        
        # Set the application style to be dark
        self.setStyleSheet(f"""
            QMainWindow, QWidget {{
                background-color: {DARK_BG};
                color: {WHITE};
            }}
            QPushButton {{
                background-color: {ORANGE};
                color: {WHITE};
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #d35400;
            }}
            QLineEdit {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                border: 1px solid {LIGHT_BG};
                border-radius: 4px;
                padding: 5px;
                font-size: 16px;
            }}
            QHeaderView::section {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                padding: 5px;
                border: 1px solid {LIGHT_BG};
            }}
            QToolButton {{
                background-color: transparent;
                border: none;
                color: {WHITE};
                padding: 5px;
            }}
            QToolButton:hover {{
                background-color: {LIGHT_BG};
            }}
            QFrame#sidebar {{
                background-color: {DARKER_BG};
                border-right: 1px solid {LIGHT_BG};
            }}
            QFrame#stats_sidebar {{
                background-color: {DARKER_BG};
                border-left: 1px solid {LIGHT_BG};
            }}
            QLabel#menuItem {{
                padding: 10px;
                font-size: 16px;
                font-weight: bold;
            }}
            QLabel#menuItemSelected {{
                padding: 10px;
                font-size: 16px;
                font-weight: bold;
                background-color: {ORANGE};
                border-radius: 4px;
            }}
            
            QPushButton#logoutBtn {{
                background-color: #ffc107;
                color: #333;
                font-weight: bold;
                border-radius: 4px;
                padding: 10px;
            }}
            QLabel#username {{
                font-size: 16px;
                font-weight: bold;
            }}
            QLabel#userrole {{
                font-size: 12px;
                color: #bdc3c7;
            }}
            QWidget#menuItemWidget {{
                padding: 5px 15px;
                margin: 2px 10px;
            }}
            QWidget#menuItemWidgetSelected {{
                padding: 5px 15px;
                margin: 2px 10px;
                background-color: {ORANGE};
                border-radius: 4px;
            }}
            QWidget#menuItemWidget:hover {{
                background-color: {LIGHT_BG};
                border-radius: 4px;
            }}
            QScrollBar:horizontal {{
                width: 0px;
                height: 0px;
        }}
            QScrollBar::handle:horizontal {{
                background: #607D8B;
                min-width: 30px;
                border-radius: 7px;
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                border: none;
                background: none;
            }}
            QWidget#searchContainer {{
                background-color: {MEDIUM_BG};
                border-radius: 4px;
                padding: 5px;
            }}
        """)

        # Create central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Main layout
        self.main_layout = QHBoxLayout(central_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)
        
        self.sidebar_toggle = external_sidebar_toggle
        self.logout_btn = logout_btn
        # DO NOT add external_sidebar to self.main_layout
        self.sidebar = external_sidebar  # Store the external sidebar
        self.external_sidebar_mode = True  # Flag to indicate we're using external sidebar
         # Start hidden
        # Create sidebar using the imported function
        # Define callback functions for the sidebar
        def on_module_change(module_name):
            print(f"Module changed to: {module_name}")
            # Here you would handle module changes
        
        def on_logout():
            print("Logout clicked")
            # Here you would handle logout

        # Content area
        self.content_area = QWidget()
        content_layout = QVBoxLayout(self.content_area)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)
        self.main_layout.addWidget(self.content_area)
        
        self.main_layout.addSpacing(15)  # Add 15px spacing between content and sidebar
        
        # Top toolbar
        toolbar = QWidget()
        toolbar.setFixedHeight(75)
        toolbar.setStyleSheet("background-color: {DARK_BG}; border-bottom: 1px solid {LIGHT_BG};")
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(20, 15, 20, 15)  # Increased margins
        toolbar_layout.setSpacing(15)
        # Left toolbar buttons (appears on right in RTL)
        left_buttons = QWidget()
        left_buttons.setStyleSheet(f"""
            QWidget {{
                background-color: {DARK_BG};
                border-bottom: 1px solid {LIGHT_BG};
            }}
        """)
        left_buttons_layout = QHBoxLayout(left_buttons)
        left_buttons_layout.setContentsMargins(0, 0, 0, 0)

        self.menu_btn = QToolButton()
        self.menu_btn.setIcon(QIcon("pics/sidebar_icon.png"))
        self.menu_btn.setIconSize(QSize(20, 20))
        self.menu_btn.setFixedSize(35, 35)
        self.menu_btn.setStyleSheet(f"""
            QToolButton {{
                background-color: {LIGHT_BG};  /* Background color like hover */
                border: none;
                color: {WHITE};
                border-radius: 17px;  /* Circular background */
                font-size: 16px;
            }}
            QToolButton:hover {{
                background-color: {DARK_BG};
            }}
        """)
        
    
        # Connect to external sidebar toggle if available
        if self.sidebar_toggle:
            self.menu_btn.clicked.connect(self.sidebar_toggle.click)
        else:
            self.menu_btn.clicked.connect(self.toggle_sidebar)

        # Import button
        import_btn = QToolButton()
        import_btn.setFixedSize(45, 45)  # Increased size
        import_btn.setStyleSheet(f"""
        QToolButton {{
            background-color: transparent;
            border: none;
            color: {WHITE};
            font-size: 16px;
        }}
        QToolButton:hover {{
            background-color: {LIGHT_BG};
            border-radius: 22px;
        }}
    """)
        export_icon = QPixmap("pics/export.png")
        if not export_icon.isNull():
            export_icon = export_icon.scaled(25, 25, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            import_btn.setIcon(QIcon(export_icon))
            import_btn.setIconSize(QSize(25, 25))
        else:
            import_btn.setText("↓")
            print("Failed to load export icon: pics/export.png")
    
        import_btn.setToolTip("تصدير")
        
        import_btn.clicked.connect(self.export_to_excel)
        
        
        # Print button
        print_btn = QToolButton()
        print_btn.setFixedSize(45, 45)  # Increased size
        print_btn.setStyleSheet(f"""
        QToolButton {{
            background-color: transparent;
            border: none;
            color: {WHITE};
            font-size: 16px;
        }}
        QToolButton:hover {{
            background-color: {LIGHT_BG};
            border-radius: 22px;
        }}
    """)
    
        # Load print icon from pics folder
        print_icon = QPixmap("pics/imprime.png")
        if not print_icon.isNull():
            print_icon = print_icon.scaled(25, 25, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            print_btn.setIcon(QIcon(print_icon))
            print_btn.setIconSize(QSize(25, 25))
        else:
            print_btn.setText("🖨️")
            print("Failed to load print icon: pics/imprime.png")

        print_btn.setToolTip("طباعة")
        print_btn.clicked.connect(self.generate_current_table_as_pdf_action)
        left_buttons_layout.addWidget(self.menu_btn)
        left_buttons_layout.addWidget(import_btn)
        left_buttons_layout.addWidget(print_btn)
        

        # Right toolbar (appears on left in RTL)
        right_buttons = QWidget()
        right_buttons.setStyleSheet(f"""
            QWidget {{
                background-color: {DARK_BG};
                border-bottom: 1px solid {LIGHT_BG};
            }}
        """)
        right_buttons_layout = QHBoxLayout(right_buttons)
        right_buttons_layout.setContentsMargins(0, 0, 0, 0)
        right_buttons.setFixedWidth(60)
        
        # Stats sidebar toggle button
        self.stats_sidebar_widget = StatisticsSidebar(session=self.session, parent=self)
        
        self.stats_btn = QToolButton()
        self.stats_btn.setFixedSize(45, 45)
        self.stats_btn.setStyleSheet(f"""
        QToolButton {{
            background-color: transparent;
            border: none;
            color: {WHITE};
            font-size: 16px;
        }}
        QToolButton:hover {{
            background-color: {LIGHT_BG};
            border-radius: 22px;
        }}
    """)
        stats_icon = QPixmap("pics/stats_icon.png")                             
        self.stats_sidebar_widget.hide()
        if not export_icon.isNull():
            export_icon = export_icon.scaled(25, 25, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.stats_btn.setIcon(QIcon(stats_icon))
            self.stats_btn.setIconSize(QSize(25, 25))
        else:
            self.stats_btn.setText("↓")
            print("Failed to load export icon: pics/export.png")
        self.stats_btn.clicked.connect(self.toggle_stats_sidebar)

        right_buttons_layout.addWidget(self.stats_btn)
     

        # Improved search bar with original dark colors
        search_container = QWidget()
        search_container.setObjectName("searchContainer")
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(8, 8, 8, 8)
        search_layout.setSpacing(0)
        search_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        search_container.setMinimumWidth(500)  # Minimum width instead of fixed
      # Maximum width to prevent it from getting too large
  # Add this line for preferred size

        self.search_box = QLineEdit()
        
        self.search_box.setFixedHeight(30)
        self.search_box.setPlaceholderText("بحث...")
        self.search_box.setStyleSheet(f"""
            QLineEdit {{
                background-color: #e3e3e3;
                border: none;
                border-radius: 5px;
                padding: 10px 15px;
                font-size: 14px;
                color: {BLACK};
                text-align: right;
            }}
        """)

        self.search_box.setFixedHeight(40)

        search_btn = QPushButton("بحث")
        search_btn.setFixedWidth(80)
        search_btn.setFixedHeight(30)
        search_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                border: none;
                border-radius: 5px;
                padding: 10px 15px;
                font-size: 14px;
                text-align: center;
            }}
            QPushButton:hover {{
                background-color: {LIGHT_BG};
            }}
        """)

        search_btn.clicked.connect(self.search_employees)

        search_layout.addWidget(self.search_box)
        search_layout.addWidget(search_btn)

        # Add to toolbar with search bar centered
        toolbar_layout.addWidget(left_buttons, 0)
        toolbar_layout.addStretch(1) # Add 50px spacing
        toolbar_layout.addWidget(search_container, 0)
        toolbar_layout.addStretch(1)  # Add 50px spacing
        toolbar_layout.addWidget(right_buttons, 0)
        
        # Filter section - left aligned
        filter_section = QWidget()
        filter_section_layout = QHBoxLayout(filter_section)
        filter_section_layout.setContentsMargins(15, 10, 15, 10)
        
        filter_text = QPushButton()
        filter_text.setText("ترشيح")
        filter_text.setStyleSheet(f"""
            QPushButton {{
                background-color: {MEDIUM_BG};
                border: none;
                border-radius: 8px;
                color: {WHITE};
                font-weight: bold;
                font-size: 16px;
                padding: 8px 15px;
                min-width: 80px;
            }}""")
        filter_text.setIcon(QIcon("pics/filter icon.png"))
        filter_text.clicked.connect(self.open_filter_window)

        filter_columns_text = QPushButton()
        filter_columns_text.setText("ترشيح أعمدة الجدول")
        filter_columns_text.setStyleSheet(f"""
            QPushButton {{
                background-color: {MEDIUM_BG};
                border: none;
                border-radius: 8px;
                color: {WHITE};
                font-weight: bold;
                font-size: 16px;
                padding: 8px 15px;
                min-width: 80px;
            }}""")
        filter_columns_text.setIcon(QIcon("pics/filter.png"))
        filter_columns_text.clicked.connect(self.open_filter_columns_window)
        
        # Refresh button as requested
        refresh_btn = QPushButton("⟳")
        refresh_btn.setFixedSize(42, 42)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #37474f;
                border: none;
                border-radius: 20px;
                color: #ffffff;
                font-weight: bold;
                font-size: 17px;
            }
            QPushButton:hover {
                background-color: #455a64;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_data)
        filter_section_layout.addWidget(filter_text)
        filter_section_layout.addWidget(filter_columns_text)
        filter_section_layout.addWidget(refresh_btn)  
        filter_section_layout.addStretch()  # Push buttons to left

        # Table title
        table_title = QLabel(" جدول الموظفين")
        table_title.setStyleSheet(f"""
            QLabel {{
                color: {WHITE};
                font-size: 24px;
                font-weight: bold;
            }}
        """)
        table_title.setAlignment(Qt.AlignLeft)

        # Create a scroll area for the table
        table_scroll_area = QScrollArea()
        table_scroll_area.setWidgetResizable(True)
        table_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Table container
        table_container = QWidget()
        table_container.setStyleSheet(f"""
            background-color: {MEDIUM_BG};
            border-radius: 10px;
        """)        
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(15, 15, 15, 15)  # Increased margins
        table_scroll_area.setFrameShape(QFrame.NoFrame)  # no frame that specify the scroll frame

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(19)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {MEDIUM_BG};
                color: {WHITE};
                border: none;
                gridline-color: {LIGHT_BG};
                font-size: 16px;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {LIGHT_BG};
                color: {WHITE};
                border: none;
                font-weight: bold;
                font-size: 16px;
                text-align: right;
            }}
            QTableWidget::item:selected {{
                background-color: {ORANGE};
                color: {WHITE};
            }}
            QHeaderView::section {{
                background-color: {DARKER_BG};
                color: {WHITE};
                padding: 10px;
                border: none;
                font-weight: bold;
                font-size: 16px;
                text-align: right;
            }}
            QScrollBar:vertical {{
                background: {MEDIUM_BG};
                width: 0px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {LIGHT_BG};
                min-height: 0px;
                border-radius: 0px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)

        # Set Arabic headers exactly as in the image
        self.headers = [
            "رقم الموظف",
            "التفعيل",
            "الاسم",
            "اللقب",
            "لقب الزوج ",
            "تاريخ الميلاد",
            "ولاية الميلاد",
            "الجنس",
            " الوضعية العائلية",
            " الوضعية تجاه الخدمة الوطنية",
            "الشهادة التي تم على أساسهاالتوظيف الأصلي",
            "الشهادة الحالية ",
            "رتبة التوظيف الأصلي",
            "الرتبة أو منصب الشغل الحالي ",
            "الصنف الحالي ",
            "الدرجة الحالية",
            "تاريخ المفعول ",
            "التبعية",
            "المصلحة"
        ]

        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)

        # Set text alignment for all columns to right-aligned
        for i in range(len(self.headers)):
            self.table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)

        # Add table to layout
        table_layout.addWidget(self.table)
        self.main_layout.addWidget(table_container)

        # Adjust table properties with variable column widths based on content
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)  # Fixed width columns

        # Define column widths based on content length
        column_widths = {
            "رقم الموظف": 100,
            "التفعيل": 100,  # Added (was not in original widths)
            "الاسم": 150,
            "اللقب": 150,
            "لقب الزوج ": 150,  # Added (was not in original widths)
            "تاريخ الميلاد": 150,
            "ولاية الميلاد": 150,
            "الجنس": 100,
            " الوضعية العائلية": 150,  # Match spacing exactly
            " الوضعية تجاه الخدمة الوطنية": 300,  # Added (not in original)
            "الشهادة التي تم على أساسهاالتوظيف الأصلي": 400,  # Match exact header spacing
            "الشهادة الحالية ": 150,
            "رتبة التوظيف الأصلي": 200,
            "الرتبة أو منصب الشغل الحالي ": 300,
            "الصنف الحالي ": 150,
            "الدرجة الحالية": 150,
            "تاريخ المفعول ": 150,
            "التبعية": 150,
            "المصلحة": 150
        }

        # Set column widths
        for i, header in enumerate(self.headers):
            if i < len(self.headers):
                width = column_widths.get(header, 200)  # Default to 200 if not specified
                self.table.setColumnWidth(i, width)

        self.table.setAlternatingRowColors(False)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)

        # Connect double-click event to open employee details
        self.table.cellDoubleClicked.connect(self.open_employee_details)

        table_layout.addWidget(self.table)
        table_scroll_area.setWidget(table_container)

        # Action buttons at bottom
        action_buttons = QWidget()
        action_buttons_layout = QHBoxLayout(action_buttons)
        action_buttons_layout.setContentsMargins(20, 20, 20, 20)  # Increased top margin
        action_buttons_layout.setSpacing(25) 
        add_button = QPushButton("إضافة")
        add_button.setFixedSize(160, 45)
        add_button.setStyleSheet(f"""
                                 QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
                QPushButton:disabled {{
                    background-color: #a0a0a0;
                    color: #e0e0e0;
                }}
            """)
        # Connect the Add button to open the Add Employee window
        add_button.clicked.connect(self.open_add_employee_window)

        additional_info_btn = QPushButton("معلومات إضافية")
        additional_info_btn.setFixedSize(160, 45)
        additional_info_btn.setStyleSheet(f"""
                                 QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
                QPushButton:disabled {{
                    background-color: #a0a0a0;
                    color: #e0e0e0;
                }}
            """)
        additional_info_btn.clicked.connect(self.open_employee_details_from_button)

        cancel_btn = QPushButton("إلغاء التفعيل")
        cancel_btn.setFixedSize(160, 45)
        cancel_btn.setStyleSheet(f"""
                                 QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
                QPushButton:disabled {{
                    background-color: #a0a0a0;
                    color: #e0e0e0;
                }}
            """)
        cancel_btn.clicked.connect(self.open_departT_from_button)    

        archive_btn = QPushButton("أرشفة")
        archive_btn.setFixedSize(160, 45)
        archive_btn.setStyleSheet(f"""
                                 QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
                QPushButton:disabled {{
                    background-color: #a0a0a0;
                    color: #e0e0e0;
                }}
            """)
        archive_btn.clicked.connect(self.open_archive_form_button)

        activity_log_btn = QPushButton("سجل الأنشطة")
        activity_log_btn.setFixedSize(160, 45)
        activity_log_btn.setStyleSheet(f"""
                                 QPushButton {{
                    background-color: {ORANGE};
                    color: {WHITE};
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #e05d00;
                }}
                QPushButton:pressed {{
                    background-color: #cc5200;
                }}
                QPushButton:disabled {{
                    background-color: #a0a0a0;
                    color: #e0e0e0;
                }}
            """)
        activity_log_btn.clicked.connect(self.show_history)

        action_buttons_layout.addWidget(add_button)
        
        action_buttons_layout.addWidget(additional_info_btn)
        
        action_buttons_layout.addWidget(cancel_btn)
        
        action_buttons_layout.addWidget(archive_btn)
       
        action_buttons_layout.addWidget(activity_log_btn)

        # Content area
        self.content_area = QWidget()
        content_layout = QVBoxLayout(self.content_area)
        content_layout.setContentsMargins(0, 0, 15, 0)  # Add 15px right margin
        content_layout.setSpacing(5)
        
        # Add widgets in the order: toolbar, filters, title, table, buttons
        content_layout.addWidget(toolbar)
        content_layout.addWidget(filter_section)
        content_layout.addWidget(table_title)
        content_layout.addWidget(table_scroll_area)
        content_layout.addWidget(action_buttons)

        # Add sidebars and content to main layout
        self.main_layout.addWidget(self.content_area)
        self.main_layout.addWidget(self.stats_sidebar_widget)
        self.main_layout.addSpacing(15)  # Add 15px spacing between content and sidebar

        # Initialize with some sample data
        #self.initialize_sample_data()

        #get the data fom the DB
        self.load_employees_to_table()

        # Update statistics
        #self.update_statistics()

    def refresh_data(self):
        """Refresh the table data from database"""
        self.load_employees_to_table()
        # Reset any applied filters
        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)
        self.selected_filter_columns = []
    def update_data(self):
        self.refresh_data()
    # for getting the data into the main table 
    def load_employees_to_table(self):
        self.session.expire_all()
        self.table.setRowCount(0)  # Vider les lignes existantes

        # Récupérer les employés avec informations de carrière
        employees_data = self.controller.get_employees_with_career()

        # Récupérer la liste des ID des employés ayant quitté définitivement
        departed_ids = set(depart.idemploye for depart in self.controller.get_final_departures())

        # Vérifier si des colonnes filtrées sont définies
        has_filtered_columns = hasattr(self, 'visible_headers') and hasattr(self, 'selected_columns')

        for emp, carriere in employees_data:
            # ⚠️ Ignorer les employés archivés (départ définitif)
            if emp.idemploye in departed_ids:
                continue

            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            emp_type = "موظف" if emp.type == "permanent" else "عون متعاقد"

            all_data = [
                str(emp.idemploye),                         # رقم الموظف
                "مفعل" if emp.Statut else "غير مفعل",       # التفعيل
                emp.Prenom or "",                              # الاسم
                emp.Nom or "",                           # اللقب
                emp.NomEpoux or "",                         # لقب الزوج
                emp.Datedenaissance.strftime("%Y-%m-%d") if emp.Datedenaissance else "",  # تاريخ الميلاد
                emp.Lieudenaissance or "",                  # ولاية الميلاد
                emp.Sexe or "",                             # الجنس
                emp.Statutfamilial or "",                   # الوضعية العائلية
                emp.Servicesnationale or "",                # الوضعية تجاه الخدمة الوطنية
                carriere.Dipinitial or "",                  # الشهادة الأصلية
                carriere.Dipactuel or "",                   # الشهادة الحالية
                carriere.GRec or "",                        # رتبة التوظيف الأصلي
                carriere.Nomposte or "",                    # الرتبة أو المنصب الحالي
                carriere.current_class or "",               # الصنف الحالي
                str(carriere.current_reference_number or ""), # الرقم الاستدلالي
                carriere.effectiveDate.strftime("%Y-%m-%d") if carriere.effectiveDate else "",  # تاريخ المفعول
                carriere.dependency or "",                  # التبعية
                carriere.service or "",                     # المصلحة
            ]

            if has_filtered_columns:
                for col, header in enumerate(self.visible_headers):
                    original_index = self.headers.index(header)
                    if original_index < len(all_data):
                        self.table.setItem(row_position, col, QTableWidgetItem(all_data[original_index]))
            else:
                for col, value in enumerate(all_data):
                    if col < self.table.columnCount():
                        self.table.setItem(row_position, col, QTableWidgetItem(value))

        # Mettre à jour les statistiques dans la sidebar
        self.stats_sidebar_widget.refresh_statistics()


    def show_history(self):
        """Show the activity history using the  history dialog with shared session"""
        try:
            from Views.Historique import HistoryDialog
            
            print(f"DEBUG - Ouverture historique avec user_data: {self.current_user_data}")
            
            self.history_dialog = HistoryDialog(
                parent=self,
                current_user_data=self.current_user_data,
                session=self.session,  # AJOUT: Passer la session partagée
                module_name="إدارة الموظفين" , # Module name in Arabic
                gestion_filter="إدارة الموظفين"  # Module name in Arabic
            )
            self.history_dialog.show()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "خطأ",
                f"حدث خطأ غير متوقع:\n{str(e)}"
            )
            print(f"Erreur dans show_history: {e}")   
    # filter columns for the table
    def open_filter_columns_window(self):
        """Open the filter columns window"""
        self.filter_columns_window = FilterTableColumnsWindow(self, self.headers)
        self.filter_columns_window.exec_()

    # for reordering the table
    def apply_column_filter(self, selected_columns, column_order):
        """Apply column filter to the table - with proper reordering"""
        # Save the current data
        saved_data = []
        for row in range(self.table.rowCount()):
            row_data = {}
            for col, header in enumerate(self.headers):
                item = self.table.item(row, col)
                if item:
                    row_data[header] = item.text()
                else:
                    row_data[header] = ""
            saved_data.append(row_data)

        # Remember the current row count
        row_count = self.table.rowCount()

        # Clear the table
        self.table.clear()

        # Set up the table with the new column order
        self.table.setColumnCount(len(selected_columns))

        # Create new headers list based on the selected columns and their order
        new_headers = []
        for i, original_index in enumerate(selected_columns):
            new_headers.append(self.headers[original_index])

        self.table.setHorizontalHeaderLabels(new_headers)

        # Set text alignment for all columns to right-aligned (for Arabic)
        for i in range(len(new_headers)):
            header_item = self.table.horizontalHeaderItem(i)
            if header_item:
                header_item.setTextAlignment(Qt.AlignVCenter)

        # Define column widths based on content length
        column_widths = {
            "رقم الموظف": 100,
            "التفعيل": 100,
            "الاسم": 150,
            "اللقب": 150,
            "لقب الزوج ": 150,
            "تاريخ الميلاد": 150,
            "ولاية الميلاد": 150,
            "الجنس": 100,
            " الوضعية العائلية": 150,
            " الوضعية تجاه الخدمة الوطنية": 300,
            "الشهادة التي تم على أساسهاالتوظيف الأصلي": 400,
            "الشهادة الحالية ": 150,
            "رتبة التوظيف الأصلي": 200,
            "الرتبة أو منصب الشغل الحالي ": 300,
            "الصنف الحالي ": 150,
            "الدرجة الحالية": 150,
            "تاريخ المفعول ": 150,
            "التبعية": 150,
            "المصلحة": 150
        }

        # Set column widths
        for i, header in enumerate(new_headers):
            width = column_widths.get(header, 200)  # Default to 200 if not specified
            self.table.setColumnWidth(i, width)

        # Restore the data in the new order
        self.table.setRowCount(row_count)
        for row, row_data in enumerate(saved_data):
            for col, header in enumerate(new_headers):
                if header in row_data:
                    self.table.setItem(row, col, QTableWidgetItem(row_data[header]))

        # Make the last column stretch
        self.table.horizontalHeader().setStretchLastSection(True)

        # Store the new column configuration
        self.visible_headers = new_headers
        self.selected_columns = selected_columns
        
        # Update the table after applying filters
        self.load_employees_to_table()

    def update_statistics(self):
        """Update the statistics based on the current table data"""
        # Count total employees in the table
        #total_employees = self.table.rowCount()
        #self.stats_widgets_dict["total_employees"].setText(str(total_employees))

        # Count employees by type
        employees_count = 0
        contractors_count = 0
        active_count = 0
        inactive_count = 0

        for row in range(self.table.rowCount()):
            # Check employment type
            job_type_col = -1  # Not directly in the table, need to get from employee data
            activation_col = 1  # "التفعيل" column index

            # Get the employee data
            employee_data = self.get_employee_data_from_row(row)

            # Count by employment type
            if "طبيعة علاقة العمل (موظف عون متعاقد)" in employee_data:
                if employee_data["طبيعة علاقة العمل (موظف عون متعاقد)"] == "موظف":
                    employees_count += 1
                elif employee_data["طبيعة علاقة العمل (موظف عون متعاقد)"] == "عون متعاقد":
                    contractors_count += 1

            # Count by activation status
            activation_item = self.table.item(row, activation_col)
            if activation_item:
                if activation_item.text() == "مفعل":
                    active_count += 1
                elif activation_item.text() == "غير مفعل":
                    inactive_count += 1

        # Update statistics widgets
        #self.stats_widgets_dict["employees"].setText(str(employees_count))
        #self.stats_widgets_dict["contractors"].setText(str(contractors_count))
        #self.stats_widgets_dict["active_employees"].setText(str(active_count))
        #self.stats_widgets_dict["inactive_employees"].setText(str(inactive_count))

    def toggle_stats_sidebar(self):
        if self.stats_sidebar_widget.isVisible():
            self.stats_sidebar_widget.hide()
        else:
            self.stats_sidebar_widget.refresh_statistics()  # Appel correct
            self.stats_sidebar_widget.show()

    def toggle_sidebar(self):
        """Toggle menu sidebar visibility"""
        if self.sidebar_visible:
            # Hide sidebar
            self.sidebar.setFixedWidth(0)
            self.sidebar_visible = False
        else:
            # Show sidebar
            self.sidebar.setFixedWidth(self.sidebar_width)
            self.sidebar_visible = True
            
    def open_add_employee_window(self):
        """Open the Add Employee window"""
        # Pass current user data to the AddEmployeeWindow
        self.add_window = AddEmployeeWindow(self, self.current_user_data, session=self.session)
        self.add_window.exec_()
        self.refresh_data()  # Refresh the table after adding a new employee

    def open_employee_details(self, row, column):
        """Open employee details window when double-clicking a table cell"""
        if row < 0 or row >= self.table.rowCount():
            return
        
        # Get employee ID from the correct column regardless of order
        employee_id = self.get_employee_id_from_row(row)
        
        if not employee_id:
            QMessageBox.warning(self, "خطأ", "لا يمكن العثور على رقم الموظف")
            return
        
        # Create employee data dictionary with current visible columns
        employee_data = self.get_employee_data_from_row(row)
        employee_data["رقم الموظف"] = employee_id
        
        # Open the employee details window with current user data
        details_window = EmployeeDetailsWindow(self, employee_data, self.current_user_data,session=self.session)
        details_window.exec_()
        
        # Refresh the table after closing the details window
        self.load_employees_to_table()

    def open_employee_details_from_button(self):
        """Open employee details window from the button"""
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد موظف من الجدول")
            return
        
        # Get employee ID from the correct column regardless of order
        employee_id = self.get_employee_id_from_row(current_row)
        
        if not employee_id:
            QMessageBox.warning(self, "خطأ", "لا يمكن العثور على رقم الموظف")
            return
        
        # Create employee data dictionary with current visible columns
        employee_data = self.get_employee_data_from_row(current_row)
        employee_data["رقم الموظف"] = employee_id
        
        # Open the employee details window with current user data
        details_window = EmployeeDetailsWindow(self, employee_data, self.current_user_data,session=self.session)
        details_window.exec_()
        
        # Refresh the table after closing the details window
        self.load_employees_to_table()

    def open_departT_from_button(self):
        """Ouvre le formulaire de départ temporaire en modal, après vérification qu'une ligne est sélectionnée"""
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد موظف من الجدول")
            return
        # Récupère l'ID de l'employé à partir de la table (colonne 0 par exemple)
        employee_id = int(self.table.item(current_row, 0).text())
        employe = self.session.query(Employe).filter_by(idemploye=employee_id).first()
        if not employe:
            QMessageBox.warning(self, "خطأ", "الموظف غير موجود في قاعدة البيانات")
            return
        if not employe.Statut:
            QMessageBox.information(self, "تنبيه", "هذا الموظف غير نشط بالفعل ولا يمكن تسجيل رحيله المؤقت.")
            return
        # Crée une instance du formulaire
        self.departT_form = TemporaryDepartureForm(
            session=self.session,
            employe=employe,
            refresh_callback=self.load_employees_to_table  
        )
        # Affiche la fenêtre du formulaire
        self.departT_form.show()

    def open_archive_form_button(self):
        """Ouvre le formulaire de départ temporaire en modal, après vérification qu'une ligne est sélectionnée"""
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد موظف من الجدول")
            return
        employee_id = int(self.table.item(current_row, 0).text())
        employe = self.session.query(Employe).filter_by(idemploye=employee_id).first()
        if not employe:
            QMessageBox.warning(self, "خطأ", "الموظف غير موجود في قاعدة البيانات")
            return
        # Crée une instance du formulaire
        self.archive_form = FinalDepartureForm(
            session=self.session,
            employe=employe,
            refresh_callback=self.load_employees_to_table  
        )
        # Affiche la fenêtre du formulaire
        self.archive_form.show()

    def get_employee_id_from_row(self, row):
        """Get employee ID from a table row regardless of column order"""
        # Check if we have filtered columns
        has_filtered_columns = hasattr(self, 'visible_headers') and hasattr(self, 'selected_columns')
        
        if has_filtered_columns:
            # Find the column index for "رقم الموظف" in the visible headers
            try:
                id_column_index = self.visible_headers.index("رقم الموظف")
                item = self.table.item(row, id_column_index)
                if item:
                    return item.text()
            except ValueError:
                # "رقم الموظف" is not in visible columns, need to get it from database
                # Get any available data to identify the employee
                employee_data = self.get_employee_data_from_row(row)
                return self.find_employee_id_by_data(employee_data)
        else:
            # Use original column order
            id_column_index = 0  # "رقم الموظف" is the first column in original order
            item = self.table.item(row, id_column_index)
            if item:
                return item.text()
        
        return None

    def get_employee_data_from_row(self, row):
        """Get all employee data from a table row"""
        employee_data = {}
        
        # Check if we have filtered columns
        has_filtered_columns = hasattr(self, 'visible_headers') and hasattr(self, 'selected_columns')
        
        if has_filtered_columns:
            # Use visible headers
            headers_to_use = self.visible_headers
        else:
            # Use all headers
            headers_to_use = self.headers
        
        # Get data from each visible column
        for col, header in enumerate(headers_to_use):
            if col < self.table.columnCount():
                item = self.table.item(row, col)
                if item:
                    employee_data[header] = item.text()
                else:
                    employee_data[header] = ""
        
        return employee_data
    
    def find_employee_id_by_data(self, employee_data):
        """Find employee ID using other available data when ID column is not visible"""
        try:
            # Try to find employee by name and surname combination
            name = employee_data.get("الاسم", "")
            surname = employee_data.get("اللقب", "")
            
            if name and surname:
                # Query database to find employee by name and surname
                employees_data = self.controller.get_employees_with_career()
                for emp, carriere in employees_data:
                    if emp.Prenom.strip().lower() == name.strip().lower() and emp.Nom.strip().lower() == surname.strip().lower():
                        return str(emp.idemploye)
                print(f"Comparing: DB({emp.Prenom}, {emp.Nom}) vs UI({name}, {surname})")
            # Try to find by other unique fields if available
            position = employee_data.get("الرتبة أو منصب الشغل الحالي ", "")
            service = employee_data.get("المصلحة", "")
            
            if position and service:
                employees_data = self.controller.get_employees_with_career()
                for emp, carriere in employees_data:
                    if carriere.Nomposte == position and carriere.service == service:
                        # Additional check to ensure uniqueness
                        if name and surname:
                            if emp.Prenom == name and emp.Nom == surname:
                                return str(emp.idemploye)
                        else:
                            return str(emp.idemploye)
            
            return None
        except Exception as e:
            print(f"Error finding employee ID: {e}")
            return None

    def add_employee_to_table(self, employee_data):
        """Add a new employee to the table"""
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)

        # Map the employee data to the correct columns based on the headers
        # This ensures data is displayed in the correct order
        column_mapping = {
            0: 0,  # رقم الموظف
            1: 2,  # الاسم
            2: 3,  # اللقب
            3: 5,  # تاريخ الميلاد
            4: 6,  # ولاية الميلاد
            5: 7,  # الجنس
            10: 8,  # الوضعية العائلية
            # Add more mappings as needed
        }

        # Fill in the data we have
        for src_col, dest_col in column_mapping.items():
            if src_col < len(employee_data) and dest_col < self.table.columnCount():
                self.table.setItem(row_position, dest_col, QTableWidgetItem(str(employee_data[src_col])))

        # Set default value for activation status (مفعل)
        self.table.setItem(row_position, 1, QTableWidgetItem("مفعل"))

        # Update statistics
        self.update_statistics()
        if hasattr(self, 'stats_sidebar') and self.stats_sidebar_visible:
            self.stats_sidebar.refresh_statistics()

    def open_filter_window(self):
        """Open the filter window"""
        self.filter_window = FilterWindow(self, self.headers)
        self.filter_window.exec_()

    def apply_filter_criteria(self, criteria):
        """Apply filter criteria to the table using database query"""
        # Clear the table
        self.table.setRowCount(0)

        # If no criteria, show all employees
        if not criteria:
            self.load_employees_to_table()
            return

        # Use the controller to filter employees from the database
        filtered_employees = self.controller.filter_employees(criteria)

                    
        # Check if we have filtered columns
        has_filtered_columns = hasattr(self, 'visible_headers') and hasattr(self, 'selected_columns')

        # Add filtered employees to the table
        for emp, carriere in filtered_employees:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            all_data = [
                str(emp.idemploye),
                "مفعل" if emp.Statut else "غير مفعل",
                emp.Nom or "",
                emp.Prenom or "",
                emp.NomEpoux or "",
                emp.Datedenaissance.strftime("%Y-%m-%d") if emp.Datedenaissance else "",
                emp.Lieudenaissance or "",
                emp.Sexe or "",
                emp.Statutfamilial or "",
                emp.Servicesnationale or "",
                carriere.Dipinitial or "",
                carriere.Dipactuel or "",
                carriere.GRec or "",
                carriere.Nomposte or "",
                carriere.current_class or "",
                str(carriere.current_reference_number or ""),
                carriere.effectiveDate.strftime("%Y-%m-%d") if carriere.effectiveDate else "",
                carriere.dependency or "",
                carriere.service or "",
            ]

            if has_filtered_columns:
                for col, header in enumerate(self.visible_headers):
                    original_index = self.headers.index(header)
                    if original_index < len(all_data):
                        self.table.setItem(row_position, col, QTableWidgetItem(all_data[original_index]))
            else:
                for col, value in enumerate(all_data):
                    if col < self.table.columnCount():
                        self.table.setItem(row_position, col, QTableWidgetItem(value))


        
        self.update_statistics()

    def apply_filter(self, filter_criteria):
        """Apply filter criteria to the table"""
        try:
            # Clear the table
            self.table.setRowCount(0)
            
            # Get filtered employees from controller
            filtered_employees = self.controller.filter_employees(filter_criteria)
            
            # Check if we have filtered columns
            has_filtered_columns = hasattr(self, 'visible_headers') and hasattr(self, 'selected_columns')
            
            # Populate table with filtered data
            for emp, carriere in filtered_employees:
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                
                # Prepare all possible data
                all_data = [
                    str(emp.idemploye),                         # رقم الموظف
                    "مفعل" if emp.Statut else "غير مفعل",       # التفعيل
                    emp.Nom or "",                              # الاسم
                    emp.Prenom or "",                           # اللقب
                    emp.NomEpoux or "",                         # لقب الزوج
                    emp.Datedenaissance.strftime("%Y-%m-%d") if emp.Datedenaissance else "",  # تاريخ الميلاد
                    emp.Lieudenaissance or "",                  # ولاية الميلاد
                    emp.Sexe or "",                             # الجنس
                    emp.Statutfamilial or "",                   # الوضعية العائلية
                    emp.Servicesnationale or "",                # الوضعية تجاه الخدمة الوطنية
                    carriere.Dipinitial or "",                  # الشهادة الأصلية
                    carriere.Dipactuel or "",                   # الشهادة الحالية
                    carriere.GRec or "",                        # رتبة التوظيف الأصلي
                    carriere.Nomposte or "",                    # الرتبة أو المنصب الحالي
                    carriere.current_class or "",               # الصنف الحالي
                    str(carriere.current_reference_number or ""), # الرقم الاستدلالي
                    carriere.effectiveDate.strftime("%Y-%m-%d") if carriere.effectiveDate else "",  # تاريخ المفعول
                    carriere.dependency or "",                  # التبعية
                    carriere.service or "",                     # المصلحة
                ]
                
                # If we have filtered columns, use them
                if has_filtered_columns:
                    for col, header in enumerate(self.visible_headers):
                        original_index = self.headers.index(header)
                        if original_index < len(all_data):
                            self.table.setItem(row_position, col, QTableWidgetItem(all_data[original_index]))
                else:
                    # Otherwise use all columns
                    for col, value in enumerate(all_data):
                        if col < self.table.columnCount():
                            self.table.setItem(row_position, col, QTableWidgetItem(value))
                            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء تطبيق الترشيح: {str(e)}")

    def search_employees(self):
        """Search employees based on search box input"""
        search_term = self.search_box.text().strip()
        
        if not search_term:
            # If search is empty, reload all employees
            self.load_employees_to_table()
            return
        
        # Clear the table
        self.table.setRowCount(0)
        
        # Get all employees with career information
        employees_data = self.controller.get_employees_with_career()
        
        # Check if we have filtered columns
        has_filtered_columns = hasattr(self, 'visible_headers') and hasattr(self, 'selected_columns')
        
        # Filter employees based on search term
        for emp, carriere in employees_data:
            # Check if search term matches any field
            if (search_term in str(emp.idemploye) or
                search_term in (emp.Nom or "") or
                search_term in (emp.Prenom or "") or
                search_term in (carriere.Nomposte or "") or
                search_term in (carriere.service or "") or
                search_term in (carriere.dependency or "")):
                
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                
                # Prepare all possible data
                all_data = [
                    str(emp.idemploye),                         # رقم الموظف
                    "مفعل" if emp.Statut else "غير مفعل",       # التفعيل
                    emp.Nom or "",                              # الاسم
                    emp.Prenom or "",                           # اللقب
                    emp.NomEpoux or "",                         # لقب الزوج
                    emp.Datedenaissance.strftime("%Y-%m-%d") if emp.Datedenaissance else "",  # تاريخ الميلاد
                    emp.Lieudenaissance or "",                  # ولاية الميلاد
                    emp.Sexe or "",                             # الجنس
                    emp.Statutfamilial or "",                   # الوضعية العائلية
                    emp.Servicesnationale or "",                # الوضعية تجاه الخدمة الوطنية
                    carriere.Dipinitial or "",                  # الشهادة الأصلية
                    carriere.Dipactuel or "",                   # الشهادة الحالية
                    carriere.GRec or "",                        # رتبة التوظيف الأصلي
                    carriere.Nomposte or "",                    # الرتبة أو المنصب الحالي
                    carriere.current_class or "",               # الصنف الحالي
                    str(carriere.current_reference_number or ""), # الرقم الاستدلالي
                    carriere.effectiveDate.strftime("%Y-%m-%d") if carriere.effectiveDate else "",  # تاريخ المفعول
                    carriere.dependency or "",                  # التبعية
                    carriere.service or "",                     # المصلحة
                ]
                
                # If we have filtered columns, use them
                if has_filtered_columns:
                    for col, header in enumerate(self.visible_headers):
                        original_index = self.headers.index(header)
                        if original_index < len(all_data):
                            self.table.setItem(row_position, col, QTableWidgetItem(all_data[original_index]))
                else:
                    # Otherwise use all columns
                    for col, value in enumerate(all_data):
                        if col < self.table.columnCount():
                            self.table.setItem(row_position, col, QTableWidgetItem(value))

    def export_to_excel(self):
        """Exporte les données actuelles du QTableWidget vers un fichier Excel,
           en utilisant une boîte de dialogue de sauvegarde simplifiée."""
        if self.table.rowCount() == 0:
            QMessageBox.information(self, "لا توجد بيانات", "لا توجد بيانات في الجدول للتصدير.")
            return

        # Proposer un nom de fichier par défaut avec horodatage
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_file_name = f"rapport_employes_{timestamp}.xlsx"

        # Essayer de proposer le dossier Téléchargements comme emplacement par défaut
        try:
            if os.name == 'nt': # Windows
                default_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            elif os.name == 'posix': # macOS, Linux
                default_dir = os.path.join(os.path.expanduser("~"), "Downloads")
                if not os.path.exists(default_dir):
                    localized_downloads_path = os.path.join(os.path.expanduser("~"), "Téléchargements")
                    if os.path.exists(localized_downloads_path):
                        default_dir = localized_downloads_path
            else:
                default_dir = os.path.expanduser("~") # Fallback vers le dossier utilisateur

            if not os.path.exists(default_dir): # Si "Downloads" n'existe vraiment pas, fallback au dossier utilisateur
                default_dir = os.path.expanduser("~")

            default_path = os.path.join(default_dir, default_file_name)

        except Exception:
            default_path = os.path.join(os.path.expanduser("~"), default_file_name) # Fallback simple

        # Utiliser QFileDialog pour obtenir le chemin de sauvegarde
        # La boîte de dialogue native du système sera utilisée.
        filePath, _ = QFileDialog.getSaveFileName(
            self,
            "حفظ التقرير كـ Excel",  # Titre de la boîte de dialogue
            default_path,          # Chemin et nom de fichier suggérés par défaut
            "Fichiers Excel (*.xlsx)" # Filtre de type de fichier simplifié
        )

        if not filePath: # Si l'utilisateur annule la boîte de dialogue
            return

        # S'assurer que l'extension est .xlsx (QFileDialog devrait s'en charger avec le filtre, mais par sécurité)
        if not filePath.lower().endswith(".xlsx"):
            filePath += ".xlsx"

        # Le reste de la logique d'exportation est identique
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.sheet_view.rightToLeft = True
            headers_to_export = []
            if hasattr(self, 'visible_headers') and self.visible_headers:
                headers_to_export = self.visible_headers
            else:
                for col in range(self.table.columnCount()):
                    header_item = self.table.horizontalHeaderItem(col)
                    if header_item:
                        headers_to_export.append(header_item.text())
                    else:
                        headers_to_export.append(f"Colonne {col+1}")

            for col_num, header_title in enumerate(headers_to_export, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header_title
                cell.font = openpyxl.styles.Font(bold=True)

            for row_num in range(self.table.rowCount()):
                for col_num_table in range(len(headers_to_export)):
                    item = self.table.item(row_num, col_num_table)
                    value = item.text() if item else ""
                    sheet.cell(row=row_num + 2, column=col_num_table + 1).value = value

            for col_idx, column_cells in enumerate(sheet.columns, 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(filePath)

        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ أثناء تصدير الملف:\n{e}")
            print(f"Erreur d'exportation Excel : {e}")
            import traceback
            traceback.print_exc()

    def generate_current_table_as_pdf_action(self):
        """Action appelée par le bouton pour générer et sauvegarder le PDF du tableau."""
        if self.table.rowCount() == 0:
            QMessageBox.information(self, "لا توجد بيانات", "لا توجد بيانات في الجدول لتوليد PDF.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_file_name = f"rapport_tableau_employes_{timestamp}.pdf"
        try:
            if os.name == 'nt': default_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            elif os.name == 'posix':
                default_dir = os.path.join(os.path.expanduser("~"), "Downloads")
                if not os.path.exists(default_dir):
                    localized_path = os.path.join(os.path.expanduser("~"), "Téléchargements")
                    if os.path.exists(localized_path): default_dir = localized_path
            else: default_dir = os.path.expanduser("~")
            if not os.path.exists(default_dir): default_dir = os.path.expanduser("~")
            default_path = os.path.join(default_dir, default_file_name)
        except Exception:
            default_path = os.path.join(os.path.expanduser("~"), default_file_name)

        filePath, _ = QFileDialog.getSaveFileName(
            self, "حفظ التقرير كـ PDF", default_path, "Fichiers PDF (*.pdf)"
        )

        if not filePath:
            return
        if not filePath.lower().endswith(".pdf"):
            filePath += ".pdf"

        success = self._generate_pdf_from_table_data(filePath) # Appel à la méthode de génération
        if success:
            QMessageBox.information(self, "نجاح", f"تم توليد ملف PDF بنجاح:\n{filePath}")
        else:
            QMessageBox.critical(self, "فشل", "حدث خطأ أثناء توليد ملف PDF.")

    def _generate_pdf_from_table_data(self, output_pdf_path, report_title="تقرير الموظفين"):
        """
        Génère un fichier PDF à partir des données actuelles du QTableWidget.

        Args:
            output_pdf_path (str): Chemin complet où sauvegarder le fichier PDF.
            report_title (str): Titre à afficher sur le rapport PDF.

        Returns:
            bool: True si la génération a réussi, False sinon.
        """
        if self.table.rowCount() == 0:
            print("Avertissement: Tentative de génération de PDF avec une table vide.")
            return False

        # 1. Extraire les données et les en-têtes de la table
        headers_for_pdf = []
        # Utiliser les en-têtes visibles s'ils sont définis (après filtrage de colonnes)
        if hasattr(self, 'visible_headers') and self.visible_headers:
            headers_for_pdf = self.visible_headers
        else: # Sinon, prendre tous les en-têtes de la table actuelle
            for col in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(col)
                headers_for_pdf.append(header_item.text() if header_item else f"العمود {col+1}")

        employee_data_rows = []
        for row_num in range(self.table.rowCount()):
            row_data = []
            # Itérer sur le nombre de colonnes défini par les en-têtes que nous allons utiliser
            for col_num_in_headers in range(len(headers_for_pdf)):
                item = self.table.item(row_num, col_num_in_headers) # Lire la cellule à la position actuelle de la table
                row_data.append(item.text() if item else "")
            employee_data_rows.append(row_data)

        # 2. Préparer le contexte pour Jinja2
        context = {
            "report_title": report_title,
            "headers": headers_for_pdf,
            "employee_data_rows": employee_data_rows, # Renommé pour plus de clarté dans le template
            "generation_date": datetime.now().strftime("%Y-%m-%d %H:%M")
        }

        # 3. Charger le template et rendre l'HTML
        try:
            # Déterminer le chemin du dossier des templates
            # S'attend à ce que le dossier 'templates' soit au même niveau que le script exécuté
            # ou que sys.argv[0] pointe vers le script principal si packagé.
            if getattr(sys, 'frozen', False): # Si l'application est "gelée" (ex: PyInstaller)
                script_dir = os.path.dirname(sys.executable)
            else: # En mode script normal
                script_dir = os.path.dirname(os.path.abspath(sys.argv[0] or __file__))
            
            template_folder = os.path.join(script_dir, "templates")

            if not os.path.isdir(template_folder):
                QMessageBox.critical(self, "خطأ", f"لم يتم العثور على مجلد القوالب: {template_folder}")
                print(f"ERREUR: Dossier de template non trouvé: {template_folder}")
                return False
            
            template_name = "employee_report_template.html"
            if not os.path.isfile(os.path.join(template_folder, template_name)):
                QMessageBox.critical(self, "خطأ", f"لم يتم العثور على ملف القالب '{template_name}' في '{template_folder}'.")
                print(f"ERREUR: Fichier template '{template_name}' non trouvé dans '{template_folder}'.")
                return False

            env = Environment(loader=FileSystemLoader(template_folder), autoescape=True) # autoescape est une bonne pratique
            template = env.get_template(template_name)
            html_output = template.render(context)

        except Exception as e_template:
            QMessageBox.critical(self, "خطأ في القالب", f"حدث خطأ أثناء معالجة القالب:\n{e_template}")
            print(f"Erreur de template Jinja2: {e_template}")
            traceback.print_exc()
            return False

        # 4. Convertir l'HTML en PDF avec WeasyPrint
        try:
            # S'assurer que le dossier de destination pour le PDF existe
            output_dir = os.path.dirname(output_pdf_path)
            if output_dir: # S'il y a un chemin de dossier (pas juste un nom de fichier)
                os.makedirs(output_dir, exist_ok=True)

            font_config = FontConfiguration() # Pour une meilleure gestion des polices si nécessaire
            
            html_doc = WeasyHTML(string=html_output)
            html_doc.write_pdf(output_pdf_path, font_config=font_config)

            print(f"PDF généré avec succès : {output_pdf_path}")
            return True

        except ImportError: # Au cas où WeasyPrint ne serait pas trouvé au runtime
            QMessageBox.critical(self, "خطأ", "مكتبة WeasyPrint غير مثبتة بشكل صحيح لتوليد PDF.")
            print("ERREUR: WeasyPrint n'est pas installé pour la génération PDF.")
            return False
        except Exception as e_pdf:
            QMessageBox.critical(self, "خطأ في توليد PDF", f"حدث خطأ أثناء توليد ملف PDF:\n{e_pdf}")
            print(f"Erreur WeasyPrint: {e_pdf}")
            traceback.print_exc()
            return False

    # Si vous voulez utiliser cette fonction pour l'impression (Option 3 de la discussion précédente)
    def print_current_table_via_pdf(self):
        # Créer un nom de fichier PDF temporaire
        try:
            temp_dir = tempfile.gettempdir()
            temp_pdf_path = os.path.join(temp_dir, f"temp_employee_report_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")

            success_pdf = self._generate_pdf_from_table_data(temp_pdf_path, "Rapport Employés pour Impression")

            if not success_pdf:
                # Le message d'erreur aura déjà été affiché par _generate_pdf_from_table_data
                return

            # Logique d'impression du PDF (copiée de la discussion précédente)
            if sys.platform == "win32":
                try:
                    os.startfile(temp_pdf_path, "print")
                    # QMessageBox.information(self, "الطباعة", "تم إرسال المستند إلى الطابعة.")
                except Exception as e:
                    QMessageBox.critical(self, "Erreur d'impression", f"Impossible d'imprimer le PDF (Windows):\n{e}\nFichier: {temp_pdf_path}")
            elif sys.platform == "darwin":
                try:
                    subprocess.call(["open", temp_pdf_path])
                except Exception as e:
                    QMessageBox.critical(self, "Erreur d'impression", f"Impossible d'ouvrir le PDF (macOS):\n{e}")
            else: # Linux
                try: subprocess.call(["lp", temp_pdf_path])
                except FileNotFoundError:
                    try: subprocess.call(["xdg-open", temp_pdf_path])
                    except Exception as e_xdg: QMessageBox.critical(self, "Erreur d'impression", f"Impossible d'imprimer/ouvrir (Linux):\n{e_xdg}")
                except Exception as e: QMessageBox.critical(self, "Erreur d'impression", f"Impossible d'imprimer (Linux):\n{e}")
            
            # Ne pas supprimer le fichier temporaire immédiatement pour permettre à la commande d'impression de le lire
            # La suppression des fichiers temporaires est un sujet plus large (ex: à la fermeture de l'app)

        except Exception as e:
            QMessageBox.critical(self, "خطأ عام", f"حدث خطأ:\n{e}")
            traceback.print_exc()
            
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setLayoutDirection(Qt.RightToLeft)  # Set RTL layout for Arabic

    # Set application font
    font = QFont("Arial", 16)
    app.setFont(font)

    window = MainEmployeeWindow(session=None)  # Replace None with your actual session object
    window.show()
    sys.exit(app.exec_())